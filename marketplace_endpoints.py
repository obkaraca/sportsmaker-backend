from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
from pydantic import BaseModel
import uuid
import logging
from motor.motor_asyncio import AsyncIOMotorDatabase, AsyncIOMotorClient
import os
from dotenv import load_dotenv

load_dotenv()

from models import (
    Listing, ListingCreate, ListingType, ListingStatus,
    Offer, OfferCreate, OfferStatus,
    RentalBooking, Transaction, Review, MarketplaceStats,
    ProductCondition, ServiceType, UserType, NotificationType, NotificationRelatedType
)
from auth import get_current_user
from iyzico_service import IyzicoService
from notification_endpoints import create_notification_helper
from geliver_endpoints import create_geliver_shipment_after_payment, get_provider_name, create_geliver_return_shipment

# Initialize Iyzico service
iyzico_service = IyzicoService()


router = APIRouter()
logger = logging.getLogger(__name__)

# Database connection
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "sports_management")
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]


# ============================================
# HELPER FUNCTIONS
# ============================================

def add_business_days(start_date: datetime, days: int) -> datetime:
    """Başlangıç tarihinden itibaren iş günü ekle (hafta sonları hariç)"""
    current_date = start_date
    added_days = 0
    
    while added_days < days:
        current_date += timedelta(days=1)
        # 0 = Pazartesi, 5 = Cumartesi, 6 = Pazar
        if current_date.weekday() < 5:  # Pazartesi-Cuma
            added_days += 1
    
    return current_date


# ============================================
# CATEGORIES
# ============================================

DEFAULT_CATEGORIES = [
    {"id": "futbol", "name": "Futbol", "sport": "Futbol", "icon": "football"},
    {"id": "basketbol", "name": "Basketbol", "sport": "Basketbol", "icon": "basketball"},
    {"id": "voleybol", "name": "Voleybol", "sport": "Voleybol", "icon": "american-football"},
    {"id": "tenis", "name": "Tenis", "sport": "Tenis", "icon": "tennisball"},
    {"id": "yuzme", "name": "Yüzme", "sport": "Yüzme", "icon": "water"},
    {"id": "fitness", "name": "Fitness", "sport": "Fitness", "icon": "fitness"},
    {"id": "diger", "name": "Diğer", "sport": None, "icon": "ellipsis-horizontal"},
]

@router.get("/categories")
async def get_categories():
    """Get all marketplace categories"""
    return {"categories": DEFAULT_CATEGORIES}

# ============================================
# BRANDS - Custom Brand Management
# ============================================

class BrandCreate(BaseModel):
    name: str

@router.get("/brands")
async def get_brands():
    """Get all custom brands added by users"""
    try:
        brands = await db.marketplace_brands.find().sort("name", 1).to_list(500)
        brand_names = [b["name"] for b in brands]
        return {"brands": brand_names}
    except Exception as e:
        logger.error(f"Error fetching brands: {str(e)}")
        return {"brands": []}

@router.post("/brands")
async def add_brand(
    data: BrandCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add a new custom brand to the list"""
    try:
        brand_name = data.name.strip()
        if not brand_name:
            raise HTTPException(status_code=400, detail="Marka adı boş olamaz")
        
        # Check if brand already exists (case-insensitive)
        existing = await db.marketplace_brands.find_one({
            "name": {"$regex": f"^{brand_name}$", "$options": "i"}
        })
        
        if existing:
            return {"message": "Marka zaten mevcut", "brand": existing["name"]}
        
        # Add new brand
        brand = {
            "id": str(uuid.uuid4()),
            "name": brand_name,
            "added_by": current_user["id"],
            "created_at": datetime.utcnow()
        }
        await db.marketplace_brands.insert_one(brand)
        
        logger.info(f"✅ New brand added: {brand_name} by user {current_user['id']}")
        return {"message": "Marka eklendi", "brand": brand_name}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding brand: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# LISTINGS - CREATE, READ, UPDATE, DELETE
# ============================================

@router.post("/listings")
async def create_listing(
    listing_data: ListingCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new marketplace listing"""
    try:
        current_user_id = current_user["id"]
        # Get user info
        user = await db.users.find_one({"id": current_user_id})
        if not user:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
        listing = {
            "id": str(uuid.uuid4()),
            "seller_id": current_user_id,
            "seller_name": user.get("full_name", ""),
            "seller_type": user.get("user_type", "player"),
            "seller_rating": user.get("rating", 0.0),
            **listing_data.dict(),
            "status": ListingStatus.ACTIVE,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
            "expires_at": datetime.utcnow() + timedelta(days=30),
            "offer_count": 0,
            "favorite_count": 0,
            "views_count": 0
        }
        
        # Set status to pending for admin approval
        listing["status"] = "pending"
        
        await db.marketplace_listings.insert_one(listing)
        listing.pop("_id", None)
        
        # Send notification to all admins
        admins = await db.users.find({"user_type": {"$in": ["admin", "super_admin"]}}).to_list(100)
        for admin in admins:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": admin["id"],
                "type": "listing_approval_request",
                "title": "Yeni İlan Onay Talebi",
                "message": f"{user.get('full_name', 'Kullanıcı')} yeni ilan oluşturdu: {listing['title']}",
                "data": {"listing_id": listing["id"], "seller_id": current_user_id, "seller_name": user.get('full_name', '')},
                "action_url": f"/marketplace/approve/{listing['id']}",
                "read": False,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        # Update user stats
        await update_user_marketplace_stats(db, current_user_id)
        
        # Marketplace ilan oluşturma log'u
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "marketplace_listing_create", "success", {
                "listing_id": listing["id"],
                "title": listing["title"],
                "listing_type": listing.get("listing_type"),
                "price": listing.get("price"),
                "category": listing.get("category_id")
            })
        except Exception as log_err:
            logger.error(f"Log error: {log_err}")
        
        logger.info(f"✅ Created listing {listing['id']} (pending approval) by user {current_user_id}")
        return listing
        
    except Exception as e:
        logger.error(f"❌ Error creating listing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/listings")
async def get_listings(
    listing_type: Optional[ListingType] = None,
    category_id: Optional[str] = None,
    condition: Optional[ProductCondition] = None,
    service_type: Optional[ServiceType] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    location: Optional[str] = None,
    seller_id: Optional[str] = None,
    search: Optional[str] = None,
    status: Optional[ListingStatus] = ListingStatus.ACTIVE,
    brand: Optional[str] = None,
    size: Optional[str] = None,
    gender: Optional[str] = None,
    product_type: Optional[str] = None,
    sport: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    skip: int = 0,
    limit: int = 20,
    
):
    """Get marketplace listings with filters"""
    try:
        query = {}
        
        # Default olarak sadece aktif ilanları göster (sold olanları gizle)
        if status:
            query["status"] = status
        else:
            query["status"] = "active"  # Default: sadece aktif ilanlar
        if listing_type:
            query["listing_type"] = listing_type
        if category_id:
            query["category_id"] = category_id
        if condition:
            query["condition"] = condition
        if service_type:
            query["service_type"] = service_type
        if seller_id:
            query["seller_id"] = seller_id
        if location:
            query["location"] = {"$regex": location, "$options": "i"}
        
        # New filters - using $and to avoid $or conflicts
        and_conditions = []
        
        if brand:
            query["brand"] = brand
            
        if size:
            and_conditions.append({
                "$or": [
                    {"size": size},
                    {"tags": {"$regex": f"Beden:.*{size}", "$options": "i"}}
                ]
            })
            
        if gender:
            and_conditions.append({
                "$or": [
                    {"gender": gender.lower()},
                    {"tags": {"$regex": gender, "$options": "i"}}
                ]
            })
            
        if product_type:
            and_conditions.append({
                "$or": [
                    {"model": product_type},
                    {"tags": {"$regex": product_type, "$options": "i"}}
                ]
            })
            
        if sport:
            # Map sport name back to category_id
            sport_category_map = {
                'Futbol': 'futbol',
                'Basketbol': 'basketbol',
                'Voleybol': 'voleybol',
                'Tenis': 'tenis',
                'Yüzme': 'yüzme',
                'Koşu': 'koşu',
                'Fitness': 'fitness',
                'Yoga': 'yoga',
                'Bisiklet': 'bisiklet',
                'Dağ Tırmanışı': 'dağ-tırmanışı',
            }
            sport_category = sport_category_map.get(sport)
            if sport_category:
                # Don't override existing category_id in query
                # Instead add sport filter to $and conditions
                and_conditions.append({"category_id": sport_category})
        
        if min_price is not None or max_price is not None:
            price_query = {}
            if min_price is not None:
                price_query["$gte"] = min_price
            if max_price is not None:
                price_query["$lte"] = max_price
            query["price"] = price_query
        
        if search:
            and_conditions.append({
                "$or": [
                    {"title": {"$regex": search, "$options": "i"}},
                    {"description": {"$regex": search, "$options": "i"}},
                    {"tags": {"$regex": search, "$options": "i"}}
                ]
            })
        
        # Add all $and conditions if any
        if and_conditions:
            query["$and"] = and_conditions
        
        total = await db.marketplace_listings.count_documents(query)
        
        sort_direction = -1 if sort_order == "desc" else 1
        listings = await db.marketplace_listings.find(query).sort(
            sort_by, sort_direction
        ).skip(skip).limit(limit).to_list(limit)
        
        for listing in listings:
            listing.pop("_id", None)
        
        return {
            "listings": listings,
            "total": total,
            "skip": skip,
            "limit": limit
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting listings: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/listings/{listing_id}")
async def get_listing(
    listing_id: str,
    
):
    """Get single listing by ID"""
    try:
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Listing not found")
        
        # Increment views
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$inc": {"views_count": 1}}
        )
        listing["views_count"] += 1
        
        listing.pop("_id", None)
        return listing
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error getting listing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Not: update_listing fonksiyonu satır 2590'da UpdateListingRequest ile tanımlı
# Bu eski endpoint kaldırıldı çünkü çakışma yapıyordu

@router.delete("/listings/{listing_id}")
async def delete_listing(
    listing_id: str,
    current_user: dict = Depends(get_current_user),
    
):
    """Delete listing"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Listing not found")
        
        if listing["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        # Soft delete - change status
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$set": {
                "status": ListingStatus.CANCELLED,
                "updated_at": datetime.utcnow()
            }}
        )
        
        await update_user_marketplace_stats(db, current_user_id)
        
        return {"message": "Listing deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error deleting listing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# OFFERS / BIDS
# ============================================

@router.get("/offers/{offer_id}")
async def get_offer(
    offer_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get offer details"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        offer = await db.marketplace_offers.find_one({"id": offer_id})
        if not offer:
            raise HTTPException(status_code=404, detail="Offer not found")
        
        # Check authorization - only buyer or seller can see offer
        if offer["buyer_id"] != current_user_id and offer["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Not authorized to view this offer")
        
        # Remove MongoDB _id
        offer.pop("_id", None)
        
        return offer
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error getting offer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/offers")
async def create_offer(
    offer_data: OfferCreate,
    current_user: dict = Depends(get_current_user),
    
):
    """Create offer on listing"""
    try:
        # Extract user_id from current_user dict
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        logger.info(f"🔵 Creating offer - User ID: {current_user_id}, Listing: {offer_data.listing_id}")
        
        # Get listing
        listing = await db.marketplace_listings.find_one({"id": offer_data.listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Listing not found")
        
        if listing["status"] != ListingStatus.ACTIVE:
            raise HTTPException(status_code=400, detail="Listing is not active")
        
        if listing["seller_id"] == current_user_id:
            raise HTTPException(status_code=400, detail="Cannot offer on own listing")
        
        if not listing.get("allow_offers", True):
            raise HTTPException(status_code=400, detail="Offers not allowed on this listing")
        
        # Get buyer info
        buyer = await db.users.find_one({"id": current_user_id})
        if not buyer:
            raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
        offer = {
            "id": str(uuid.uuid4()),
            "buyer_id": current_user_id,
            "buyer_name": buyer.get("full_name", ""),
            "seller_id": listing["seller_id"],
            **offer_data.dict(),
            "status": OfferStatus.PENDING,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
            "expires_at": datetime.utcnow() + timedelta(days=3)
        }
        
        await db.marketplace_offers.insert_one(offer)
        offer.pop("_id", None)
        
        # Update listing offer count
        await db.marketplace_listings.update_one(
            {"id": offer_data.listing_id},
            {"$inc": {"offer_count": 1}}
        )
        
        # Send notification to seller
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": listing["seller_id"],
            "type": NotificationType.OFFER_RECEIVED,
            "title": "Yeni Teklif Aldınız",
            "message": f"{buyer.get('full_name', 'Bir kullanıcı')} '{listing['title']}' ilanınıza {offer_data.amount} TL teklif yaptı.",
            "related_id": offer['id'],
            "related_type": NotificationRelatedType.MARKETPLACE_OFFER,
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Created offer {offer['id']} on listing {offer_data.listing_id}")
        return offer
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error creating offer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/offers")
async def get_offers(
    listing_id: Optional[str] = None,
    status: Optional[OfferStatus] = None,
    as_buyer: bool = False,
    as_seller: bool = False,
    skip: int = 0,
    limit: int = 20,
    current_user: dict = Depends(get_current_user),
    
):
    """Get offers (sent or received)"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        query = {}
        
        if listing_id:
            query["listing_id"] = listing_id
        
        if status:
            query["status"] = status
        
        if as_buyer:
            query["buyer_id"] = current_user_id
        elif as_seller:
            query["seller_id"] = current_user_id
        else:
            # Both
            query["$or"] = [
                {"buyer_id": current_user_id},
                {"seller_id": current_user_id}
            ]
        
        total = await db.marketplace_offers.count_documents(query)
        offers = await db.marketplace_offers.find(query).sort(
            "created_at", -1
        ).skip(skip).limit(limit).to_list(limit)
        
        for offer in offers:
            offer.pop("_id", None)
        
        return {
            "offers": offers,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting offers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/offers/{offer_id}/accept")
async def accept_offer(
    offer_id: str,
    current_user: dict = Depends(get_current_user),
    
):
    """Accept offer - locks price for buyer for 12 hours"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        offer = await db.marketplace_offers.find_one({"id": offer_id})
        if not offer:
            raise HTTPException(status_code=404, detail="Offer not found")
        
        if offer["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        if offer["status"] != OfferStatus.PENDING:
            raise HTTPException(status_code=400, detail="Offer is not pending")
        
        # Calculate 12 hour lock time
        price_locked_until = datetime.utcnow() + timedelta(hours=12)
        
        # Update offer with price lock
        await db.marketplace_offers.update_one(
            {"id": offer_id},
            {"$set": {
                "status": OfferStatus.ACCEPTED,
                "updated_at": datetime.utcnow(),
                "price_locked_until": price_locked_until,
                "price_locked_for_buyer": offer["buyer_id"],
                "reminder_sent": False
            }}
        )
        
        # Get listing (keep it active, don't reserve yet)
        listing = await db.marketplace_listings.find_one({"id": offer["listing_id"]})
        
        # Send notification to buyer
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": offer["buyer_id"],
            "type": NotificationType.OFFER_ACCEPTED,
            "title": "Teklifiniz Kabul Edildi! 🎉",
            "message": f"'{listing['title']}' ilanı için yaptığınız {offer['amount']} TL teklif kabul edildi. Fiyat 12 saat boyunca sizin için kilitlendi.",
            "related_id": offer_id,
            "related_type": NotificationRelatedType.MARKETPLACE_OFFER,
            "data": {
                "listing_id": offer["listing_id"],
                "offer_id": offer_id,
                "accepted_price": offer["amount"]
            },
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Accepted offer {offer_id}, price locked until {price_locked_until}")
        return {
            "message": "Offer accepted successfully",
            "price_locked_until": price_locked_until.isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error accepting offer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/offers/{offer_id}/reject")
async def reject_offer(
    offer_id: str,
    current_user: dict = Depends(get_current_user),
    
):
    """Reject offer"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        offer = await db.marketplace_offers.find_one({"id": offer_id})
        if not offer:
            raise HTTPException(status_code=404, detail="Offer not found")
        
        if offer["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        await db.marketplace_offers.update_one(
            {"id": offer_id},
            {"$set": {
                "status": OfferStatus.REJECTED,
                "updated_at": datetime.utcnow()
            }}
        )
        
        # Send notification to buyer
        listing = await db.marketplace_listings.find_one({"id": offer["listing_id"]})
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": offer["buyer_id"],
            "type": NotificationType.OFFER_REJECTED,
            "title": "Teklifiniz Reddedildi",
            "message": f"'{listing['title']}' ilanı için yaptığınız {offer['amount']} TL teklif reddedildi.",
            "related_id": offer_id,
            "related_type": NotificationRelatedType.MARKETPLACE_OFFER,
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Rejected offer {offer_id}")
        return {"message": "Offer rejected"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error rejecting offer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/offers/{offer_id}/counter")
async def counter_offer(
    offer_id: str,
    counter_amount: float,
    counter_message: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    
):
    """Make counter offer"""
    try:
        current_user_id = current_user["id"] if isinstance(current_user, dict) else current_user
        offer = await db.marketplace_offers.find_one({"id": offer_id})
        if not offer:
            raise HTTPException(status_code=404, detail="Offer not found")
        
        if offer["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        await db.marketplace_offers.update_one(
            {"id": offer_id},
            {"$set": {
                "status": OfferStatus.COUNTER_OFFERED,
                "counter_amount": counter_amount,
                "counter_message": counter_message,
                "updated_at": datetime.utcnow()
            }}
        )
        
        # Send notification to buyer
        listing = await db.marketplace_listings.find_one({"id": offer["listing_id"]})
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": offer["buyer_id"],
            "type": NotificationType.OFFER_COUNTER,
            "title": "Karşı Teklif Yapıldı 💬",
            "message": f"'{listing['title']}' ilanı için yaptığınız {offer['amount']} TL teklife karşı {counter_amount} TL teklif yapıldı.",
            "related_id": offer_id,
            "related_type": NotificationRelatedType.MARKETPLACE_OFFER,
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Counter offer sent for offer {offer_id}: {counter_amount} TL")
        return {"message": "Counter offer sent"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error counter offering: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# FAVORITES
# ============================================

@router.post("/favorites/{listing_id}")
async def add_to_favorites(
    listing_id: str,
    current_user: dict = Depends(get_current_user),
    
):
    """Add listing to favorites"""
    try:
        current_user_id = current_user["id"]
        favorite = {
            "id": str(uuid.uuid4()),
            "user_id": current_user_id,
            "listing_id": listing_id,
            "created_at": datetime.utcnow()
        }
        
        # Check if already favorited
        existing = await db.marketplace_favorites.find_one({
            "user_id": current_user_id,
            "listing_id": listing_id
        })
        
        if existing:
            return {"message": "Zaten favorilerde"}
        
        await db.marketplace_favorites.insert_one(favorite)
        
        # Update listing favorite count
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$inc": {"favorite_count": 1}}
        )
        
        return {"message": "Added to favorites"}
        
    except Exception as e:
        logger.error(f"❌ Error adding favorite: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/favorites/{listing_id}")
async def remove_from_favorites(
    listing_id: str,
    current_user: dict = Depends(get_current_user),
    
):
    """Remove from favorites"""
    try:
        current_user_id = current_user["id"]
        result = await db.marketplace_favorites.delete_one({
            "user_id": current_user_id,
            "listing_id": listing_id
        })
        
        if result.deleted_count > 0:
            await db.marketplace_listings.update_one(
                {"id": listing_id},
                {"$inc": {"favorite_count": -1}}
            )
        
        return {"message": "Removed from favorites"}
        
    except Exception as e:
        logger.error(f"❌ Error removing favorite: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/favorites")
async def get_favorites(
    skip: int = 0,
    limit: int = 20,
    current_user: dict = Depends(get_current_user),
    
):
    """Get user's favorite listings"""
    try:
        current_user_id = current_user["id"]
        favorites = await db.marketplace_favorites.find(
            {"user_id": current_user_id}
        ).skip(skip).limit(limit).to_list(limit)
        
        listing_ids = [f["listing_id"] for f in favorites]
        
        listings = await db.marketplace_listings.find(
            {"id": {"$in": listing_ids}}
        ).to_list(limit)
        
        for listing in listings:
            listing.pop("_id", None)
        
        return listings  # Direkt array döndür, reports.tsx düz array bekliyor
        
    except Exception as e:
        logger.error(f"❌ Error getting favorites: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# STATS & ANALYTICS
# ============================================

@router.get("/stats/me")
async def get_my_stats(
    current_user: dict = Depends(get_current_user),
    
):
    """Get marketplace stats for current user"""
    try:
        current_user_id = current_user["id"]
        stats = await db.marketplace_stats.find_one({"user_id": current_user_id})
        
        if not stats:
            stats = await update_user_marketplace_stats(db, current_user_id)
        else:
            stats.pop("_id", None)
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Error getting stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/stats/trending")
async def get_trending(
    category_id: Optional[str] = None,
    limit: int = 10,
    
):
    """Get trending listings"""
    try:
        query = {"status": ListingStatus.ACTIVE}
        if category_id:
            query["category_id"] = category_id
        
        # Sort by views and offers
        listings = await db.marketplace_listings.find(query).sort([
            ("views_count", -1),
            ("offer_count", -1)
        ]).limit(limit).to_list(limit)
        
        for listing in listings:
            listing.pop("_id", None)
        
        return {"trending": listings}
        
    except Exception as e:
        logger.error(f"❌ Error getting trending: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# HELPER FUNCTIONS
# ============================================

async def update_user_marketplace_stats(db: AsyncIOMotorDatabase, user_id: str):
    """Update user marketplace statistics"""
    try:
        # Count listings
        total_listings = await db.marketplace_listings.count_documents({"seller_id": user_id})
        active_listings = await db.marketplace_listings.count_documents({
            "seller_id": user_id,
            "status": ListingStatus.ACTIVE
        })
        sold_items = await db.marketplace_listings.count_documents({
            "seller_id": user_id,
            "status": ListingStatus.SOLD
        })
        
        # Count offers
        offers_received = await db.marketplace_offers.count_documents({"seller_id": user_id})
        offers_sent = await db.marketplace_offers.count_documents({"buyer_id": user_id})
        
        # Calculate revenue
        transactions = await db.marketplace_transactions.find(
            {"seller_id": user_id, "status": "completed"}
        ).to_list(1000)
        total_revenue = sum(t.get("seller_amount", 0) for t in transactions)
        
        # Get reviews
        reviews = await db.marketplace_reviews.find({"reviewee_id": user_id}).to_list(1000)
        total_reviews = len(reviews)
        average_rating = sum(r.get("rating", 0) for r in reviews) / total_reviews if total_reviews > 0 else 0
        
        # Get views and favorites
        listings = await db.marketplace_listings.find({"seller_id": user_id}).to_list(1000)
        views_count = sum(l.get("views_count", 0) for l in listings)
        favorite_count = sum(l.get("favorite_count", 0) for l in listings)
        
        stats = {
            "user_id": user_id,
            "total_listings": total_listings,
            "active_listings": active_listings,
            "sold_items": sold_items,
            "total_revenue": total_revenue,
            "total_offers_received": offers_received,
            "total_offers_sent": offers_sent,
            "average_rating": round(average_rating, 2),
            "total_reviews": total_reviews,
            "views_count": views_count,
            "favorite_count": favorite_count
        }
        
        await db.marketplace_stats.update_one(
            {"user_id": user_id},
            {"$set": stats},
            upsert=True
        )
        
        return stats
        
    except Exception as e:
        logger.error(f"❌ Error updating stats: {str(e)}")
        return {}


# ============================================
# ADMIN APPROVAL SYSTEM
# ============================================

@router.put("/listings/{listing_id}/approve")
async def approve_listing(listing_id: str, current_user: dict = Depends(get_current_user)):
    """Approve listing (admin only)"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Listing not found")
        
        await db.marketplace_listings.update_one({"id": listing_id}, {"$set": {"status": "active", "updated_at": datetime.utcnow()}})
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": listing["seller_id"],
            "type": "listing_approved",
            "title": "İlan Onaylandı",
            "message": f"'{listing['title']}' adlı ilanınız onaylandı ve yayınlandı.",
            "data": {"listing_id": listing_id},
            "action_url": f"/marketplace/{listing_id}",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        return {"message": "Listing approved"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/listings/{listing_id}/reject")
async def reject_listing(listing_id: str, reason: str, current_user: dict = Depends(get_current_user)):
    """Reject listing (admin only)"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") not in ["admin", "super_admin"]:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Listing not found")
        
        await db.marketplace_listings.update_one({"id": listing_id}, {"$set": {"status": "rejected", "rejection_reason": reason, "updated_at": datetime.utcnow()}})
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": listing["seller_id"],
            "type": "listing_rejected",
            "title": "İlan Reddedildi",
            "message": f"'{listing['title']}' adlı ilanınız reddedildi.\\n\\nRed Gerekçesi: {reason}",
            "data": {"listing_id": listing_id, "reason": reason},
            "action_url": f"/marketplace/create",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        return {"message": "Listing rejected"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/filter-options")
async def get_filter_options():
    """Get dynamic filter options from active listings"""
    try:
        # Get all active (unsold) listings
        listings = await db.marketplace_listings.find({
            "status": "active"
        }).to_list(length=None)
        
        # Extract unique values from tags and direct fields
        product_types = set()
        brands = set()
        sizes = set()
        genders = set()
        sports = set()
        listing_types = set()
        
        # Known product types for classification
        clothing_shoe_types = ['ayakkabı', 'tişört', 'şort', 'etek', 'tayt', 'forma', 'eşofman', 'çorap', 'eldiven']
        equipment_types = ['top', 'raket', 'file', 'kale', 'kask', 'koruyucu', 'çanta']
        
        # Category to sport name mapping
        category_sport_map = {
            'futbol': 'Futbol',
            'basketbol': 'Basketbol',
            'voleybol': 'Voleybol',
            'tenis': 'Tenis',
            'yüzme': 'Yüzme',
            'koşu': 'Koşu',
            'fitness': 'Fitness',
            'yoga': 'Yoga',
            'bisiklet': 'Bisiklet',
            'dağ-tırmanışı': 'Dağ Tırmanışı',
        }
        
        for listing in listings:
            # Extract from brand field directly
            if listing.get("brand"):
                brands.add(listing["brand"])
            
            # Extract sport from category_id
            if listing.get("category_id"):
                sport_name = category_sport_map.get(listing["category_id"], listing["category_id"].title())
                sports.add(sport_name)
            
            # Extract listing type
            if listing.get("listing_type"):
                listing_type_map = {
                    'product': 'Satılık',
                    'rental': 'Kiralık',
                    'service': 'Hizmet'
                }
                listing_type_name = listing_type_map.get(listing["listing_type"], listing["listing_type"])
                listing_types.add(listing_type_name)
            
            # Extract from model field as product type if relevant
            if listing.get("model"):
                model_lower = listing["model"].lower()
                # Check if model is a product type
                for known_type in clothing_shoe_types + equipment_types:
                    if known_type in model_lower:
                        product_types.add(listing["model"])
                        break
            
            # Extract from tags
            if listing.get("tags"):
                for tag in listing["tags"]:
                    tag_lower = tag.lower()
                    
                    # Check for prefix format (old listings)
                    if tag.startswith("Ürün Çeşidi:"):
                        product_types.add(tag.replace("Ürün Çeşidi:", "").strip())
                    elif tag.startswith("Marka:"):
                        brands.add(tag.replace("Marka:", "").strip())
                    elif tag.startswith("Beden:"):
                        sizes.add(tag.replace("Beden:", "").strip())
                    elif tag.startswith("Cinsiyet:"):
                        genders.add(tag.replace("Cinsiyet:", "").strip())
                    # Check for direct values (new listings)
                    elif tag in ['Erkek', 'Kadın', 'Unisex']:
                        genders.add(tag)
                    elif 'beden:' in tag_lower:
                        # Extract size from "Beden: 42" format
                        sizes.add(tag.split(':')[-1].strip())
                    else:
                        # Try to classify tag
                        # Check if it's a product type
                        for known_type in clothing_shoe_types + equipment_types:
                            if known_type in tag_lower:
                                product_types.add(tag)
                                break
        
        return {
            "product_types": sorted(list(product_types)),
            "brands": sorted(list(brands)),
            "sizes": sorted(list(sizes)),
            "genders": sorted(list(genders)),
            "sports": sorted(list(sports)),
            "listing_types": sorted(list(listing_types))
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting filter options: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# MARKETPLACE PAYMENT WITH IYZICO
# ============================================

@router.post("/listings/{listing_id}/purchase")
async def initiate_marketplace_purchase(
    listing_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """
    Initiate marketplace purchase with Iyzico
    - Direct purchase: use listing price + shipping (100 TL)
    - Via offer: use accepted offer price + shipping (100 TL)
    - Commission: 10% from seller
    """
    try:
        buyer_id = current_user.get("id")
        
        # Get offer_id and shipping info from request body
        try:
            body = await request.json() if request.method == "POST" else {}
            offer_id = body.get("offer_id")
            shipping_address = body.get("shipping_address")
            shipping_provider = body.get("shipping_provider")
            shipping_service = body.get("shipping_service")
            custom_shipping_fee = body.get("shipping_fee")
            logger.info(f"💳 Request body: {body}")
        except Exception as e:
            logger.error(f"❌ Error parsing request body: {e}")
            offer_id = None
            shipping_address = None
            shipping_provider = None
            shipping_service = None
            custom_shipping_fee = None
        
        logger.info(f"💳 Purchase request: listing_id={listing_id}, offer_id={offer_id}, buyer={buyer_id}")
        
        # Get listing
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="İlan bulunamadı")
        
        if listing["status"] != "active":
            raise HTTPException(status_code=400, detail="İlan aktif değil")
        
        if listing["seller_id"] == buyer_id:
            raise HTTPException(status_code=400, detail="Kendi ilanınızı satın alamazsınız")
        
        # Get buyer info
        buyer = await db.users.find_one({"id": buyer_id})
        if not buyer:
            raise HTTPException(status_code=404, detail="Alıcı bulunamadı")
        
        # Determine purchase price
        if offer_id:
            # Purchase via accepted offer
            offer = await db.marketplace_offers.find_one({"id": offer_id})
            if not offer:
                raise HTTPException(status_code=404, detail="Teklif bulunamadı")
            if offer["status"] != "accepted":
                raise HTTPException(status_code=400, detail="Teklif kabul edilmemiş")
            if offer["buyer_id"] != buyer_id:
                raise HTTPException(status_code=403, detail="Bu teklif size ait değil")
            
            item_price = offer["amount"]
            purchase_type = "offer"
        else:
            # Direct purchase
            item_price = listing["price"]
            purchase_type = "direct"
        
        # Get commission settings for shipping fee and commission rate
        commission_settings = await db.settings.find_one({"type": "commission"})
        
        # Calculate prices
        # Kiralık ürünlerde kargo ücreti yok
        is_rental = listing.get("listing_type") == "rental"
        
        # Get shipping fee from commission settings or use custom shipping fee from Geliver
        if commission_settings:
            default_shipping_fee = 0.0 if is_rental else float(commission_settings.get("marketplace_commission_fixed", 100.0))
            COMMISSION_RATE = float(commission_settings.get("marketplace_commission_rate", 10.0)) / 100.0
        else:
            default_shipping_fee = 0.0 if is_rental else 100.0  # Default shipping fee
            COMMISSION_RATE = 0.10  # Default 10% commission
        
        # Use custom shipping fee from Geliver if provided
        SHIPPING_FEE = float(custom_shipping_fee) if custom_shipping_fee is not None else default_shipping_fee
        
        total_price_for_buyer = item_price + SHIPPING_FEE
        commission_amount = item_price * COMMISSION_RATE
        seller_receives = item_price - commission_amount
        
        # Create transaction record
        transaction_id = str(uuid.uuid4())
        transaction = {
            "id": transaction_id,
            "listing_id": listing_id,
            "offer_id": offer_id,
            "buyer_id": buyer_id,
            "seller_id": listing["seller_id"],
            "purchase_type": purchase_type,
            "item_price": item_price,
            "shipping_fee": SHIPPING_FEE,
            "shipping_provider": shipping_provider,
            "shipping_service": shipping_service,
            "shipping_address": shipping_address,
            "commission_amount": commission_amount,
            "total_paid_by_buyer": total_price_for_buyer,
            "seller_receives": seller_receives,
            "status": "pending_payment",
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        await db.marketplace_transactions.insert_one(transaction)
        logger.info(f"💰 Transaction {transaction_id} created for listing {listing_id}")
        
        # Initialize Iyzico checkout
        backend_base_url = os.getenv("BACKEND_BASE_URL", "http://localhost:8001")
        callback_url = f"{backend_base_url}/api/marketplace/payment-callback"
        
        logger.info(f"💳 İyzico callback URL: {callback_url}")
        
        result = iyzico_service.initialize_checkout_form(
            user={
                'id': buyer["id"],
                'email': buyer["email"],
                'full_name': buyer["full_name"],
                'phone_number': buyer.get('phone', '+905000000000'),
                'tc_kimlik': buyer.get('tckn', '11111111111'),
                'created_at': buyer.get('created_at', datetime.utcnow())
            },
            amount=total_price_for_buyer,
            related_type='marketplace',
            related_id=transaction_id,
            related_name=f"{listing['title']} - Market Alışverişi",
            callback_url=callback_url
        )
        
        if result.get('status') == 'success':
            # Update transaction with payment token
            await db.marketplace_transactions.update_one(
                {"id": transaction_id},
                {"$set": {"payment_token": result["token"], "updated_at": datetime.utcnow()}}
            )
            
            logger.info(f"✅ Iyzico checkout initialized for transaction {transaction_id}")
            
            # Convert camelCase keys from Iyzico to snake_case for frontend
            return {
                "success": True,
                "transaction_id": transaction_id,
                "payment_page_url": result.get("paymentPageUrl"),  # camelCase -> snake_case
                "token": result["token"],
                "checkout_form_content": result.get("checkoutFormContent"),  # camelCase -> snake_case
                "price_breakdown": {
                    "item_price": item_price,
                    "shipping_fee": SHIPPING_FEE,
                    "total": total_price_for_buyer,
                    "seller_receives": seller_receives,
                    "commission": commission_amount,
                    "is_rental": is_rental  # Kiralık mı bilgisini de ekleyelim
                }
            }
        else:
            logger.error(f"❌ Iyzico checkout failed: {result.get('error_message')}")
            raise HTTPException(status_code=500, detail=f"Ödeme başlatılamadı: {result.get('error_message')}")
            
    except HTTPException as he:
        logger.error(f"❌ HTTPException in purchase: {he.status_code} - {he.detail}")
        raise
    except Exception as e:
        logger.error(f"❌ Purchase initiation error: {str(e)}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/payment-callback")
@router.get("/payment-callback")
async def marketplace_payment_callback(
    request: Request,
    token: Optional[str] = None
):
    """Handle Iyzico payment callback for marketplace purchase"""
    try:
        # Try to get token from query params or form data
        if not token:
            form_data = await request.form()
            token = form_data.get("token")
        
        if not token:
            logger.error("❌ Token not found in callback request")
            raise HTTPException(status_code=400, detail="Token parametresi eksik")
        
        logger.info(f"📥 Marketplace payment callback received, token: {token}")
        
        # Retrieve payment result from Iyzico
        result = iyzico_service.retrieve_checkout_form_result(token)
        
        if result.get('status') != 'success':
            logger.error(f"❌ Payment failed: {result.get('error_message')}")
            raise HTTPException(status_code=400, detail="Ödeme başarısız")
        
        # Get transaction
        transaction = await db.marketplace_transactions.find_one({"payment_token": token})
        if not transaction:
            logger.error(f"❌ Transaction not found for token: {token}")
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        transaction_id = transaction["id"]
        listing_id = transaction["listing_id"]
        buyer_id = transaction["buyer_id"]
        seller_id = transaction["seller_id"]
        
        # Update transaction status
        await db.marketplace_transactions.update_one(
            {"id": transaction_id},
            {
                "$set": {
                    "status": "completed",
                    "paid_at": datetime.utcnow(),
                    "payment_details": result,
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        # Update listing status to sold
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {
                "$set": {
                    "status": "sold",
                    "sold_at": datetime.utcnow(),
                    "buyer_id": buyer_id,
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        logger.info(f"✅ Marketplace payment completed for transaction {transaction_id}")
        
        # Get listing and users for notifications
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        buyer = await db.users.find_one({"id": buyer_id})
        seller = await db.users.find_one({"id": seller_id})
        
        # Re-fetch updated transaction
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        
        # ===== GELIVER GÖNDERI OLUŞTUR =====
        # Ödeme tamamlandıktan sonra Geliver'da gerçek gönderi oluştur
        geliver_result = await create_geliver_shipment_after_payment(
            transaction=transaction,
            listing=listing,
            buyer=buyer,
            seller=seller,
            database=db
        )
        
        logger.info(f"📦 Geliver gönderi sonucu: {geliver_result}")
        
        # Geliver sonuçlarını al
        tracking_code = geliver_result.get("barcode", "") or geliver_result.get("tracking_code", "")
        label_url = geliver_result.get("label_url", "")
        shipping_provider = transaction.get("shipping_provider", "")
        shipping_provider_name = get_provider_name(shipping_provider)
        
        if not geliver_result.get("success"):
            logger.warning(f"Geliver gönderi oluşturulamadı: {geliver_result.get('error')}")
            # Yine de devam et, sadece kargo kodu olmadan
        
        # 1. Notification to BUYER
        await create_notification_helper(
            db=db,
            notification_data={
                "user_id": buyer_id,
                "notification_type": "purchase_success",
                "title": "Satın Alma Başarılı!",
                "message": f"'{listing['title']}' ürününü başarıyla satın aldınız. Satıcı: {seller.get('full_name')}. Telefon: {seller.get('phone', 'Belirtilmemiş')}",
                "related_type": "marketplace_listing",
                "related_id": listing_id,
                "is_read": False,
                "data": {
                    "transaction_id": transaction_id,
                    "listing_id": listing_id,
                    "seller_name": seller.get('full_name'),
                    "seller_phone": seller.get('phone'),
                    "total_paid": transaction["total_paid_by_buyer"]
                }
            }
        )
        
        # 2. Notification to SELLER with shipping info
        seller_message = f"'{listing['title']}' ürününüz satıldı!\n\nAlıcı: {buyer.get('full_name')}\nTelefon: {buyer.get('phone', 'Belirtilmemiş')}\nKazanç: ₺{transaction['seller_receives']:.2f}"
        
        if shipping_provider_name:
            seller_message += f"\n\n📦 KARGO BİLGİLERİ:"
            seller_message += f"\nKargo Firması: {shipping_provider_name}"
            if tracking_code:
                seller_message += f"\n🔢 Kargo Kodu: {tracking_code}"
            if label_url:
                seller_message += f"\n🏷️ Etiket: {label_url}"
            seller_message += "\n\n⚠️ Bu kargo kodu ile 2 iş günü içerisinde kargo gönderimi yapınız."
        else:
            seller_message += "\n\n⚠️ Lütfen alıcı ile iletişime geçip kargo bilgilerini paylaşınız."
        
        await create_notification_helper(
            db=db,
            notification_data={
                "user_id": seller_id,
                "notification_type": "item_sold",
                "title": "Ürün Satıldı! 📦 Kargo Gönderiniz",
                "message": seller_message,
                "related_type": "marketplace_listing",
                "related_id": listing_id,
                "is_read": False,
                "data": {
                    "transaction_id": transaction_id,
                    "listing_id": listing_id,
                    "buyer_name": buyer.get('full_name'),
                    "buyer_phone": buyer.get('phone'),
                    "seller_receives": transaction["seller_receives"],
                    "commission": transaction["commission_amount"],
                    "shipping_provider": shipping_provider,
                    "shipping_provider_name": shipping_provider_name,
                    "tracking_code": tracking_code,
                    "label_url": label_url
                }
            }
        )
        
        # 3. Notifications to ALL ADMINS
        admins = await db.users.find({"user_type": {"$in": ["admin", "super_admin"]}}).to_list(100)
        for admin in admins:
            await create_notification_helper(
                db=db,
                notification_data={
                    "user_id": admin["id"],
                    "notification_type": "marketplace_sale",
                    "title": "Market Satışı Tamamlandı",
                    "message": f"'{listing['title']}' satıldı. Alıcı: {buyer.get('full_name')}, Satıcı: {seller.get('full_name')}. Komisyon: ₺{transaction['commission_amount']:.2f}",
                    "related_type": "marketplace_listing",
                    "related_id": listing_id,
                    "is_read": False,
                    "data": {
                        "transaction_id": transaction_id,
                        "listing_id": listing_id,
                        "buyer_name": buyer.get('full_name'),
                        "seller_name": seller.get('full_name'),
                        "commission": transaction["commission_amount"],
                        "total_amount": transaction["total_paid_by_buyer"]
                    }
                }
            )
        
        logger.info(f"📧 Notifications sent: Buyer, Seller, {len(admins)} Admins")
        
        # Marketplace satın alma log'u (Alıcı için)
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(buyer_id, "marketplace_purchase", "success", {
                "transaction_id": transaction_id,
                "listing_id": listing_id,
                "listing_title": listing.get("title"),
                "amount": transaction.get("total_paid_by_buyer"),
                "seller_id": seller_id,
                "seller_name": seller.get("full_name")
            })
            # Marketplace satış log'u (Satıcı için)
            await log_user_activity(seller_id, "marketplace_sale", "success", {
                "transaction_id": transaction_id,
                "listing_id": listing_id,
                "listing_title": listing.get("title"),
                "amount": transaction.get("seller_receives"),
                "commission": transaction.get("commission_amount"),
                "buyer_id": buyer_id,
                "buyer_name": buyer.get("full_name")
            })
        except Exception as log_err:
            logger.error(f"Log error: {log_err}")
        
        # Redirect to success page using HTML (for iframe compatibility)
        frontend_url = os.getenv("FRONTEND_URL", "http://localhost:3000")
        success_url = f"{frontend_url}/marketplace/purchase-success?transaction_id={transaction_id}"
        
        # Return HTML that redirects the parent window
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Ödeme Başarılı</title>
            <script>
                // Try to redirect parent window (for iframe)
                if (window.parent && window.parent !== window) {{
                    window.parent.postMessage({{
                        type: 'payment_success',
                        status: 'success',
                        transaction_id: '{transaction_id}',
                        redirect_url: '{success_url}'
                    }}, '*');
                    // Also try direct redirect of parent
                    try {{
                        window.parent.location.href = '{success_url}';
                    }} catch(e) {{
                        // If blocked by CORS, redirect self
                        window.location.href = '{success_url}';
                    }}
                }} else {{
                    // Direct window redirect
                    window.location.href = '{success_url}';
                }}
            </script>
        </head>
        <body style="display:flex;align-items:center;justify-content:center;height:100vh;font-family:system-ui;background:#1a1a2e;color:white;">
            <div style="text-align:center;">
                <div style="font-size:48px;margin-bottom:16px;">✅</div>
                <h2 style="margin-bottom:8px;">Ödeme Başarılı!</h2>
                <p style="color:#888;">Yönlendiriliyorsunuz...</p>
                <p style="margin-top:16px;"><a href="{success_url}" style="color:#00bcd4;">Tıklayın</a> otomatik yönlendirme olmazsa.</p>
            </div>
        </body>
        </html>
        """
        
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html_content, status_code=200)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Payment callback error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/transactions/{transaction_id}")
async def get_transaction(
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get marketplace transaction details with listing info"""
    try:
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        # Check access permission
        user_id = current_user.get("id")
        if transaction["buyer_id"] != user_id and transaction["seller_id"] != user_id:
            # Check if admin
            user = await db.users.find_one({"id": user_id})
            if not user or user.get("user_type") not in ["admin", "super_admin"]:
                raise HTTPException(status_code=403, detail="Bu işlemi görüntüleme yetkiniz yok")
        
        transaction.pop("_id", None)
        
        # Listing bilgisini ekle
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        if listing:
            listing.pop("_id", None)
            transaction["listing"] = listing
        
        return transaction
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Get transaction error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/transactions/{transaction_id}/status")
async def get_transaction_status(
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get marketplace transaction status (for polling)"""
    try:
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        # Check access permission
        user_id = current_user.get("id")
        if transaction["buyer_id"] != user_id and transaction["seller_id"] != user_id:
            # Check if admin
            user = await db.users.find_one({"id": user_id})
            if not user or user.get("user_type") not in ["admin", "super_admin"]:
                raise HTTPException(status_code=403, detail="Bu işlemi görüntüleme yetkiniz yok")
        
        transaction.pop("_id", None)
        return transaction
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Get transaction status error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== SİPARİŞ DURUM YÖNETİMİ ====================

@router.put("/transactions/{transaction_id}/confirm")
async def confirm_order(transaction_id: str, current_user: dict = Depends(get_current_user)):
    """Satıcı siparişi onaylar"""
    try:
        current_user_id = current_user["id"]
        
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        if transaction.get("seller_id") != current_user_id:
            raise HTTPException(status_code=403, detail="Bu işlemi onaylama yetkiniz yok")
        
        if transaction.get("status") != "pending":
            raise HTTPException(status_code=400, detail="Bu sipariş zaten işlem görmüş")
        
        await db.marketplace_transactions.update_one(
            {"id": transaction_id},
            {"$set": {"status": "confirmed", "confirmed_at": datetime.utcnow()}}
        )
        
        # Alıcıya bildirim gönder
        buyer_id = transaction.get("buyer_id")
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        
        await send_notification(
            user_id=buyer_id,
            title="Sipariş Onaylandı 🎉",
            message=f"'{listing.get('title', 'Ürün')}' siparişiniz satıcı tarafından onaylandı.",
            notification_type="order_confirmed",
            data={"transaction_id": transaction_id}
        )
        
        return {"message": "Sipariş onaylandı", "status": "confirmed"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Confirm order error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/transactions/{transaction_id}/ship")
async def mark_shipped(transaction_id: str, current_user: dict = Depends(get_current_user)):
    """Satıcı siparişi kargoya verir"""
    try:
        current_user_id = current_user["id"]
        
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        if transaction.get("seller_id") != current_user_id:
            raise HTTPException(status_code=403, detail="Bu işlemi güncelleme yetkiniz yok")
        
        if transaction.get("status") not in ["pending", "confirmed"]:
            raise HTTPException(status_code=400, detail="Bu sipariş kargoya verilemez")
        
        await db.marketplace_transactions.update_one(
            {"id": transaction_id},
            {"$set": {"status": "shipped", "shipped_at": datetime.utcnow()}}
        )
        
        # Alıcıya bildirim gönder
        buyer_id = transaction.get("buyer_id")
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        tracking_code = transaction.get("tracking_code", "")
        
        await send_notification(
            user_id=buyer_id,
            title="Kargonuz Yola Çıktı 📦",
            message=f"'{listing.get('title', 'Ürün')}' kargoya verildi. Takip kodu: {tracking_code}",
            notification_type="order_shipped",
            data={"transaction_id": transaction_id, "tracking_code": tracking_code}
        )
        
        return {"message": "Kargoya verildi olarak işaretlendi", "status": "shipped"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Mark shipped error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/transactions/{transaction_id}/deliver")
async def mark_delivered(transaction_id: str, current_user: dict = Depends(get_current_user)):
    """Alıcı ürünü teslim aldı"""
    try:
        current_user_id = current_user["id"]
        
        transaction = await db.marketplace_transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="İşlem bulunamadı")
        
        if transaction.get("buyer_id") != current_user_id:
            raise HTTPException(status_code=403, detail="Bu işlemi güncelleme yetkiniz yok")
        
        if transaction.get("status") != "shipped":
            raise HTTPException(status_code=400, detail="Bu sipariş henüz kargoda değil")
        
        await db.marketplace_transactions.update_one(
            {"id": transaction_id},
            {"$set": {"status": "delivered", "delivered_at": datetime.utcnow()}}
        )
        
        # Satıcıya bildirim gönder
        seller_id = transaction.get("seller_id")
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        
        await send_notification(
            user_id=seller_id,
            title="Ürün Teslim Edildi ✅",
            message=f"'{listing.get('title', 'Ürün')}' alıcıya teslim edildi.",
            notification_type="order_delivered",
            data={"transaction_id": transaction_id}
        )
        
        return {"message": "Teslim alındı olarak işaretlendi", "status": "delivered"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Mark delivered error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== RAPORLAMA ENDPOINTS ====================

@router.get("/my-purchases")
async def get_my_purchases(current_user: dict = Depends(get_current_user)):
    """Kullanıcının satın aldığı ürünler"""
    try:
        current_user_id = current_user["id"]
        transactions = await db.marketplace_transactions.find({
            "buyer_id": current_user_id
        }).sort("created_at", -1).to_list(100)
        
        # Her transaction için listing bilgisini ekle
        result = []
        for tx in transactions:
            listing = await db.marketplace_listings.find_one({"id": tx["listing_id"]})
            if listing:
                listing.pop("_id", None)
                tx.pop("_id", None)
                result.append({
                    "transaction": tx,
                    "listing": listing,
                    "id": tx.get("id"),
                    "status": tx.get("status"),
                    "tracking_code": tx.get("tracking_code"),
                    "barcode": tx.get("barcode"),
                    "buyer_id": tx.get("buyer_id"),
                    "seller_id": tx.get("seller_id"),
                })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get purchases error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/my-sales")
async def get_my_sales(current_user: dict = Depends(get_current_user)):
    """Kullanıcının sattığı ürünler (tüm durumlar)"""
    try:
        current_user_id = current_user["id"]
        
        # Transaction'ları seller_id ile bul
        transactions = await db.marketplace_transactions.find({
            "seller_id": current_user_id
        }).sort("created_at", -1).to_list(100)
        
        result = []
        for tx in transactions:
            listing = await db.marketplace_listings.find_one({"id": tx["listing_id"]})
            if listing:
                listing.pop("_id", None)
                tx.pop("_id", None)
                result.append({
                    "transaction": tx,
                    "listing": listing,
                    "id": tx.get("id"),
                    "status": tx.get("status"),
                    "tracking_code": tx.get("tracking_code"),
                    "barcode": tx.get("barcode"),
                    "buyer_id": tx.get("buyer_id"),
                    "seller_id": tx.get("seller_id"),
                })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get sales error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/my-rentals")
async def get_my_rentals(current_user: dict = Depends(get_current_user)):
    """Kullanıcının kiraladığı ürünler"""
    try:
        current_user_id = current_user["id"]
        transactions = await db.marketplace_transactions.find({
            "buyer_id": current_user_id,
            "status": "completed"
        }).sort("created_at", -1).to_list(100)
        
        # Sadece kiralık ürünleri filtrele
        result = []
        for tx in transactions:
            listing = await db.marketplace_listings.find_one({
                "id": tx["listing_id"],
                "listing_type": "rental"
            })
            if listing:
                listing.pop("_id", None)
                tx.pop("_id", None)
                result.append({
                    "transaction": tx,
                    "listing": listing
                })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get rentals error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/my-offers")
async def get_my_offers(current_user: dict = Depends(get_current_user)):
    """Kullanıcının verdiği teklifler"""
    try:
        current_user_id = current_user["id"]
        offers = await db.marketplace_offers.find({
            "buyer_id": current_user_id
        }).sort("created_at", -1).to_list(100)
        
        # Her teklif için listing bilgisini ekle
        result = []
        for offer in offers:
            listing = await db.marketplace_listings.find_one({"id": offer["listing_id"]})
            if listing:
                listing.pop("_id", None)
                offer.pop("_id", None)
                result.append({
                    "offer": offer,
                    "listing": listing
                })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get offers error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ============================================
# ORDERS & RETURNS SYSTEM
# ============================================

class ReturnRequest(BaseModel):
    reason: str
    description: Optional[str] = None

@router.get("/my-orders")
async def get_my_orders(current_user: dict = Depends(get_current_user)):
    """Kullanıcının siparişleri"""
    try:
        current_user_id = current_user["id"]
        
        # Tüm satın alma işlemlerini getir (pending_payment dahil)
        transactions = await db.marketplace_transactions.find({
            "buyer_id": current_user_id,
            "status": {"$in": ["pending_payment", "completed", "pending", "confirmed", "shipped", "delivered", "approved", "cancelled", 
                               "return_requested", "return_approved", "return_shipped", "return_completed", "return_rejected"]}
        }).sort("created_at", -1).to_list(100)
        
        orders = []
        for tx in transactions:
            listing = await db.marketplace_listings.find_one({"id": tx["listing_id"]})
            if listing:
                # İade süresi kontrolü (14 gün)
                created_at = tx.get("delivered_at") or tx.get("created_at")
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                
                return_deadline = created_at + timedelta(days=14) if created_at else None
                can_return = tx.get("status") == "delivered" and return_deadline and datetime.utcnow() < return_deadline
                
                # Get first image safely
                images = listing.get("images", []) or []
                first_image = images[0] if images else None
                
                orders.append({
                    "id": tx.get("id"),
                    "transaction_id": tx.get("id"),
                    "listing_id": tx.get("listing_id"),
                    "listing_title": listing.get("title"),
                    "listing_image": first_image,
                    "seller_name": listing.get("seller_name"),
                    "seller_id": listing.get("seller_id"),
                    "price": tx.get("amount") or listing.get("price"),
                    "shipping_address": tx.get("shipping_address"),
                    "status": tx.get("status", "pending"),
                    "created_at": tx.get("created_at"),
                    "shipped_at": tx.get("shipped_at"),
                    "delivered_at": tx.get("delivered_at"),
                    "tracking_code": tx.get("tracking_code"),
                    "can_return": can_return,
                    "return_deadline": return_deadline.isoformat() if return_deadline else None
                })
        
        return orders
    except Exception as e:
        logger.error(f"❌ Get orders error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/orders/{order_id}/return")
async def request_return(order_id: str, return_data: ReturnRequest, current_user: dict = Depends(get_current_user)):
    """İade talebi oluştur (ESKİ - /returns/create kullanın)"""
    try:
        current_user_id = current_user["id"]
        
        # Siparişi bul
        transaction = await db.marketplace_transactions.find_one({
            "id": order_id,
            "buyer_id": current_user_id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        if transaction.get("status") not in ["delivered", "completed"]:
            raise HTTPException(status_code=400, detail="Sadece teslim edilmiş siparişler için iade talep edilebilir")
        
        # İade süresi kontrolü - 1 iş günü
        delivered_at = transaction.get("delivered_at") or transaction.get("created_at")
        if isinstance(delivered_at, str):
            delivered_at = datetime.fromisoformat(delivered_at.replace('Z', '+00:00'))
        
        if delivered_at:
            # 1 iş günü hesapla (hafta sonlarını say)
            deadline = add_business_days(delivered_at, 1)
            if datetime.utcnow() > deadline:
                raise HTTPException(status_code=400, detail="İade süresi dolmuş (teslimattan sonra 1 iş günü)")
        
        # İade talebini kaydet
        return_request = {
            "id": str(uuid.uuid4()),
            "order_id": order_id,
            "transaction_id": order_id,
            "buyer_id": current_user_id,
            "seller_id": transaction.get("seller_id"),
            "listing_id": transaction.get("listing_id"),
            "reason": return_data.reason,
            "description": return_data.description,
            "status": "pending",
            "created_at": datetime.utcnow()
        }
        
        await db.marketplace_returns.insert_one(return_request)
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": order_id},
            {"$set": {"status": "return_requested", "return_request_id": return_request["id"]}}
        )
        
        # Satıcıya bildirim gönder
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        
        # İade sebebi açıklamaları
        reason_labels = {
            "size_issue": "Beden Uygunsuz",
            "fake_product": "Taklit Ürün",
            "undisclosed_defect": "Bildirilmeyen Defo",
        }
        reason_text = reason_labels.get(return_data.reason, return_data.reason)
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": transaction.get("seller_id"),
            "type": "return_requested",
            "title": "📦 İade Talebi Alındı",
            "message": f"'{listing.get('title', 'Ürün')}' için iade talebi alındı.\n\nSebep: {reason_text}\n\n2 iş günü içinde yanıt vermeniz gerekmektedir.",
            "data": {"order_id": order_id, "return_id": return_request["id"]},
            "action_url": f"/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        # İade talebi log'u
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "marketplace_return_request", "success", {
                "order_id": order_id,
                "return_id": return_request["id"],
                "listing_title": listing.get("title") if listing else "",
                "reason": return_data.reason,
                "seller_id": transaction.get("seller_id")
            })
        except Exception as log_err:
            logger.error(f"Log error: {log_err}")
        
        return {"message": "İade talebi oluşturuldu", "return_id": return_request["id"]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Return request error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class CreateReturnRequest(BaseModel):
    transaction_id: str
    reason: str
    description: Optional[str] = None
    images: Optional[List[str]] = []  # Fotoğraflar (base64)


@router.post("/returns/create")
async def create_return_request(data: CreateReturnRequest, current_user: dict = Depends(get_current_user)):
    """
    Yeni iade talebi oluştur
    - Teslimattan sonra 1 iş günü içinde yapılmalı
    - "Beğenmedim" sebebiyle iade kabul edilmez
    - En az 1 fotoğraf zorunlu
    """
    try:
        current_user_id = current_user["id"]
        
        # Fotoğraf kontrolü
        if not data.images or len(data.images) < 1:
            raise HTTPException(status_code=400, detail="En az 1 fotoğraf eklemeniz gerekiyor")
        
        if len(data.images) > 5:
            raise HTTPException(status_code=400, detail="En fazla 5 fotoğraf ekleyebilirsiniz")
        
        # Siparişi bul
        transaction = await db.marketplace_transactions.find_one({
            "id": data.transaction_id,
            "buyer_id": current_user_id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        if transaction.get("status") not in ["delivered", "completed"]:
            raise HTTPException(status_code=400, detail="Sadece teslim edilmiş siparişler için iade talep edilebilir")
        
        # Daha önce iade talebi var mı kontrol et
        existing_return = await db.marketplace_returns.find_one({
            "order_id": data.transaction_id,
            "buyer_id": current_user_id
        })
        if existing_return:
            raise HTTPException(status_code=400, detail="Bu sipariş için zaten bir iade talebi mevcut")
        
        # İade süresi kontrolü - 1 iş günü
        delivered_at = transaction.get("delivered_at") or transaction.get("created_at")
        if isinstance(delivered_at, str):
            delivered_at = datetime.fromisoformat(delivered_at.replace('Z', '+00:00'))
        
        if delivered_at:
            deadline = add_business_days(delivered_at, 1)
            now = datetime.utcnow()
            if now > deadline:
                raise HTTPException(
                    status_code=400, 
                    detail=f"İade süresi dolmuş. Teslimattan sonra 1 iş günü ({deadline.strftime('%d.%m.%Y %H:%M')}) içinde iade talebi oluşturabilirsiniz."
                )
        
        # Geçerli sebep mi kontrol et
        valid_reasons = ["size_issue", "fake_product", "undisclosed_defect", "not_liked"]
        if data.reason not in valid_reasons:
            raise HTTPException(status_code=400, detail="Geçersiz iade sebebi")
        
        # İade talebini kaydet
        return_request = {
            "id": str(uuid.uuid4()),
            "order_id": data.transaction_id,
            "transaction_id": data.transaction_id,
            "buyer_id": current_user_id,
            "seller_id": transaction.get("seller_id"),
            "listing_id": transaction.get("listing_id"),
            "reason": data.reason,
            "description": data.description,
            "images": data.images,  # Fotoğraflar
            "status": "pending",
            "created_at": datetime.utcnow(),
            "return_deadline": add_business_days(datetime.utcnow(), 4)  # Onaylanırsa 4 iş günü kargo süresi
        }
        
        await db.marketplace_returns.insert_one(return_request)
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": data.transaction_id},
            {"$set": {"status": "return_requested", "return_request_id": return_request["id"]}}
        )
        
        # Satıcıya bildirim gönder
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        
        # İade sebebi açıklamaları
        reason_labels = {
            "size_issue": "Beden Uygunsuz",
            "fake_product": "Taklit Ürün",
            "undisclosed_defect": "Bildirilmeyen Defo",
        }
        reason_text = reason_labels.get(data.reason, data.reason)
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": transaction.get("seller_id"),
            "type": "return_requested",
            "title": "📦 İade Talebi Alındı",
            "message": f"'{listing.get('title', 'Ürün')}' için iade talebi alındı.\n\nSebep: {reason_text}\n{data.description or ''}\n\n2 iş günü içinde yanıt vermeniz gerekmektedir. Onaylamazsanız konu yöneticilere iletilecektir.",
            "data": {"order_id": data.transaction_id, "return_id": return_request["id"]},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ İade talebi oluşturuldu - Return ID: {return_request['id']}, Reason: {data.reason}")
        
        return {
            "message": "İade talebi oluşturuldu", 
            "return_id": return_request["id"],
            "status": "pending"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Create return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/my-returns")
async def get_my_returns(current_user: dict = Depends(get_current_user)):
    """Kullanıcının kendi iade talepleri (alıcı ve satıcı olarak)"""
    try:
        current_user_id = current_user["id"]
        
        # Hem alıcı hem satıcı olarak iade taleplerini getir
        returns = await db.marketplace_returns.find({
            "$or": [
                {"buyer_id": current_user_id},
                {"seller_id": current_user_id}
            ]
        }).sort("created_at", -1).to_list(100)
        
        result = []
        for ret in returns:
            ret.pop("_id", None)
            listing = await db.marketplace_listings.find_one({"id": ret.get("listing_id")})
            if listing:
                listing.pop("_id", None)
            transaction = await db.marketplace_transactions.find_one({"id": ret.get("order_id")})
            if transaction:
                transaction.pop("_id", None)
            
            result.append({
                "id": ret.get("id"),
                "status": ret.get("status", "pending"),
                "reason": ret.get("reason"),
                "created_at": ret.get("created_at"),
                "listing": listing,
                "transaction": transaction,
                "buyer_id": ret.get("buyer_id"),
                "seller_id": ret.get("seller_id"),
            })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get my returns error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/return-requests")
async def get_return_requests(current_user: dict = Depends(get_current_user)):
    """Satıcının aldığı iade talepleri"""
    try:
        current_user_id = current_user["id"]
        
        returns = await db.marketplace_returns.find({
            "seller_id": current_user_id
        }).sort("created_at", -1).to_list(100)
        
        result = []
        for ret in returns:
            ret.pop("_id", None)
            listing = await db.marketplace_listings.find_one({"id": ret.get("listing_id")})
            buyer = await db.users.find_one({"id": ret.get("buyer_id")})
            
            result.append({
                **ret,
                "listing_title": listing.get("title") if listing else None,
                "listing_image": listing.get("images", [None])[0] if listing else None,
                "buyer_name": buyer.get("name") if buyer else None,
            })
        
        return result
    except Exception as e:
        logger.error(f"❌ Get return requests error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/return-requests/{return_id}/approve")
async def approve_return(return_id: str, current_user: dict = Depends(get_current_user)):
    """İade talebini onayla ve Geliver'dan iade kargo etiketi oluştur"""
    try:
        current_user_id = current_user["id"]
        
        return_request = await db.marketplace_returns.find_one({
            "id": return_id,
            "seller_id": current_user_id
        })
        
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        if return_request.get("status") != "pending":
            raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
        
        # Transaction, listing, buyer ve seller bilgilerini al
        transaction = await db.marketplace_transactions.find_one({"id": return_request.get("order_id")})
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        buyer = await db.users.find_one({"id": return_request.get("buyer_id")})
        seller = await db.users.find_one({"id": current_user_id})
        
        # Geliver'dan iade kargo etiketi oluştur
        logger.info(f"🔄 İade için Geliver gönderi oluşturuluyor - Return ID: {return_id}")
        geliver_result = await create_geliver_return_shipment(
            return_request=return_request,
            transaction=transaction,
            listing=listing or {},
            buyer=buyer or {},
            seller=seller or {}
        )
        
        # İade kargo bilgilerini hazırla
        return_tracking_code = geliver_result.get("tracking_code", "")
        return_barcode = geliver_result.get("barcode", "")
        return_label_url = geliver_result.get("label_url", "")
        return_shipping_cost = geliver_result.get("shipping_cost", 0)
        return_provider_name = geliver_result.get("provider_name", "Kargo")
        
        # İade talebini güncelle
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {
                "status": "approved", 
                "approved_at": datetime.utcnow(),
                "return_tracking_code": return_tracking_code,
                "return_barcode": return_barcode,
                "return_label_url": return_label_url,
                "return_shipping_cost": return_shipping_cost,
                "return_provider_name": return_provider_name
            }}
        )
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {
                "status": "return_approved",
                "return_tracking_code": return_tracking_code,
                "return_barcode": return_barcode
            }}
        )
        
        # Alıcıya bildirim - Kargo ücreti bilgisi ile
        shipping_cost_text = f"İade kargo ücreti: ₺{return_shipping_cost:.2f}" if return_shipping_cost > 0 else ""
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("buyer_id"),
            "type": "return_approved",
            "title": "İade Onaylandı ✅",
            "message": f"İade talebiniz onaylandı. Lütfen ürünü kargoya verin.\n\n📦 Kargo Firması: {return_provider_name}\n🔢 Takip Kodu: {return_tracking_code}\n{shipping_cost_text}\n\n⚠️ İade kargo ücreti size aittir.",
            "data": {
                "return_id": return_id,
                "tracking_code": return_tracking_code,
                "barcode": return_barcode,
                "label_url": return_label_url,
                "shipping_cost": return_shipping_cost,
                "provider_name": return_provider_name
            },
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ İade onaylandı - Return ID: {return_id}, Tracking: {return_tracking_code}")
        
        return {
            "message": "İade talebi onaylandı",
            "tracking_code": return_tracking_code,
            "barcode": return_barcode,
            "label_url": return_label_url,
            "shipping_cost": return_shipping_cost,
            "provider_name": return_provider_name
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Approve return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/return-requests/{return_id}/reject")
async def reject_return(return_id: str, reason: str = "", current_user: dict = Depends(get_current_user)):
    """İade talebini reddet - Konu SportsMarket yöneticilerine iletilir"""
    try:
        current_user_id = current_user["id"]
        
        return_request = await db.marketplace_returns.find_one({
            "id": return_id,
            "seller_id": current_user_id
        })
        
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        if return_request.get("status") != "pending":
            raise HTTPException(status_code=400, detail="Bu talep zaten işlenmiş")
        
        # İade talebini "disputed" (itirazlı) olarak güncelle
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {
                "status": "disputed", 
                "seller_rejection_reason": reason, 
                "disputed_at": datetime.utcnow()
            }}
        )
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {"status": "return_disputed"}}
        )
        
        # Listing ve kullanıcı bilgilerini al
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        buyer = await db.users.find_one({"id": return_request.get("buyer_id")})
        seller = await db.users.find_one({"id": current_user_id})
        
        # İade sebebi açıklamaları
        reason_labels = {
            "size_issue": "Beden Uygunsuz",
            "fake_product": "Taklit Ürün",
            "undisclosed_defect": "Bildirilmeyen Defo",
        }
        return_reason_text = reason_labels.get(return_request.get("reason", ""), return_request.get("reason", ""))
        
        # Alıcıya bildirim - İtiraz edildi, yöneticiler inceleyecek
        buyer_notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("buyer_id"),
            "type": "return_disputed",
            "title": "⚠️ İade Talebine İtiraz Edildi",
            "message": f"Satıcı iade talebinize itiraz etti.\n\nSebep: {reason}\n\nKonu SportsMarket yöneticileri tarafından incelenmek üzere iletildi.",
            "data": {"return_id": return_id},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(buyer_notification)
        
        # TÜM Admin kullanıcılarına bildirim gönder
        admin_users = await db.users.find({"user_type": "admin"}).to_list(100)
        for admin in admin_users:
            admin_notification = {
                "id": str(uuid.uuid4()),
                "user_id": admin["id"],
                "type": "admin_return_dispute",
                "title": "🔴 İade İhtilafı - İnceleme Gerekiyor",
                "message": f"Satıcı itiraz ettiği için incelemeniz gereken bir iade talebi var.\n\n📦 Ürün: {listing.get('title', 'Ürün')}\n💰 Tutar: ₺{listing.get('price', 0)}\n\n👤 Alıcı: {buyer.get('full_name', 'Bilinmiyor')}\n🏪 Satıcı: {seller.get('full_name', 'Bilinmiyor')}\n\n📋 İade Sebebi: {return_reason_text}\n❌ Satıcı İtirazı: {reason}",
                "data": {
                    "return_id": return_id,
                    "listing_id": return_request.get("listing_id"),
                    "buyer_id": return_request.get("buyer_id"),
                    "seller_id": current_user_id,
                    "type": "return_dispute"
                },
                "action_url": "/admin/marketplace",
                "read": False,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(admin_notification)
        
        logger.info(f"⚠️ İade itirazı - Admin'lere bildirildi. Return ID: {return_id}")
        
        return {"message": "İade talebine itiraz edildi. Konu yöneticilere iletildi."}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Reject return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/return-requests/{return_id}/ship")
async def ship_return(return_id: str, current_user: dict = Depends(get_current_user)):
    """İade ürününü kargoya ver (Alıcı tarafından)"""
    try:
        current_user_id = current_user["id"]
        
        return_request = await db.marketplace_returns.find_one({
            "id": return_id,
            "buyer_id": current_user_id
        })
        
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        if return_request.get("status") != "approved":
            raise HTTPException(status_code=400, detail="Bu iade henüz onaylanmamış veya zaten kargoda")
        
        # İade talebini güncelle
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {"status": "return_shipped", "return_shipped_at": datetime.utcnow()}}
        )
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {"status": "return_shipped"}}
        )
        
        # Satıcıya bildirim
        tracking_code = return_request.get("return_tracking_code", "")
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("seller_id"),
            "type": "return_shipped",
            "title": "İade Kargoda 📦",
            "message": f"'{listing.get('title', 'Ürün')}' iade kargosu yola çıktı.\n🔢 Takip Kodu: {tracking_code}",
            "data": {
                "return_id": return_id,
                "tracking_code": tracking_code
            },
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ İade kargoya verildi - Return ID: {return_id}")
        
        return {"message": "İade ürünü kargoya verildi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ship return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/return-requests/{return_id}/receive")
async def receive_return(return_id: str, current_user: dict = Depends(get_current_user)):
    """İade ürününü teslim al (Satıcı tarafından)"""
    try:
        current_user_id = current_user["id"]
        
        return_request = await db.marketplace_returns.find_one({
            "id": return_id,
            "seller_id": current_user_id
        })
        
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        if return_request.get("status") != "return_shipped":
            raise HTTPException(status_code=400, detail="Bu iade henüz kargoya verilmemiş")
        
        # İade talebini güncelle
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {"status": "returned", "returned_at": datetime.utcnow()}}
        )
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {"status": "returned"}}
        )
        
        # Alıcıya bildirim - İade tamamlandı
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("buyer_id"),
            "type": "return_completed",
            "title": "İade Tamamlandı ✅",
            "message": f"'{listing.get('title', 'Ürün')}' iade işlemi tamamlandı. Ödemeniz iade edilecektir.",
            "data": {"return_id": return_id},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ İade tamamlandı - Return ID: {return_id}")
        
        return {"message": "İade ürünü teslim alındı, iade tamamlandı"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Receive return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Kullanıcı adres endpoint'leri
@router.get("/user/addresses")
async def get_user_addresses(current_user: dict = Depends(get_current_user)):
    """Kullanıcının kayıtlı adresleri"""
    try:
        current_user_id = current_user["id"]
        
        user = await db.users.find_one({"id": current_user_id})
        if not user:
            return []
        
        return user.get("saved_addresses", [])
    except Exception as e:
        logger.error(f"❌ Get addresses error: {str(e)}")
        return []


@router.post("/user/addresses")
async def save_user_address(address: dict, current_user: dict = Depends(get_current_user)):
    """Adres kaydet"""
    try:
        current_user_id = current_user["id"]
        
        # Adrese ID ekle
        address["id"] = str(uuid.uuid4())
        address["created_at"] = datetime.utcnow().isoformat()
        
        # Kullanıcının adreslerine ekle
        await db.users.update_one(
            {"id": current_user_id},
            {"$push": {"saved_addresses": address}}
        )
        
        return {"message": "Adres kaydedildi", "address_id": address["id"]}
    except Exception as e:
        logger.error(f"❌ Save address error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# KARGO TAKİP SİSTEMİ (Satıcı için Sipariş Takibi)
# ============================================

# Türkiye'deki popüler kargo şirketleri
CARGO_COMPANIES = [
    {"id": "aras", "name": "Aras Kargo", "tracking_url": "https://www.araskargo.com.tr/trmTakip.aspx?q="},
    {"id": "yurtici", "name": "Yurtiçi Kargo", "tracking_url": "https://www.yurticikargo.com/tr/online-servisler/gonderi-sorgula?code="},
    {"id": "mng", "name": "MNG Kargo", "tracking_url": "https://www.mngkargo.com.tr/gonderi-takip/?q="},
    {"id": "ptt", "name": "PTT Kargo", "tracking_url": "https://gonderitakip.ptt.gov.tr/Track/Verify?q="},
    {"id": "surat", "name": "Sürat Kargo", "tracking_url": "https://www.suratkargo.com.tr/gonderi-takip?barcode="},
    {"id": "ups", "name": "UPS", "tracking_url": "https://www.ups.com/track?tracknum="},
    {"id": "fedex", "name": "FedEx", "tracking_url": "https://www.fedex.com/fedextrack/?trknbr="},
    {"id": "dhl", "name": "DHL", "tracking_url": "https://www.dhl.com/tr-tr/home/tracking.html?tracking-id="},
    {"id": "other", "name": "Diğer", "tracking_url": ""}
]

class ShippingInfo(BaseModel):
    cargo_company: str
    tracking_code: str
    notes: Optional[str] = None

@router.get("/cargo-companies")
async def get_cargo_companies():
    """Kargo şirketleri listesi"""
    return {"companies": CARGO_COMPANIES}

@router.get("/seller/orders")
async def get_seller_orders(current_user: dict = Depends(get_current_user)):
    """Satıcının siparişleri (sipariş takibi için)"""
    try:
        current_user_id = current_user["id"]
        
        # Satıcının tüm satışlarını getir
        transactions = await db.marketplace_transactions.find({
            "seller_id": current_user_id,
            "status": {"$in": ["completed", "pending", "confirmed", "shipped", "delivered", 
                               "return_requested", "return_approved", "return_shipped", "return_completed", "return_rejected"]}
        }).sort("created_at", -1).to_list(100)
        
        orders = []
        for tx in transactions:
            listing = await db.marketplace_listings.find_one({"id": tx["listing_id"]})
            buyer = await db.users.find_one({"id": tx["buyer_id"]})
            
            if listing:
                orders.append({
                    "id": tx.get("id"),
                    "transaction_id": tx.get("id"),
                    "listing_id": tx.get("listing_id"),
                    "listing_title": listing.get("title"),
                    "listing_image": listing.get("images", [None])[0],
                    "buyer_id": tx.get("buyer_id"),
                    "buyer_name": buyer.get("full_name") if buyer else "Bilinmeyen",
                    "buyer_phone": buyer.get("phone") if buyer else None,
                    "price": tx.get("amount") or listing.get("price"),
                    "shipping_address": tx.get("shipping_address"),
                    "status": tx.get("status", "pending"),
                    "created_at": tx.get("created_at"),
                    "cargo_company": tx.get("cargo_company"),
                    "tracking_code": tx.get("tracking_code"),
                    "shipped_at": tx.get("shipped_at"),
                    "delivered_at": tx.get("delivered_at"),
                    "confirmed_at": tx.get("confirmed_at"),
                    "shipping_notes": tx.get("shipping_notes")
                })
        
        return orders
    except Exception as e:
        logger.error(f"❌ Get seller orders error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/seller/orders/{order_id}/ship")
async def ship_order(order_id: str, shipping_info: ShippingInfo, current_user: dict = Depends(get_current_user)):
    """Siparişi kargoya ver (kargo bilgisi gir)"""
    try:
        current_user_id = current_user["id"]
        
        # Siparişi bul
        transaction = await db.marketplace_transactions.find_one({
            "id": order_id,
            "seller_id": current_user_id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        if transaction.get("status") not in ["completed", "pending", "confirmed"]:
            raise HTTPException(status_code=400, detail="Bu sipariş kargoya verilemez")
        
        # Kargo şirketini bul
        cargo_company_info = next((c for c in CARGO_COMPANIES if c["id"] == shipping_info.cargo_company), None)
        cargo_company_name = cargo_company_info["name"] if cargo_company_info else shipping_info.cargo_company
        
        # Siparişi güncelle
        await db.marketplace_transactions.update_one(
            {"id": order_id},
            {"$set": {
                "status": "shipped",
                "cargo_company": shipping_info.cargo_company,
                "cargo_company_name": cargo_company_name,
                "tracking_code": shipping_info.tracking_code,
                "shipping_notes": shipping_info.notes,
                "shipped_at": datetime.utcnow()
            }}
        )
        
        # Alıcıya kargo bildirimi gönder
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        tracking_url = cargo_company_info["tracking_url"] + shipping_info.tracking_code if cargo_company_info else ""
        
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": transaction.get("buyer_id"),
            "type": "order_shipped",
            "title": "Kargonuz Yola Çıktı! 📦",
            "message": f"'{listing.get('title', 'Ürün')}' kargoya verildi.\nKargo: {cargo_company_name}\nTakip No: {shipping_info.tracking_code}",
            "data": {
                "order_id": order_id,
                "cargo_company": shipping_info.cargo_company,
                "cargo_company_name": cargo_company_name,
                "tracking_code": shipping_info.tracking_code,
                "tracking_url": tracking_url
            },
            "action_url": f"/marketplace/my-orders",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Order {order_id} shipped with tracking {shipping_info.tracking_code}")
        
        # Kargo bilgisi ekleme log'u
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "marketplace_ship", "success", {
                "order_id": order_id,
                "listing_title": listing.get("title") if listing else "",
                "cargo_company": cargo_company_name,
                "tracking_code": shipping_info.tracking_code,
                "buyer_id": transaction.get("buyer_id")
            })
        except Exception as log_err:
            logger.error(f"Log error: {log_err}")
        
        return {"message": "Kargo bilgisi kaydedildi ve alıcıya bildirim gönderildi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ship order error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/orders/{order_id}/confirm-delivery")
async def confirm_delivery(order_id: str, current_user: dict = Depends(get_current_user)):
    """Alıcı kargoyu teslim aldığını onaylar"""
    try:
        current_user_id = current_user["id"]
        
        # Siparişi bul
        transaction = await db.marketplace_transactions.find_one({
            "id": order_id,
            "buyer_id": current_user_id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        if transaction.get("status") != "shipped":
            raise HTTPException(status_code=400, detail="Bu sipariş henüz kargoya verilmemiş veya zaten teslim edilmiş")
        
        # Siparişi teslim edildi olarak güncelle
        delivered_at = datetime.utcnow()
        auto_approve_deadline = delivered_at + timedelta(days=1)  # 1 gün sonra otomatik onay
        
        await db.marketplace_transactions.update_one(
            {"id": order_id},
            {"$set": {
                "status": "delivered",
                "delivered_at": delivered_at,
                "auto_approve_deadline": auto_approve_deadline
            }}
        )
        
        # Listing bilgisini al
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        listing_title = listing.get("title", "Ürün") if listing else "Ürün"
        
        # Alıcıya "1 gün içinde onayla veya iade et" bildirimi
        buyer_notification = {
            "id": str(uuid.uuid4()),
            "user_id": current_user_id,
            "type": "delivery_confirmation_required",
            "title": "📦 Ürün Teslim Alındı!",
            "message": f"'{listing_title}' ürününüzü aldınız. 1 gün içinde onaylayın veya iade talep edin. Aksi halde otomatik onaylanacaktır.",
            "related_id": order_id,
            "related_type": "marketplace_order",
            "data": {
                "order_id": order_id,
                "listing_id": transaction.get("listing_id"),
                "listing_title": listing_title,
                "auto_approve_deadline": auto_approve_deadline.isoformat(),
                "action_required": True,
                "actions": ["approve", "return"]
            },
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(buyer_notification)
        
        logger.info(f"✅ Order {order_id} delivery confirmed by buyer, auto-approve deadline: {auto_approve_deadline}")
        return {
            "message": "Teslimat onaylandı. 1 gün içinde ürünü onaylayın veya iade talep edin.",
            "auto_approve_deadline": auto_approve_deadline.isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Confirm delivery error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/orders/{order_id}/approve")
async def approve_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """Alıcı ürünü onaylar (satıcıya ve yöneticiye bildirim + değerlendirme)"""
    try:
        current_user_id = current_user["id"]
        
        # Siparişi bul
        transaction = await db.marketplace_transactions.find_one({
            "id": order_id,
            "buyer_id": current_user_id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        
        if transaction.get("status") != "delivered":
            raise HTTPException(status_code=400, detail="Sadece teslim edilmiş siparişler onaylanabilir")
        
        # Siparişi onaylı olarak güncelle
        await db.marketplace_transactions.update_one(
            {"id": order_id},
            {"$set": {
                "status": "approved",
                "approved_at": datetime.utcnow(),
                "can_review": True
            }}
        )
        
        # Listing ve kullanıcı bilgilerini al
        listing = await db.marketplace_listings.find_one({"id": transaction.get("listing_id")})
        buyer = await db.users.find_one({"id": current_user_id})
        buyer_name = buyer.get("full_name", "Alıcı") if buyer else "Alıcı"
        
        # Satıcıya bildirim gönder
        seller_notification = {
            "id": str(uuid.uuid4()),
            "user_id": transaction.get("seller_id"),
            "type": "order_approved",
            "title": "Sipariş Onaylandı! ✅",
            "message": f"'{listing.get('title', 'Ürün')}' siparişi {buyer_name} tarafından onaylandı.",
            "data": {
                "order_id": order_id,
                "listing_id": transaction.get("listing_id")
            },
            "action_url": f"/marketplace/seller/orders",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(seller_notification)
        
        # Yöneticilere bildirim gönder
        admins = await db.users.find({"user_type": {"$in": ["admin", "super_admin"]}}).to_list(100)
        for admin in admins:
            admin_notification = {
                "id": str(uuid.uuid4()),
                "user_id": admin["id"],
                "type": "order_approved_admin",
                "title": "Sipariş Onaylandı",
                "message": f"'{listing.get('title', 'Ürün')}' siparişi alıcı tarafından onaylandı.",
                "data": {
                    "order_id": order_id,
                    "listing_id": transaction.get("listing_id"),
                    "seller_id": transaction.get("seller_id"),
                    "buyer_id": current_user_id
                },
                "action_url": f"/marketplace/admin-panel",
                "read": False,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(admin_notification)
        
        logger.info(f"✅ Order {order_id} approved by buyer {current_user_id}")
        return {
            "message": "Sipariş onaylandı! Satıcıyı değerlendirebilirsiniz.",
            "can_review": True,
            "seller_id": transaction.get("seller_id"),
            "listing_id": transaction.get("listing_id")
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Approve order error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# DEĞERLENDİRME (PUAN VE YORUM) SİSTEMİ
# ============================================

class ReviewCreate(BaseModel):
    rating: int  # 1-5 arası
    comment: str  # max 500 karakter
    
class ReviewResponse(BaseModel):
    id: str
    reviewer_id: str
    reviewer_name: str
    reviewer_avatar: Optional[str]
    target_id: str
    target_type: str  # seller, coach, player, facility, referee
    rating: int
    comment: str
    order_id: Optional[str]
    reservation_id: Optional[str]
    created_at: datetime

@router.post("/reviews/seller/{seller_id}")
async def create_seller_review(
    seller_id: str, 
    review_data: ReviewCreate, 
    order_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Satıcıya puan ve yorum ver"""
    try:
        current_user_id = current_user["id"]
        
        # Validasyon
        if review_data.rating < 1 or review_data.rating > 5:
            raise HTTPException(status_code=400, detail="Puan 1-5 arası olmalıdır")
        
        if len(review_data.comment) > 500:
            raise HTTPException(status_code=400, detail="Yorum en fazla 500 karakter olabilir")
        
        # Eğer order_id verilmişse, bu siparişin alıcısı olmalı
        if order_id:
            transaction = await db.marketplace_transactions.find_one({
                "id": order_id,
                "buyer_id": current_user_id,
                "seller_id": seller_id
            })
            if not transaction:
                raise HTTPException(status_code=400, detail="Bu siparişi değerlendirme yetkiniz yok")
            
            # Daha önce değerlendirme yapılmış mı kontrol et
            existing_review = await db.reviews.find_one({
                "reviewer_id": current_user_id,
                "order_id": order_id
            })
            if existing_review:
                raise HTTPException(status_code=400, detail="Bu siparişi zaten değerlendirdiniz")
        
        # Kullanıcı bilgilerini al
        reviewer = await db.users.find_one({"id": current_user_id})
        seller = await db.users.find_one({"id": seller_id})
        
        if not seller:
            raise HTTPException(status_code=404, detail="Satıcı bulunamadı")
        
        # Review oluştur
        review = {
            "id": str(uuid.uuid4()),
            "reviewer_id": current_user_id,
            "reviewer_name": reviewer.get("full_name", "Anonim") if reviewer else "Anonim",
            "reviewer_avatar": reviewer.get("profile_image") if reviewer else None,
            "target_id": seller_id,
            "target_type": "seller",
            "rating": review_data.rating,
            "comment": review_data.comment,
            "order_id": order_id,
            "created_at": datetime.utcnow()
        }
        
        await db.reviews.insert_one(review)
        
        # Satıcının ortalama puanını güncelle
        all_reviews = await db.reviews.find({"target_id": seller_id, "target_type": "seller"}).to_list(1000)
        avg_rating = sum(r["rating"] for r in all_reviews) / len(all_reviews) if all_reviews else 0
        review_count = len(all_reviews)
        
        await db.users.update_one(
            {"id": seller_id},
            {"$set": {"seller_rating": round(avg_rating, 1), "seller_review_count": review_count}}
        )
        
        # Siparişi değerlendirildi olarak işaretle
        if order_id:
            await db.marketplace_transactions.update_one(
                {"id": order_id},
                {"$set": {"reviewed": True, "review_id": review["id"]}}
            )
        
        # Satıcıya bildirim gönder
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": seller_id,
            "type": "new_review",
            "title": "Yeni Değerlendirme ⭐",
            "message": f"{reviewer.get('full_name', 'Bir kullanıcı')} size {review_data.rating} yıldız verdi.",
            "data": {"review_id": review["id"], "rating": review_data.rating},
            "action_url": f"/profile/{seller_id}/reviews",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Review created for seller {seller_id} by {current_user_id}")
        return {"message": "Değerlendirmeniz kaydedildi", "review_id": review["id"]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Create seller review error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/reviews/{target_type}/{target_id}")
async def create_review(
    target_type: str,  # coach, player, facility, referee
    target_id: str, 
    review_data: ReviewCreate,
    reservation_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Antrenör, oyuncu, tesis veya hakeme puan ve yorum ver"""
    try:
        current_user_id = current_user["id"]
        
        # Target type validasyonu
        valid_types = ["coach", "player", "facility", "referee"]
        if target_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Geçersiz hedef tipi. Geçerli tipler: {valid_types}")
        
        # Validasyon
        if review_data.rating < 1 or review_data.rating > 5:
            raise HTTPException(status_code=400, detail="Puan 1-5 arası olmalıdır")
        
        if len(review_data.comment) > 500:
            raise HTTPException(status_code=400, detail="Yorum en fazla 500 karakter olabilir")
        
        # Hedef kullanıcıyı/tesisi bul
        if target_type == "facility":
            target = await db.facilities.find_one({"id": target_id})
            target_name = target.get("name") if target else "Tesis"
        else:
            target = await db.users.find_one({"id": target_id})
            target_name = target.get("full_name") if target else "Kullanıcı"
        
        if not target:
            raise HTTPException(status_code=404, detail="Hedef bulunamadı")
        
        # Rezervasyon kontrolü (opsiyonel)
        if reservation_id:
            # Daha önce değerlendirme yapılmış mı kontrol et
            existing_review = await db.reviews.find_one({
                "reviewer_id": current_user_id,
                "reservation_id": reservation_id
            })
            if existing_review:
                raise HTTPException(status_code=400, detail="Bu rezervasyonu zaten değerlendirdiniz")
        
        # Kullanıcı bilgilerini al
        reviewer = await db.users.find_one({"id": current_user_id})
        
        # Review oluştur
        review = {
            "id": str(uuid.uuid4()),
            "reviewer_id": current_user_id,
            "reviewer_name": reviewer.get("full_name", "Anonim") if reviewer else "Anonim",
            "reviewer_avatar": reviewer.get("profile_image") if reviewer else None,
            "target_id": target_id,
            "target_type": target_type,
            "rating": review_data.rating,
            "comment": review_data.comment,
            "reservation_id": reservation_id,
            "created_at": datetime.utcnow()
        }
        
        await db.reviews.insert_one(review)
        
        # Hedefin ortalama puanını güncelle
        all_reviews = await db.reviews.find({"target_id": target_id, "target_type": target_type}).to_list(1000)
        avg_rating = sum(r["rating"] for r in all_reviews) / len(all_reviews) if all_reviews else 0
        review_count = len(all_reviews)
        
        rating_field = f"{target_type}_rating" if target_type != "facility" else "rating"
        count_field = f"{target_type}_review_count" if target_type != "facility" else "review_count"
        
        if target_type == "facility":
            await db.facilities.update_one(
                {"id": target_id},
                {"$set": {"rating": round(avg_rating, 1), "review_count": review_count}}
            )
        else:
            await db.users.update_one(
                {"id": target_id},
                {"$set": {rating_field: round(avg_rating, 1), count_field: review_count}}
            )
        
        # Hedefe bildirim gönder
        notification_user_id = target.get("owner_id") if target_type == "facility" else target_id
        if notification_user_id:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": notification_user_id,
                "type": "new_review",
                "title": "Yeni Değerlendirme ⭐",
                "message": f"{reviewer.get('full_name', 'Bir kullanıcı')} size {review_data.rating} yıldız verdi.",
                "data": {
                    "review_id": review["id"], 
                    "rating": review_data.rating,
                    "target_type": target_type
                },
                "action_url": f"/reviews/{target_type}/{target_id}",
                "read": False,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Review created for {target_type} {target_id} by {current_user_id}")
        return {"message": "Değerlendirmeniz kaydedildi", "review_id": review["id"]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Create review error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/reviews/{target_type}/{target_id}")
async def get_reviews(
    target_type: str,
    target_id: str,
    skip: int = 0,
    limit: int = 20
):
    """Bir hedefin tüm değerlendirmelerini getir (herkese açık)"""
    try:
        reviews = await db.reviews.find({
            "target_id": target_id,
            "target_type": target_type
        }).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        
        total = await db.reviews.count_documents({
            "target_id": target_id,
            "target_type": target_type
        })
        
        # Ortalama puanı hesapla
        all_reviews = await db.reviews.find({
            "target_id": target_id,
            "target_type": target_type
        }).to_list(1000)
        
        avg_rating = sum(r["rating"] for r in all_reviews) / len(all_reviews) if all_reviews else 0
        
        # _id'leri kaldır
        for review in reviews:
            review.pop("_id", None)
        
        return {
            "reviews": reviews,
            "total": total,
            "average_rating": round(avg_rating, 1),
            "review_count": len(all_reviews)
        }
    except Exception as e:
        logger.error(f"❌ Get reviews error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/reviews/pending")
async def get_pending_reviews(current_user: dict = Depends(get_current_user)):
    """Kullanıcının değerlendirme bekleyen siparişleri ve rezervasyonları"""
    try:
        current_user_id = current_user["id"]
        pending_reviews = []
        
        # Onaylanmış ama değerlendirilmemiş siparişler
        orders = await db.marketplace_transactions.find({
            "buyer_id": current_user_id,
            "status": "approved",
            "reviewed": {"$ne": True}
        }).to_list(100)
        
        for order in orders:
            listing = await db.marketplace_listings.find_one({"id": order.get("listing_id")})
            seller = await db.users.find_one({"id": order.get("seller_id")})
            pending_reviews.append({
                "type": "order",
                "id": order.get("id"),
                "target_id": order.get("seller_id"),
                "target_type": "seller",
                "target_name": seller.get("full_name") if seller else "Satıcı",
                "target_avatar": seller.get("profile_image") if seller else None,
                "title": listing.get("title") if listing else "Ürün",
                "image": listing.get("images", [None])[0] if listing else None,
                "created_at": order.get("approved_at") or order.get("created_at")
            })
        
        # Bitmiş rezervasyonlar (tesis, antrenör, oyuncu, hakem)
        # Bu kısım mevcut rezervasyon yapısına göre genişletilebilir
        
        return {"pending_reviews": pending_reviews}
    except Exception as e:
        logger.error(f"❌ Get pending reviews error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# MY PRODUCTS - Kullanıcının Ürünleri
# ============================================

@router.get("/my-products")
async def get_my_products(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Kullanıcının kendi ürünlerini getir"""
    try:
        current_user_id = current_user["id"]
        
        logger.info(f"📦 /my-products called by user_id: {current_user_id}, status filter: {status}")
        
        query = {"seller_id": current_user_id}
        
        # Status filter
        if status and status != "all":
            if status == "active":
                query["status"] = "active"
            elif status == "paused":
                query["status"] = "paused"
            elif status == "in_progress":
                # Satış süreci devam eden ürünler
                query["status"] = {"$in": ["pending_payment", "sold", "shipped"]}
            elif status == "completed":
                query["status"] = {"$in": ["delivered", "completed"]}
            elif status == "archived":
                query["status"] = "archived"
        
        logger.info(f"📦 Query: {query}")
        
        listings = await db.marketplace_listings.find(query).sort("created_at", -1).to_list(500)
        
        logger.info(f"📦 Found {len(listings)} listings")
        
        # Her ürün için ek bilgileri getir
        products = []
        for listing in listings:
            listing.pop("_id", None)
            
            # Satış bilgilerini getir
            transaction = await db.marketplace_transactions.find_one({
                "listing_id": listing["id"],
                "status": {"$nin": ["cancelled", "refunded"]}
            })
            
            # Favori sayısını getir
            favorites_count = await db.marketplace_favorites.count_documents({
                "listing_id": listing["id"]
            })
            
            products.append({
                **listing,
                "favorites_count": favorites_count,
                "has_active_transaction": transaction is not None,
                "transaction_status": transaction.get("status") if transaction else None
            })
        
        return {"products": products, "total": len(products)}
    except Exception as e:
        logger.error(f"❌ Get my products error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class UpdateListingRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    price: Optional[float] = None
    original_price: Optional[float] = None
    stock: Optional[int] = None
    condition: Optional[str] = None
    brand: Optional[str] = None
    images: Optional[List[str]] = None


@router.put("/listings/{listing_id}")
async def update_listing(
    listing_id: str,
    data: UpdateListingRequest,
    current_user: dict = Depends(get_current_user)
):
    """Ürün bilgilerini güncelle - Fiyat düşüşünde favorilere bildirim gönder"""
    try:
        current_user_id = current_user["id"]
        
        # Ürünü bul
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Yetki kontrolü
        if listing["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Bu ürünü düzenleme yetkiniz yok")
        
        # Fiyat düşüşü kontrolü
        old_price = listing.get("price", 0)
        new_price = data.price if data.price is not None else old_price
        price_dropped = new_price < old_price
        
        # Güncellenecek alanları hazırla
        update_data = {"updated_at": datetime.utcnow()}
        
        if data.title is not None:
            update_data["title"] = data.title
        if data.description is not None:
            update_data["description"] = data.description
        if data.price is not None:
            update_data["price"] = data.price
            # Eski fiyatı kaydet (indirim gösterimi için)
            if price_dropped:
                update_data["original_price"] = old_price
        if data.original_price is not None:
            update_data["original_price"] = data.original_price
        if data.stock is not None:
            update_data["stock"] = data.stock
        if data.condition is not None:
            update_data["condition"] = data.condition
        if data.brand is not None:
            update_data["brand"] = data.brand
        if data.images is not None:
            update_data["images"] = data.images
        
        # Güncelle
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$set": update_data}
        )
        
        # Fiyat düşüşü bildirimi gönder
        if price_dropped:
            await send_price_drop_notifications(
                listing_id=listing_id,
                listing_title=data.title or listing.get("title"),
                old_price=old_price,
                new_price=new_price,
                seller_id=current_user_id
            )
        
        # Güncellenmiş ürünü getir
        updated_listing = await db.marketplace_listings.find_one({"id": listing_id})
        updated_listing.pop("_id", None)
        
        return {
            "success": True,
            "message": "Ürün güncellendi",
            "listing": updated_listing,
            "price_drop_notification_sent": price_dropped
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Update listing error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


async def send_price_drop_notifications(
    listing_id: str,
    listing_title: str,
    old_price: float,
    new_price: float,
    seller_id: str
):
    """Fiyat düşüşünde favorilere ekleyenlere bildirim gönder"""
    try:
        # Favorilere ekleyen kullanıcıları bul
        favorites = await db.marketplace_favorites.find({
            "listing_id": listing_id
        }).to_list(1000)
        
        if not favorites:
            logger.info(f"📢 No favorites found for listing {listing_id}")
            return
        
        logger.info(f"📢 Sending price drop notification to {len(favorites)} users")
        
        # Her favoriye ekleyen kullanıcıya bildirim gönder
        for fav in favorites:
            user_id = fav.get("user_id")
            if user_id and user_id != seller_id:  # Satıcıya gönderme
                notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": user_id,
                    "type": "price_drop",
                    "title": "🏷️ Fiyat Düştü!",
                    "message": f"{listing_title}: ₺{old_price:.0f} → ₺{new_price:.0f}",
                    "related_id": listing_id,
                    "related_type": "marketplace_listing",
                    "data": {
                        "listing_id": listing_id,
                        "listing_title": listing_title,
                        "old_price": old_price,
                        "new_price": new_price,
                        "discount_percent": round((1 - new_price / old_price) * 100)
                    },
                    "read": False,
                    "created_at": datetime.utcnow()
                }
                await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Price drop notifications sent for listing {listing_id}")
    except Exception as e:
        logger.error(f"❌ Error sending price drop notifications: {str(e)}")


class UpdateListingStatusRequest(BaseModel):
    status: str  # active, paused, archived


@router.put("/listings/{listing_id}/status")
async def update_listing_status(
    listing_id: str,
    data: UpdateListingStatusRequest,
    current_user: dict = Depends(get_current_user)
):
    """Ürün durumunu değiştir (yayından kaldır, arşivle, tekrar yayına al)"""
    try:
        current_user_id = current_user["id"]
        
        # Ürünü bul
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Yetki kontrolü
        if listing["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Bu ürünü düzenleme yetkiniz yok")
        
        # Geçerli status kontrolü
        valid_statuses = ["active", "paused", "archived"]
        if data.status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Geçersiz durum. Geçerli değerler: {valid_statuses}")
        
        # Aktif işlem varsa arşivlemeye izin verme
        if data.status == "archived":
            active_transaction = await db.marketplace_transactions.find_one({
                "listing_id": listing_id,
                "status": {"$nin": ["cancelled", "refunded", "completed", "delivered"]}
            })
            if active_transaction:
                raise HTTPException(status_code=400, detail="Aktif satış işlemi olan ürün arşivlenemez")
        
        # Güncelle
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$set": {
                "status": data.status,
                "updated_at": datetime.utcnow()
            }}
        )
        
        status_messages = {
            "active": "Ürün yayına alındı",
            "paused": "Ürün yayından kaldırıldı",
            "archived": "Ürün arşivlendi"
        }
        
        return {
            "success": True,
            "message": status_messages.get(data.status, "Durum güncellendi"),
            "new_status": data.status
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Update listing status error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/listings/{listing_id}/restore")
async def restore_listing(
    listing_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Arşivlenmiş ürünü geri yükle"""
    try:
        current_user_id = current_user["id"]
        
        # Ürünü bul
        listing = await db.marketplace_listings.find_one({"id": listing_id})
        if not listing:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Yetki kontrolü
        if listing["seller_id"] != current_user_id:
            raise HTTPException(status_code=403, detail="Bu ürünü düzenleme yetkiniz yok")
        
        if listing.get("status") != "archived":
            raise HTTPException(status_code=400, detail="Sadece arşivlenmiş ürünler geri yüklenebilir")
        
        # Askıda olarak geri yükle (kullanıcı isterse yayına alabilir)
        await db.marketplace_listings.update_one(
            {"id": listing_id},
            {"$set": {
                "status": "paused",
                "updated_at": datetime.utcnow()
            }}
        )
        
        return {
            "success": True,
            "message": "Ürün arşivden çıkarıldı (askıda)",
            "new_status": "paused"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Restore listing error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ==================== ADMIN ÖDEME RAPORU ====================

@router.get("/admin/seller-payments-report")
async def get_seller_payments_report(
    period: str = "weekly",  # weekly, monthly, custom
    year: int = None,        # Yıl filtresi
    month: int = None,       # Ay filtresi (1-12)
    week: int = None,        # Hafta filtresi (yılın kaçıncı haftası)
    current_user: dict = Depends(get_current_user)
):
    """
    Kullanıcı bazlı ödeme raporu - Admin için
    Tüm ödeme türleri: Etkinlik, Spor Market, Rezervasyon
    """
    try:
        current_user_id = current_user["id"]
        
        # Admin kontrolü
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        now = datetime.utcnow()
        
        # Varsayılan değerler
        if year is None:
            year = now.year
        if month is None and period == "monthly":
            month = now.month
        
        # Tarih aralığını hesapla
        if period == "weekly":
            if week is not None:
                # Belirli bir hafta
                first_day_of_year = datetime(year, 1, 1)
                # Yılın ilk pazartesisini bul
                days_to_monday = (7 - first_day_of_year.weekday()) % 7
                first_monday = first_day_of_year + timedelta(days=days_to_monday)
                period_start = first_monday + timedelta(weeks=week-1)
                period_end = period_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
            else:
                # Bu hafta (Pazartesi'den başla)
                days_since_monday = now.weekday()
                period_start = (now - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
                period_end = period_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
        else:  # monthly
            # Belirli ay
            period_start = datetime(year, month, 1, 0, 0, 0)
            # Ayın son günü
            if month == 12:
                period_end = datetime(year + 1, 1, 1) - timedelta(seconds=1)
            else:
                period_end = datetime(year, month + 1, 1) - timedelta(seconds=1)
        
        logger.info(f"📊 Payment report: {period_start} - {period_end}")
        
        # Kullanıcı bazlı veri toplama
        user_data = {}
        
        # 1. MARKETPLACE İŞLEMLERİ - Satıcı kazançları
        marketplace_query = {
            "status": {"$in": ["completed", "delivered", "approved"]},
            "created_at": {"$gte": period_start, "$lte": period_end}
        }
        marketplace_txs = await db.marketplace_transactions.find(marketplace_query).to_list(10000)
        
        for tx in marketplace_txs:
            seller_id = tx.get("seller_id")
            if not seller_id:
                continue
            
            if seller_id not in user_data:
                user_data[seller_id] = {
                    "user_id": seller_id,
                    "marketplace_sales": 0,
                    "marketplace_commission": 0,
                    "marketplace_net": 0,
                    "marketplace_count": 0,
                    "event_earnings": 0,
                    "event_commission": 0,
                    "event_net": 0,
                    "event_count": 0,
                    "reservation_earnings": 0,
                    "reservation_commission": 0,
                    "reservation_net": 0,
                    "reservation_count": 0,
                    "total_gross": 0,
                    "total_commission": 0,
                    "total_net": 0,
                }
            
            user_data[seller_id]["marketplace_sales"] += tx.get("total_paid_by_buyer", 0)
            user_data[seller_id]["marketplace_commission"] += tx.get("commission_amount", 0)
            user_data[seller_id]["marketplace_net"] += tx.get("seller_receives", 0)
            user_data[seller_id]["marketplace_count"] += 1
        
        # 2. ETKİNLİK ÖDEMELERİ - Organizatör kazançları
        event_payments_query = {
            "payment_status": "completed",
            "created_at": {"$gte": period_start, "$lte": period_end}
        }
        participations = await db.participations.find(event_payments_query).to_list(10000)
        
        for p in participations:
            # Etkinliği bul ve organizatörü al
            event = await db.events.find_one({"id": p.get("event_id")})
            if not event:
                continue
            
            organizer_id = event.get("organizer_id")
            if not organizer_id:
                continue
            
            if organizer_id not in user_data:
                user_data[organizer_id] = {
                    "user_id": organizer_id,
                    "marketplace_sales": 0,
                    "marketplace_commission": 0,
                    "marketplace_net": 0,
                    "marketplace_count": 0,
                    "event_earnings": 0,
                    "event_commission": 0,
                    "event_net": 0,
                    "event_count": 0,
                    "reservation_earnings": 0,
                    "reservation_commission": 0,
                    "reservation_net": 0,
                    "reservation_count": 0,
                    "total_gross": 0,
                    "total_commission": 0,
                    "total_net": 0,
                }
            
            amount = p.get("amount_paid", 0) or p.get("price", 0)
            commission = amount * 0.10  # %10 komisyon varsayılan
            net = amount - commission
            
            user_data[organizer_id]["event_earnings"] += amount
            user_data[organizer_id]["event_commission"] += commission
            user_data[organizer_id]["event_net"] += net
            user_data[organizer_id]["event_count"] += 1
        
        # 3. REZERVASYON ÖDEMELERİ - Tesis sahibi kazançları
        reservation_query = {
            "payment_status": {"$in": ["completed", "paid"]},
            "created_at": {"$gte": period_start, "$lte": period_end}
        }
        reservations = await db.reservations.find(reservation_query).to_list(10000)
        
        for r in reservations:
            facility_id = r.get("facility_id")
            if not facility_id:
                continue
            
            # Tesis sahibini bul
            facility = await db.facilities.find_one({"id": facility_id})
            if not facility:
                continue
            
            owner_id = facility.get("owner_id")
            if not owner_id:
                continue
            
            if owner_id not in user_data:
                user_data[owner_id] = {
                    "user_id": owner_id,
                    "marketplace_sales": 0,
                    "marketplace_commission": 0,
                    "marketplace_net": 0,
                    "marketplace_count": 0,
                    "event_earnings": 0,
                    "event_commission": 0,
                    "event_net": 0,
                    "event_count": 0,
                    "reservation_earnings": 0,
                    "reservation_commission": 0,
                    "reservation_net": 0,
                    "reservation_count": 0,
                    "total_gross": 0,
                    "total_commission": 0,
                    "total_net": 0,
                }
            
            amount = r.get("total_price", 0) or r.get("price", 0)
            commission = amount * 0.10  # %10 komisyon varsayılan
            net = amount - commission
            
            user_data[owner_id]["reservation_earnings"] += amount
            user_data[owner_id]["reservation_commission"] += commission
            user_data[owner_id]["reservation_net"] += net
            user_data[owner_id]["reservation_count"] += 1
        
        # Toplamları hesapla ve kullanıcı bilgilerini ekle
        result = []
        for user_id, data in user_data.items():
            # Toplamları hesapla
            data["total_gross"] = data["marketplace_sales"] + data["event_earnings"] + data["reservation_earnings"]
            data["total_commission"] = data["marketplace_commission"] + data["event_commission"] + data["reservation_commission"]
            data["total_net"] = data["marketplace_net"] + data["event_net"] + data["reservation_net"]
            
            # Kullanıcı bilgilerini al
            user_info = await db.users.find_one({"id": user_id})
            
            # Ödeme durumunu kontrol et
            payment_record = await db.seller_payments.find_one({
                "seller_id": user_id,
                "period_start": period_start,
                "period_end": period_end
            })
            
            result.append({
                "user_id": user_id,
                "user_name": user_info.get("full_name", "Bilinmeyen") if user_info else "Bilinmeyen",
                "iban": user_info.get("iban", "") if user_info else "",
                "phone": user_info.get("phone", "") if user_info else "",
                "user_type": user_info.get("user_type", "") if user_info else "",
                # Marketplace
                "marketplace_sales": round(data["marketplace_sales"], 2),
                "marketplace_commission": round(data["marketplace_commission"], 2),
                "marketplace_net": round(data["marketplace_net"], 2),
                "marketplace_count": data["marketplace_count"],
                # Etkinlik
                "event_earnings": round(data["event_earnings"], 2),
                "event_commission": round(data["event_commission"], 2),
                "event_net": round(data["event_net"], 2),
                "event_count": data["event_count"],
                # Rezervasyon
                "reservation_earnings": round(data["reservation_earnings"], 2),
                "reservation_commission": round(data["reservation_commission"], 2),
                "reservation_net": round(data["reservation_net"], 2),
                "reservation_count": data["reservation_count"],
                # Toplam
                "total_gross": round(data["total_gross"], 2),
                "total_commission": round(data["total_commission"], 2),
                "total_net": round(data["total_net"], 2),
                "total_count": data["marketplace_count"] + data["event_count"] + data["reservation_count"],
                # Durum
                "payment_status": payment_record.get("status", "pending") if payment_record else "pending",
                "payment_date": payment_record.get("payment_date").isoformat() if payment_record and payment_record.get("payment_date") else None,
                "period_start": period_start.isoformat(),
                "period_end": period_end.isoformat()
            })
        
        # Net tutara göre sırala (büyükten küçüğe)
        result.sort(key=lambda x: x["total_net"], reverse=True)
        
        # Toplam istatistikler
        total_stats = {
            "total_users": len(result),
            "total_gross": sum(r["total_gross"] for r in result),
            "total_commission": sum(r["total_commission"] for r in result),
            "total_net": sum(r["total_net"] for r in result),
            "marketplace_total": sum(r["marketplace_net"] for r in result),
            "event_total": sum(r["event_net"] for r in result),
            "reservation_total": sum(r["reservation_net"] for r in result),
            "pending_count": len([r for r in result if r["payment_status"] == "pending"]),
            "paid_count": len([r for r in result if r["payment_status"] == "paid"]),
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "period_label": f"{period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}",
            "year": year,
            "month": month,
            "week": week
        }
        
        return {
            "success": True,
            "stats": total_stats,
            "users": result
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Payment report error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/admin/mark-sellers-paid")
async def mark_sellers_paid(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Seçili satıcıları ödendi olarak işaretle"""
    try:
        current_user_id = current_user["id"]
        
        # Admin kontrolü
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        body = await request.json()
        seller_ids = body.get("seller_ids", [])
        period_start = body.get("period_start")
        period_end = body.get("period_end")
        
        if not seller_ids:
            raise HTTPException(status_code=400, detail="En az bir satıcı seçmelisiniz")
        
        period_start_dt = datetime.fromisoformat(period_start.replace("Z", ""))
        period_end_dt = datetime.fromisoformat(period_end.replace("Z", ""))
        
        marked_count = 0
        for seller_id in seller_ids:
            # Ödeme kaydı oluştur veya güncelle
            existing = await db.seller_payments.find_one({
                "seller_id": seller_id,
                "period_start": period_start_dt,
                "period_end": period_end_dt
            })
            
            if existing:
                await db.seller_payments.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {
                        "status": "paid",
                        "payment_date": datetime.utcnow(),
                        "marked_by": current_user_id
                    }}
                )
            else:
                await db.seller_payments.insert_one({
                    "id": str(uuid.uuid4()),
                    "seller_id": seller_id,
                    "period_start": period_start_dt,
                    "period_end": period_end_dt,
                    "status": "paid",
                    "payment_date": datetime.utcnow(),
                    "marked_by": current_user_id,
                    "created_at": datetime.utcnow()
                })
            
            # Satıcıya bildirim gönder
            seller = await db.users.find_one({"id": seller_id})
            if seller:
                notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": seller_id,
                    "type": "payment_completed",
                    "title": "💰 Ödemeniz Yapıldı",
                    "message": f"{period_start_dt.strftime('%d.%m.%Y')} - {period_end_dt.strftime('%d.%m.%Y')} dönemi satış kazançlarınız hesabınıza aktarıldı.",
                    "read": False,
                    "created_at": datetime.utcnow()
                }
                await db.notifications.insert_one(notification)
            
            marked_count += 1
        
        # Log aktivite
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "admin_mark_payments", "success", {
                "seller_count": marked_count,
                "period": f"{period_start_dt.strftime('%d.%m.%Y')} - {period_end_dt.strftime('%d.%m.%Y')}"
            })
        except:
            pass
        
        return {
            "success": True,
            "message": f"{marked_count} satıcının ödemesi tamamlandı olarak işaretlendi"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Mark sellers paid error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/admin/export-payments-excel")
async def export_payments_excel(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Seçili satıcıların ödeme bilgilerini Excel olarak dışa aktar"""
    from fastapi.responses import Response
    import io
    
    try:
        current_user_id = current_user["id"]
        
        # Admin kontrolü
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        data = await request.json()
        user_ids = data.get("user_ids", [])
        period_start = data.get("period_start")
        period_end = data.get("period_end")
        
        if not user_ids:
            raise HTTPException(status_code=400, detail="En az bir kullanıcı seçilmeli")
        
        # openpyxl import
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel oluşturma modülü yüklü değil")
        
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Ödeme Raporu"
        
        # Stiller
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        money_alignment = Alignment(horizontal="right")
        center_alignment = Alignment(horizontal="center")
        
        # Başlıklar
        headers = [
            "Sıra", "Kullanıcı Adı", "Telefon", "IBAN", "Kullanıcı Tipi",
            "Market Satış", "Market Komisyon", "Market Net",
            "Etkinlik Gelir", "Etkinlik Komisyon", "Etkinlik Net",
            "Rezv. Gelir", "Rezv. Komisyon", "Rezv. Net",
            "Toplam Brüt", "Toplam Komisyon", "Toplam Net", "Durum"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Satır yüksekliği
        ws.row_dimensions[1].height = 30
        
        # Kullanıcı verileri
        row_num = 2
        total_gross = 0
        total_commission = 0
        total_net = 0
        
        for idx, user_id in enumerate(user_ids, 1):
            # Kullanıcı bilgileri
            user_info = await db.users.find_one({"id": user_id})
            if not user_info:
                continue
            
            # Ödeme verilerini hesapla (basitleştirilmiş - frontend'den gelen veriler kullanılmalı)
            # Gerçek verileri almak için seller-payments-report endpoint mantığını kullanabiliriz
            # Ama burada frontend'den gelen verilerle çalışacağız
            
            user_payment = data.get("users_data", {}).get(user_id, {})
            
            ws.cell(row=row_num, column=1, value=idx).alignment = center_alignment
            ws.cell(row=row_num, column=2, value=user_info.get("full_name", "Bilinmeyen"))
            ws.cell(row=row_num, column=3, value=user_info.get("phone", ""))
            ws.cell(row=row_num, column=4, value=user_info.get("iban", ""))
            ws.cell(row=row_num, column=5, value=user_info.get("user_type", "")).alignment = center_alignment
            
            # Market
            ws.cell(row=row_num, column=6, value=user_payment.get("marketplace_sales", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=7, value=user_payment.get("marketplace_commission", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=8, value=user_payment.get("marketplace_net", 0)).alignment = money_alignment
            
            # Etkinlik
            ws.cell(row=row_num, column=9, value=user_payment.get("event_earnings", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=10, value=user_payment.get("event_commission", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=11, value=user_payment.get("event_net", 0)).alignment = money_alignment
            
            # Rezervasyon
            ws.cell(row=row_num, column=12, value=user_payment.get("reservation_earnings", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=13, value=user_payment.get("reservation_commission", 0)).alignment = money_alignment
            ws.cell(row=row_num, column=14, value=user_payment.get("reservation_net", 0)).alignment = money_alignment
            
            # Toplam
            gross = user_payment.get("total_gross", 0)
            commission = user_payment.get("total_commission", 0)
            net = user_payment.get("total_net", 0)
            
            ws.cell(row=row_num, column=15, value=gross).alignment = money_alignment
            ws.cell(row=row_num, column=16, value=commission).alignment = money_alignment
            ws.cell(row=row_num, column=17, value=net).alignment = money_alignment
            
            status = "Ödendi" if user_payment.get("payment_status") == "paid" else "Bekliyor"
            status_cell = ws.cell(row=row_num, column=18, value=status)
            status_cell.alignment = center_alignment
            if status == "Ödendi":
                status_cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            
            # Border ekle
            for col in range(1, 19):
                ws.cell(row=row_num, column=col).border = border
            
            total_gross += gross
            total_commission += commission
            total_net += net
            
            row_num += 1
        
        # Toplam satırı
        ws.cell(row=row_num, column=1, value="").border = border
        ws.cell(row=row_num, column=2, value="TOPLAM").font = Font(bold=True)
        ws.cell(row=row_num, column=2).border = border
        for col in range(3, 15):
            ws.cell(row=row_num, column=col, value="").border = border
        
        ws.cell(row=row_num, column=15, value=total_gross).font = Font(bold=True)
        ws.cell(row=row_num, column=15).alignment = money_alignment
        ws.cell(row=row_num, column=15).border = border
        
        ws.cell(row=row_num, column=16, value=total_commission).font = Font(bold=True)
        ws.cell(row=row_num, column=16).alignment = money_alignment
        ws.cell(row=row_num, column=16).border = border
        
        ws.cell(row=row_num, column=17, value=total_net).font = Font(bold=True)
        ws.cell(row=row_num, column=17).alignment = money_alignment
        ws.cell(row=row_num, column=17).border = border
        
        ws.cell(row=row_num, column=18, value="").border = border
        
        # Sütun genişlikleri
        column_widths = [6, 25, 15, 30, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14, 14, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Dönem bilgisi ekle
        ws.cell(row=row_num + 2, column=1, value="Dönem:")
        ws.cell(row=row_num + 2, column=2, value=f"{period_start} - {period_end}")
        ws.cell(row=row_num + 3, column=1, value="Oluşturulma:")
        ws.cell(row=row_num + 3, column=2, value=datetime.utcnow().strftime("%d.%m.%Y %H:%M"))
        
        # Excel dosyasını memory'ye yaz
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "EXPORT_PAYMENTS_EXCEL", "success", {
                "user_count": len(user_ids),
                "period": f"{period_start} - {period_end}"
            })
        except:
            pass
        
        filename = f"odeme_raporu_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Export payments excel error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ADMIN İADE YÖNETİMİ ====================

@router.get("/admin/returns")
async def get_admin_returns(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Admin için tüm iade taleplerini getir"""
    try:
        # Admin kontrolü
        if current_user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Admin yetkisi gerekli")
        
        # Filtre
        query = {}
        if status:
            query["status"] = status
        
        # İade taleplerini getir
        returns = await db.marketplace_returns.find(query).sort("created_at", -1).to_list(500)
        
        # İstatistikler
        all_returns = await db.marketplace_returns.find({}).to_list(1000)
        stats = {
            "disputed": len([r for r in all_returns if r.get("status") == "disputed"]),
            "pending": len([r for r in all_returns if r.get("status") == "pending"]),
            "approved": len([r for r in all_returns if r.get("status") in ["approved", "return_shipped"]]),
            "completed": len([r for r in all_returns if r.get("status") in ["returned", "completed", "admin_approved", "admin_rejected"]]),
            "total": len(all_returns)
        }
        
        # Her iade için detay bilgileri ekle
        result = []
        for ret in returns:
            ret.pop("_id", None)
            
            listing = await db.marketplace_listings.find_one({"id": ret.get("listing_id")})
            buyer = await db.users.find_one({"id": ret.get("buyer_id")})
            seller = await db.users.find_one({"id": ret.get("seller_id")})
            transaction = await db.marketplace_transactions.find_one({"id": ret.get("order_id")})
            
            if listing:
                listing.pop("_id", None)
            if buyer:
                buyer.pop("_id", None)
                buyer.pop("password", None)
            if seller:
                seller.pop("_id", None)
                seller.pop("password", None)
            if transaction:
                transaction.pop("_id", None)
            
            result.append({
                **ret,
                "listing": listing,
                "buyer": buyer,
                "seller": seller,
                "transaction": transaction
            })
        
        return {"returns": result, "stats": stats}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Admin get returns error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/admin/returns/{return_id}/approve")
async def admin_approve_return(
    return_id: str,
    admin_note: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Admin iade onayı - Alıcı lehine karar"""
    try:
        # Admin kontrolü
        if current_user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Admin yetkisi gerekli")
        
        return_request = await db.marketplace_returns.find_one({"id": return_id})
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        # İade talebini onayla
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {
                "status": "approved",
                "admin_approved": True,
                "admin_note": admin_note,
                "admin_id": current_user["id"],
                "admin_decision_at": datetime.utcnow()
            }}
        )
        
        # Sipariş durumunu güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {"status": "return_approved"}}
        )
        
        # Geliver'dan iade kargo etiketi oluştur
        transaction = await db.marketplace_transactions.find_one({"id": return_request.get("order_id")})
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        buyer = await db.users.find_one({"id": return_request.get("buyer_id")})
        seller = await db.users.find_one({"id": return_request.get("seller_id")})
        
        geliver_result = await create_geliver_return_shipment(
            return_request=return_request,
            transaction=transaction or {},
            listing=listing or {},
            buyer=buyer or {},
            seller=seller or {}
        )
        
        return_tracking_code = geliver_result.get("tracking_code", "")
        return_shipping_cost = geliver_result.get("shipping_cost", 0)
        return_provider_name = geliver_result.get("provider_name", "Kargo")
        
        # Kargo bilgilerini kaydet
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {
                "return_tracking_code": return_tracking_code,
                "return_shipping_cost": return_shipping_cost,
                "return_provider_name": return_provider_name
            }}
        )
        
        # Alıcıya bildirim
        buyer_notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("buyer_id"),
            "type": "admin_return_approved",
            "title": "✅ İade Onaylandı (Yönetici Kararı)",
            "message": f"İade talebiniz yönetici tarafından onaylandı.\n\n📦 Kargo: {return_provider_name}\n🔢 Takip Kodu: {return_tracking_code}\n💰 Kargo Ücreti: ₺{return_shipping_cost:.2f}\n\nLütfen 4 iş günü içinde ürünü kargoya verin.",
            "data": {"return_id": return_id, "tracking_code": return_tracking_code},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(buyer_notification)
        
        # Satıcıya bildirim
        seller_notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("seller_id"),
            "type": "admin_return_approved",
            "title": "ℹ️ İade Kararı",
            "message": f"İade talebine itirazınız değerlendirildi ve iade onaylandı.\n\n{admin_note}",
            "data": {"return_id": return_id},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(seller_notification)
        
        logger.info(f"✅ Admin iade onayladı - Return ID: {return_id}, Admin: {current_user['id']}")
        
        return {"message": "İade onaylandı", "tracking_code": return_tracking_code}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Admin approve return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/admin/returns/{return_id}/reject")
async def admin_reject_return(
    return_id: str,
    admin_note: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Admin iade reddi - Satıcı lehine karar"""
    try:
        # Admin kontrolü
        if current_user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Admin yetkisi gerekli")
        
        return_request = await db.marketplace_returns.find_one({"id": return_id})
        if not return_request:
            raise HTTPException(status_code=404, detail="İade talebi bulunamadı")
        
        # İade talebini reddet
        await db.marketplace_returns.update_one(
            {"id": return_id},
            {"$set": {
                "status": "admin_rejected",
                "admin_rejected": True,
                "admin_note": admin_note,
                "admin_id": current_user["id"],
                "admin_decision_at": datetime.utcnow()
            }}
        )
        
        # Sipariş durumunu tamamlandı olarak güncelle
        await db.marketplace_transactions.update_one(
            {"id": return_request.get("order_id")},
            {"$set": {"status": "completed"}}
        )
        
        listing = await db.marketplace_listings.find_one({"id": return_request.get("listing_id")})
        
        # Alıcıya bildirim
        buyer_notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("buyer_id"),
            "type": "admin_return_rejected",
            "title": "❌ İade Reddedildi (Yönetici Kararı)",
            "message": f"İade talebiniz yönetici tarafından reddedildi.\n\nGerekçe: {admin_note}",
            "data": {"return_id": return_id},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(buyer_notification)
        
        # Satıcıya bildirim
        seller_notification = {
            "id": str(uuid.uuid4()),
            "user_id": return_request.get("seller_id"),
            "type": "admin_return_rejected",
            "title": "✅ İade Talebi Reddedildi",
            "message": f"'{listing.get('title', 'Ürün')}' için yapılan iade talebi yönetici tarafından reddedildi. Sipariş tamamlandı olarak işaretlendi.",
            "data": {"return_id": return_id},
            "action_url": "/marketplace/reports",
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(seller_notification)
        
        logger.info(f"❌ Admin iade reddetti - Return ID: {return_id}, Admin: {current_user['id']}")
        
        return {"message": "İade reddedildi"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Admin reject return error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
