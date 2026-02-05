"""
Commission Settings & Earnings Endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
import logging
import io

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter()

# Database reference
db = None

def set_db(database):
    global db
    db = database

# Models
class TieredRate(BaseModel):
    """Tutara göre değişen komisyon oranı"""
    min_amount: float = 0
    max_amount: Optional[float] = None  # None = sınırsız
    rate: float = 10.0  # % oran


class CommissionSettings(BaseModel):
    coach_commission_rate: float = 10.0  # % oran
    coach_commission_fixed: float = 0.0  # Sabit tutar
    coach_free_trial_months: int = 0  # Ücretsiz deneme süresi
    coach_tiered_rates: Optional[List[dict]] = None  # Tutara göre değişen oranlar
    
    event_commission_rate: float = 10.0
    event_commission_fixed: float = 0.0
    event_free_trial_months: int = 0
    event_tiered_rates: Optional[List[dict]] = None
    
    referee_commission_rate: float = 10.0
    referee_commission_fixed: float = 0.0
    referee_free_trial_months: int = 0
    referee_tiered_rates: Optional[List[dict]] = None
    
    facility_commission_rate: float = 10.0
    facility_commission_fixed: float = 0.0
    facility_free_trial_months: int = 0
    facility_tiered_rates: Optional[List[dict]] = None
    
    membership_commission_rate: float = 10.0
    membership_commission_fixed: float = 0.0
    membership_free_trial_months: int = 0
    membership_tiered_rates: Optional[List[dict]] = None
    
    marketplace_commission_rate: float = 10.0
    marketplace_commission_fixed: float = 100.0  # Kargo ücreti (₺)
    marketplace_free_trial_months: int = 0
    marketplace_tiered_rates: Optional[List[dict]] = None


class CommissionSettingsUpdate(BaseModel):
    coach_commission_rate: Optional[float] = None
    coach_commission_fixed: Optional[float] = None
    coach_free_trial_months: Optional[int] = None
    coach_tiered_rates: Optional[List[dict]] = None
    
    event_commission_rate: Optional[float] = None
    event_commission_fixed: Optional[float] = None
    event_free_trial_months: Optional[int] = None
    event_tiered_rates: Optional[List[dict]] = None
    
    referee_commission_rate: Optional[float] = None
    referee_commission_fixed: Optional[float] = None
    referee_free_trial_months: Optional[int] = None
    referee_tiered_rates: Optional[List[dict]] = None
    
    facility_commission_rate: Optional[float] = None
    facility_commission_fixed: Optional[float] = None
    facility_free_trial_months: Optional[int] = None
    facility_tiered_rates: Optional[List[dict]] = None
    
    membership_commission_rate: Optional[float] = None
    membership_commission_fixed: Optional[float] = None
    membership_free_trial_months: Optional[int] = None
    membership_tiered_rates: Optional[List[dict]] = None
    
    marketplace_commission_rate: Optional[float] = None
    marketplace_commission_fixed: Optional[float] = None
    marketplace_free_trial_months: Optional[int] = None
    marketplace_tiered_rates: Optional[List[dict]] = None


# System user phone number - collects all commissions
SYSTEM_USER_PHONE = "+905324900472"


# Get commission settings
@router.get("/commission-settings")
async def get_commission_settings():
    """Get current commission settings"""
    try:
        settings = await db.settings.find_one({"type": "commission"})
        
        if not settings:
            # Return default settings
            default_settings = CommissionSettings()
            return {
                "success": True,
                "settings": default_settings.dict()
            }
        
        return {
            "success": True,
            "settings": {
                "coach_commission_rate": settings.get("coach_commission_rate", 10.0),
                "coach_commission_fixed": settings.get("coach_commission_fixed", 0.0),
                "coach_free_trial_months": settings.get("coach_free_trial_months", 0),
                "coach_tiered_rates": settings.get("coach_tiered_rates", []),
                
                "event_commission_rate": settings.get("event_commission_rate", 10.0),
                "event_commission_fixed": settings.get("event_commission_fixed", 0.0),
                "event_free_trial_months": settings.get("event_free_trial_months", 0),
                "event_tiered_rates": settings.get("event_tiered_rates", []),
                
                "referee_commission_rate": settings.get("referee_commission_rate", 10.0),
                "referee_commission_fixed": settings.get("referee_commission_fixed", 0.0),
                "referee_free_trial_months": settings.get("referee_free_trial_months", 0),
                "referee_tiered_rates": settings.get("referee_tiered_rates", []),
                
                "facility_commission_rate": settings.get("facility_commission_rate", 10.0),
                "facility_commission_fixed": settings.get("facility_commission_fixed", 0.0),
                "facility_free_trial_months": settings.get("facility_free_trial_months", 0),
                "facility_tiered_rates": settings.get("facility_tiered_rates", []),
                
                "membership_commission_rate": settings.get("membership_commission_rate", 10.0),
                "membership_commission_fixed": settings.get("membership_commission_fixed", 0.0),
                "membership_free_trial_months": settings.get("membership_free_trial_months", 0),
                "membership_tiered_rates": settings.get("membership_tiered_rates", []),
                
                "marketplace_commission_rate": settings.get("marketplace_commission_rate", 10.0),
                "marketplace_commission_fixed": settings.get("marketplace_commission_fixed", 100.0),  # Kargo ücreti varsayılan 100₺
                "marketplace_free_trial_months": settings.get("marketplace_free_trial_months", 0),
                "marketplace_tiered_rates": settings.get("marketplace_tiered_rates", []),
            }
        }
    except Exception as e:
        logger.error(f"Error getting commission settings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Update commission settings (Super Admin only)
@router.put("/commission-settings")
async def update_commission_settings(settings: CommissionSettingsUpdate):
    """Update commission settings"""
    try:
        update_data = {k: v for k, v in settings.dict().items() if v is not None}
        update_data["type"] = "commission"
        update_data["updated_at"] = datetime.utcnow()
        
        await db.settings.update_one(
            {"type": "commission"},
            {"$set": update_data},
            upsert=True
        )
        
        logger.info(f"✅ Commission settings updated: {update_data}")
        
        return {
            "success": True,
            "message": "Komisyon ayarları güncellendi"
        }
    except Exception as e:
        logger.error(f"Error updating commission settings: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Get user earnings/payments
@router.get("/earnings/{user_id}")
async def get_user_earnings(
    user_id: str,
    period: str = "all",  # weekly, monthly, yearly, all
    status: str = "all"   # paid, pending, all
):
    """Get user earnings and payments with commissions"""
    try:
        # Date filter
        date_filter = {}
        now = datetime.utcnow()
        
        if period == "daily":
            date_filter = {"$gte": now - timedelta(days=1)}
        elif period == "weekly":
            date_filter = {"$gte": now - timedelta(days=7)}
        elif period == "monthly":
            date_filter = {"$gte": now - timedelta(days=30)}
        elif period == "yearly":
            date_filter = {"$gte": now - timedelta(days=365)}
        
        # Check if this is the system user
        user = await db.users.find_one({"id": user_id})
        is_system_user = user and user.get("phone") == SYSTEM_USER_PHONE
        
        # Get commission settings
        settings = await db.settings.find_one({"type": "commission"})
        if not settings:
            settings = CommissionSettings().dict()
        
        # Get all relevant transactions
        earnings_list = []
        total_gross = 0
        total_commission = 0
        total_net = 0
        
        # For system user - collect ALL transactions as commission income
        if is_system_user:
            # Coach reservations - all of them
            coach_reservations = await db.coach_reservations.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in coach_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("coach_commission_rate", 10)
                commission_fixed = settings.get("coach_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "coach",
                    "type_label": "Antrenör Komisyonu",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),  # System user gets commission as income
                    "status": res.get("payment_status", "pending"),
                    "description": f"{res.get('coach_name', 'Antrenör')} - {res.get('date', '')}",
                    "user_id": res.get("coach_id"),
                    "user_name": res.get("coach_name", "Antrenör"),
                    "client_id": res.get("user_id"),
                    "client_name": res.get("client_name", res.get("user_name", "Müşteri"))
                })
                total_gross += amount
                total_commission += commission
                total_net += commission  # For system user, commission = net income
            
            # Facility reservations
            facility_reservations = await db.reservations.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in facility_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("facility_commission_rate", 10)
                commission_fixed = settings.get("facility_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "facility",
                    "type_label": "Tesis Komisyonu",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": res.get("payment_status", "pending"),
                    "description": f"{res.get('facility_name', 'Tesis')} - {res.get('date', '')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += commission
            
            # Event payments
            event_payments = await db.event_payments.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for payment in event_payments:
                amount = payment.get("amount", 0)
                commission_rate = settings.get("event_commission_rate", 10)
                commission_fixed = settings.get("event_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "id": str(payment.get("_id", payment.get("id"))),
                    "type": "event",
                    "type_label": "Etkinlik Komisyonu",
                    "date": payment.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": payment.get("status", "pending"),
                    "description": f"{payment.get('event_name', 'Etkinlik')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += commission
            
            # Marketplace orders
            marketplace_orders = await db.marketplace_orders.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for order in marketplace_orders:
                amount = order.get("total_price", 0)
                commission_rate = settings.get("marketplace_commission_rate", 10)
                commission_fixed = settings.get("marketplace_commission_fixed", 0)
                commission = tx.get("commission_amount", (amount * commission_rate / 100))
                
                # Listing bilgilerini al
                listing = await db.marketplace_listings.find_one({"id": tx.get("listing_id")})
                listing_title = listing.get("title", "Spor Market") if listing else "Spor Market"
                
                earnings_list.append({
                    "id": str(tx.get("_id", tx.get("id"))),
                    "type": "marketplace",
                    "type_label": "Spor Market Komisyonu",
                    "date": tx.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": tx.get("shipping_fee", 0),
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": tx.get("status", "pending"),
                    "description": listing_title,
                    "user_id": tx.get("seller_id"),
                    "user_name": listing.get("seller_name", "") if listing else "",
                    "client_id": tx.get("buyer_id"),
                    "client_name": ""
                })
                total_gross += amount
                total_commission += commission
                total_net += commission
            
            # Memberships
            memberships = await db.memberships.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for membership in memberships:
                amount = membership.get("price", 0)
                commission_rate = settings.get("membership_commission_rate", 10)
                commission_fixed = settings.get("membership_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "id": str(membership.get("_id", membership.get("id"))),
                    "type": "membership",
                    "type_label": "Üyelik Komisyonu",
                    "date": membership.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": membership.get("payment_status", "pending"),
                    "description": f"{membership.get('facility_name', 'Tesis')} - {membership.get('membership_type', '')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += commission
        
        else:
            # Normal user - check person_reservation_transactions first (primary source)
            person_txs = await db.person_reservation_transactions.find({
                "seller_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for tx in person_txs:
                amount = tx.get("total_amount", 0)
                commission_rate = tx.get("commission_rate", 10)
                commission = tx.get("commission_amount", 0)
                net = tx.get("seller_receives", amount - commission)
                res_type = tx.get("reservation_type", "other")
                
                # Get buyer info
                buyer_id = tx.get("buyer_id")
                buyer = await db.users.find_one({"id": buyer_id}) if buyer_id else None
                buyer_name = buyer.get("full_name", "Müşteri") if buyer else "Müşteri"
                
                type_labels = {
                    "coach": "Antrenör Rezervasyonu",
                    "player": "Oyuncu Kiralama",
                    "referee": "Hakem Rezervasyonu",
                    "facility": "Tesis Rezervasyonu"
                }
                
                earnings_list.append({
                    "id": str(tx.get("_id", tx.get("id"))),
                    "type": res_type,
                    "type_label": type_labels.get(res_type, "Rezervasyon"),
                    "date": tx.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": 0,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": tx.get("payment_status", "pending"),
                    "description": f"{buyer_name} - {tx.get('created_at', '')[:10] if tx.get('created_at') else ''}",
                    "client_id": buyer_id,
                    "client_name": buyer_name
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Also check legacy coach_reservations table
            coach_reservations = await db.coach_reservations.find({
                "coach_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in coach_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("coach_commission_rate", 10)
                commission_fixed = settings.get("coach_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "coach",
                    "type_label": "Antrenör Rezervasyonu",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": res.get("payment_status", "pending"),
                    "description": f"{res.get('client_name', 'Müşteri')} - {res.get('date', '')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Coach reservations from reservations table (using coach_id)
            coach_from_reservations = await db.reservations.find({
                "coach_id": user_id,
                "reservation_type": "coach",
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "approved"} if status == "paid" else 
                   {"status": {"$in": ["pending", "approved"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in coach_from_reservations:
                amount = res.get("total_price", 0) or res.get("hourly_rate", 0)
                commission_rate = settings.get("coach_commission_rate", 10)
                commission_fixed = settings.get("coach_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                # Get buyer info
                buyer_id = res.get("user_id")
                buyer = await db.users.find_one({"id": buyer_id}) if buyer_id else None
                buyer_name = buyer.get("full_name", "Müşteri") if buyer else "Müşteri"
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "coach",
                    "type_label": "Antrenör Rezervasyonu",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": "paid" if res.get("status") == "approved" else res.get("status", "pending"),
                    "description": f"{buyer_name} - {res.get('date', '')}",
                    "client_id": buyer_id,
                    "client_name": buyer_name
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Player reservations from reservations table (using player_id)
            player_from_reservations = await db.reservations.find({
                "player_id": user_id,
                "reservation_type": "player",
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "approved"} if status == "paid" else 
                   {"status": {"$in": ["pending", "approved"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in player_from_reservations:
                amount = res.get("total_price", 0)
                commission_rate = settings.get("player_commission_rate", 10)
                commission_fixed = settings.get("player_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                # Get buyer info
                buyer_id = res.get("user_id")
                buyer = await db.users.find_one({"id": buyer_id}) if buyer_id else None
                buyer_name = buyer.get("full_name", "Müşteri") if buyer else "Müşteri"
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "player",
                    "type_label": "Oyuncu Kiralama",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": "paid" if res.get("status") == "approved" else res.get("status", "pending"),
                    "description": f"{buyer_name} - {res.get('selected_date', '')}",
                    "client_id": buyer_id,
                    "client_name": buyer_name
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Referee reservations from reservations table (using referee_id)
            referee_from_reservations = await db.reservations.find({
                "referee_id": user_id,
                "reservation_type": "referee",
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "approved"} if status == "paid" else 
                   {"status": {"$in": ["pending", "approved"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in referee_from_reservations:
                amount = res.get("total_price", 0) or res.get("hourly_rate", 0)
                commission_rate = settings.get("referee_commission_rate", 10)
                commission_fixed = settings.get("referee_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                # Get buyer info
                buyer_id = res.get("user_id")
                buyer = await db.users.find_one({"id": buyer_id}) if buyer_id else None
                buyer_name = buyer.get("full_name", "Müşteri") if buyer else "Müşteri"
                
                earnings_list.append({
                    "id": str(res.get("_id", res.get("id"))),
                    "type": "referee",
                    "type_label": "Hakem Rezervasyonu",
                    "date": res.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": "paid" if res.get("status") == "approved" else res.get("status", "pending"),
                    "description": f"{buyer_name} - {res.get('date', '')}",
                    "client_id": buyer_id,
                    "client_name": buyer_name
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Event payments (as organizer)
            event_payments = await db.event_payments.find({
                "organizer_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for payment in event_payments:
                amount = payment.get("amount", 0)
                commission_rate = settings.get("event_commission_rate", 10)
                commission_fixed = settings.get("event_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "id": str(payment.get("_id", payment.get("id"))),
                    "type": "event",
                    "type_label": "Etkinlik Ödemesi",
                    "date": payment.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": payment.get("status", "pending"),
                    "description": f"{payment.get('event_name', 'Etkinlik')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Marketplace transactions (seller earnings)
            marketplace_txs = await db.marketplace_transactions.find({
                "seller_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "pending_payment", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for tx in marketplace_txs:
                amount = tx.get("item_price", 0)
                commission_rate = settings.get("marketplace_commission_rate", 10)
                commission = tx.get("commission_amount", (amount * commission_rate / 100))
                net = tx.get("seller_receives", amount - commission)
                
                # Listing bilgilerini al
                listing = await db.marketplace_listings.find_one({"id": tx.get("listing_id")})
                listing_title = listing.get("title", "Spor Market") if listing else "Spor Market"
                
                earnings_list.append({
                    "id": str(tx.get("_id", tx.get("id"))),
                    "type": "marketplace",
                    "type_label": "Spor Market Satışı",
                    "date": tx.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": tx.get("shipping_fee", 0),
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": tx.get("status", "pending"),
                    "description": listing_title,
                    "user_id": tx.get("seller_id"),
                    "user_name": listing.get("seller_name", "") if listing else "",
                    "client_id": tx.get("buyer_id"),
                    "client_name": ""
                })
                total_gross += amount
                total_commission += commission
                total_net += net
            
            # Memberships
            memberships = await db.memberships.find({
                "facility_owner_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for membership in memberships:
                amount = membership.get("price", 0)
                commission_rate = settings.get("membership_commission_rate", 10)
                commission_fixed = settings.get("membership_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "id": str(membership.get("_id", membership.get("id"))),
                    "type": "membership",
                    "type_label": "Üyelik Ödemesi",
                    "date": membership.get("created_at"),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": membership.get("payment_status", "pending"),
                    "description": f"{membership.get('facility_name', 'Tesis')} - {membership.get('membership_type', '')}"
                })
                total_gross += amount
                total_commission += commission
                total_net += net
        
        # Sort by date descending (handle both datetime and string dates)
        def get_sort_date(x):
            date_val = x.get("date")
            if date_val is None:
                return datetime.min
            if isinstance(date_val, str):
                try:
                    return datetime.fromisoformat(date_val.replace('Z', '+00:00'))
                except:
                    return datetime.min
            return date_val
        
        earnings_list.sort(key=get_sort_date, reverse=True)
        
        return {
            "success": True,
            "summary": {
                "total_gross": round(total_gross, 2),
                "total_commission": round(total_commission, 2),
                "total_net": round(total_net, 2),
                "transaction_count": len(earnings_list)
            },
            "earnings": earnings_list
        }
    except Exception as e:
        logger.error(f"Error getting user earnings: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# Excel Export endpoint
@router.get("/earnings/{user_id}/export")
async def export_user_earnings(
    user_id: str,
    period: str = "all",  # weekly, monthly, yearly, all
    status: str = "all"   # paid, pending, all
):
    """Export user earnings to Excel file"""
    try:
        # Reuse the existing earnings function to get data
        # Get earnings data using existing function logic
        date_filter = {}
        now = datetime.utcnow()
        
        if period == "daily":
            date_filter = {"$gte": now - timedelta(days=1)}
        elif period == "weekly":
            date_filter = {"$gte": now - timedelta(days=7)}
        elif period == "monthly":
            date_filter = {"$gte": now - timedelta(days=30)}
        elif period == "yearly":
            date_filter = {"$gte": now - timedelta(days=365)}
        
        # Check if this is the system user
        user = await db.users.find_one({"id": user_id})
        is_system_user = user and user.get("phone") == SYSTEM_USER_PHONE
        user_name = user.get("full_name", "Kullanıcı") if user else "Kullanıcı"
        
        # Get commission settings
        settings = await db.settings.find_one({"type": "commission"})
        if not settings:
            settings = CommissionSettings().dict()
        
        earnings_list = []
        
        # ======== COLLECT ALL DATA SOURCES ========
        
        if is_system_user:
            # === SYSTEM USER: Gets ALL platform commissions ===
            
            # 1. Coach reservations
            coach_reservations = await db.coach_reservations.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in coach_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("coach_commission_rate", 10)
                commission_fixed = settings.get("coach_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "date": res.get("created_at"),
                    "type_label": "Antrenör",
                    "item_name": res.get("coach_name", "Antrenör"),
                    "description": res.get("date", ""),
                    "user_name": res.get("coach_name", ""),
                    "client_name": res.get("client_name", res.get("user_name", "")),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": res.get("payment_status", "pending"),
                })
            
            # 2. Facility reservations
            facility_reservations = await db.reservations.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in facility_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("facility_commission_rate", 10)
                commission_fixed = settings.get("facility_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "date": res.get("created_at"),
                    "type_label": "Tesis",
                    "item_name": res.get("facility_name", "Tesis"),
                    "description": res.get("date", ""),
                    "user_name": res.get("facility_name", ""),
                    "client_name": res.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": res.get("payment_status", "pending"),
                })
            
            # 3. Event payments
            event_payments = await db.event_payments.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for payment in event_payments:
                amount = payment.get("amount", 0)
                commission_rate = settings.get("event_commission_rate", 10)
                commission_fixed = settings.get("event_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "date": payment.get("created_at"),
                    "type_label": "Etkinlik",
                    "item_name": payment.get("event_name", "Etkinlik"),
                    "description": payment.get("event_name", ""),
                    "user_name": payment.get("organizer_name", ""),
                    "client_name": payment.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": payment.get("status", "pending"),
                })
            
            # 4. Referee reservations
            referee_reservations = await db.referee_reservations.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in referee_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("referee_commission_rate", 10)
                commission_fixed = settings.get("referee_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "date": res.get("created_at"),
                    "type_label": "Hakem",
                    "item_name": res.get("referee_name", "Hakem"),
                    "description": res.get("date", ""),
                    "user_name": res.get("referee_name", ""),
                    "client_name": res.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": res.get("payment_status", "pending"),
                })
            
            # 5. Memberships
            memberships = await db.memberships.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for membership in memberships:
                amount = membership.get("price", 0)
                commission_rate = settings.get("membership_commission_rate", 10)
                commission_fixed = settings.get("membership_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                
                earnings_list.append({
                    "date": membership.get("created_at"),
                    "type_label": "Üyelik",
                    "item_name": membership.get("facility_name", "Tesis Üyeliği"),
                    "description": membership.get("membership_type", ""),
                    "user_name": membership.get("facility_name", ""),
                    "client_name": membership.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": membership.get("payment_status", "pending"),
                })
            
            # 6. Marketplace transactions (satışlar)
            marketplace_txs = await db.marketplace_transactions.find({
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "pending_payment", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for tx in marketplace_txs:
                amount = tx.get("item_price", 0)
                commission_rate = settings.get("marketplace_commission_rate", 10)
                commission = tx.get("commission_amount", (amount * commission_rate / 100))
                
                # Get listing info
                listing = await db.marketplace_listings.find_one({"id": tx.get("listing_id")})
                listing_title = listing.get("title", "Spor Market") if listing else "Spor Market"
                seller_name = listing.get("seller_name", "") if listing else ""
                
                earnings_list.append({
                    "date": tx.get("created_at"),
                    "type_label": "Spor Market",
                    "item_name": listing_title,
                    "description": "",
                    "user_name": seller_name,
                    "client_name": "",
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": tx.get("shipping_fee", 0),
                    "commission_amount": round(commission, 2),
                    "net_amount": round(commission, 2),
                    "status": tx.get("status", "pending"),
                })
        
        else:
            # === REGULAR USER: Gets their own earnings ===
            
            # 1. Coach reservations (as coach)
            coach_reservations = await db.coach_reservations.find({
                "coach_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in coach_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("coach_commission_rate", 10)
                commission_fixed = settings.get("coach_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "date": res.get("created_at"),
                    "type_label": "Antrenör",
                    "item_name": "Antrenörlük Hizmeti",
                    "description": res.get("date", ""),
                    "user_name": res.get("coach_name", user_name),
                    "client_name": res.get("client_name", res.get("user_name", "")),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": res.get("payment_status", "pending"),
                })
            
            # 2. Facility reservations (as facility owner)
            facility_reservations = await db.reservations.find({
                "facility_owner_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for res in facility_reservations:
                amount = res.get("total_price", 0) or res.get("price", 0)
                commission_rate = settings.get("facility_commission_rate", 10)
                commission_fixed = settings.get("facility_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "date": res.get("created_at"),
                    "type_label": "Tesis",
                    "item_name": res.get("facility_name", "Tesis"),
                    "description": res.get("date", ""),
                    "user_name": res.get("facility_name", ""),
                    "client_name": res.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": res.get("payment_status", "pending"),
                })
            
            # 3. Event payments (as organizer)
            events = await db.events.find({"organizer_id": user_id}).to_list(None)
            event_ids = [str(e.get("_id", e.get("id"))) for e in events]
            
            if event_ids:
                event_payments = await db.event_payments.find({
                    "event_id": {"$in": event_ids},
                    **({"created_at": date_filter} if date_filter else {}),
                    **({"status": "completed"} if status == "paid" else 
                       {"status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
                }).to_list(None)
                
                for payment in event_payments:
                    amount = payment.get("amount", 0)
                    commission_rate = settings.get("event_commission_rate", 10)
                    commission_fixed = settings.get("event_commission_fixed", 0)
                    commission = (amount * commission_rate / 100) + commission_fixed
                    net = amount - commission
                    
                    earnings_list.append({
                        "date": payment.get("created_at"),
                        "type_label": "Etkinlik",
                        "item_name": payment.get("event_name", "Etkinlik"),
                        "description": payment.get("event_name", ""),
                        "user_name": user_name,
                        "client_name": payment.get("user_name", ""),
                        "gross_amount": amount,
                        "commission_rate": commission_rate,
                        "commission_fixed": commission_fixed,
                        "commission_amount": round(commission, 2),
                        "net_amount": round(net, 2),
                        "status": payment.get("status", "pending"),
                    })
            
            # 4. Memberships (as facility owner)
            memberships = await db.memberships.find({
                "facility_owner_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"payment_status": "completed"} if status == "paid" else 
                   {"payment_status": {"$in": ["pending", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for membership in memberships:
                amount = membership.get("price", 0)
                commission_rate = settings.get("membership_commission_rate", 10)
                commission_fixed = settings.get("membership_commission_fixed", 0)
                commission = (amount * commission_rate / 100) + commission_fixed
                net = amount - commission
                
                earnings_list.append({
                    "date": membership.get("created_at"),
                    "type_label": "Üyelik",
                    "item_name": membership.get("facility_name", "Tesis Üyeliği"),
                    "description": membership.get("membership_type", ""),
                    "user_name": membership.get("facility_name", ""),
                    "client_name": membership.get("user_name", ""),
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": commission_fixed,
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": membership.get("payment_status", "pending"),
                })
            
            # 5. Marketplace transactions (as seller)
            marketplace_txs = await db.marketplace_transactions.find({
                "seller_id": user_id,
                **({"created_at": date_filter} if date_filter else {}),
                **({"status": "completed"} if status == "paid" else 
                   {"status": {"$in": ["pending", "pending_payment", "processing"]}} if status == "pending" else {})
            }).to_list(None)
            
            for tx in marketplace_txs:
                amount = tx.get("item_price", 0)
                commission_rate = settings.get("marketplace_commission_rate", 10)
                commission = tx.get("commission_amount", (amount * commission_rate / 100))
                net = tx.get("seller_receives", amount - commission)
                
                # Get listing info
                listing = await db.marketplace_listings.find_one({"id": tx.get("listing_id")})
                listing_title = listing.get("title", "Spor Market") if listing else "Spor Market"
                
                earnings_list.append({
                    "date": tx.get("created_at"),
                    "type_label": "Spor Market",
                    "item_name": listing_title,
                    "description": "",
                    "user_name": user_name,
                    "client_name": "",
                    "gross_amount": amount,
                    "commission_rate": commission_rate,
                    "commission_fixed": tx.get("shipping_fee", 0),
                    "commission_amount": round(commission, 2),
                    "net_amount": round(net, 2),
                    "status": tx.get("status", "pending"),
                })
        
        # Sort by date
        earnings_list.sort(key=lambda x: x.get("date") or datetime.min, reverse=True)
        
        # ======== CREATE EXCEL FILE ========
        wb = Workbook()
        ws = wb.active
        ws.title = "Kazançlar"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers - A1'den başla
        headers = [
            "Tarih",
            "İşlem Tipi", 
            "Ürün/Hizmet Adı",
            "Açıklama",
            "Hizmet Veren",
            "Müşteri",
            "Brüt Tutar (₺)", 
            "Komisyon (%)",
            "Sabit Komisyon (₺)",
            "Komisyon Tutarı (₺)",
            "Net Tutar (₺)",
            "Durum"
        ]
        
        # Write headers starting from A1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows - starting from row 2
        total_gross = 0
        total_commission = 0
        total_net = 0
        
        for row_idx, earning in enumerate(earnings_list, 2):
            # Format date
            date_str = ""
            if earning.get("date"):
                try:
                    date_obj = earning["date"]
                    if isinstance(date_obj, str):
                        date_obj = datetime.fromisoformat(date_obj.replace("Z", "+00:00"))
                    date_str = date_obj.strftime("%d.%m.%Y %H:%M")
                except Exception:
                    date_str = str(earning.get("date", ""))
            
            # Format status
            status_label = {
                "completed": "Ödendi",
                "paid": "Ödendi",
                "pending": "Bekliyor",
                "processing": "İşleniyor"
            }.get(earning.get("status", ""), earning.get("status", ""))
            
            # Row data matching headers
            row_data = [
                date_str,                              # Tarih
                earning.get("type_label", ""),         # İşlem Tipi
                earning.get("item_name", ""),          # Ürün/Hizmet Adı
                earning.get("description", ""),        # Açıklama
                earning.get("user_name", ""),          # Hizmet Veren
                earning.get("client_name", ""),        # Müşteri
                earning.get("gross_amount", 0),        # Brüt Tutar
                earning.get("commission_rate", 0),     # Komisyon %
                earning.get("commission_fixed", 0),    # Sabit Komisyon
                earning.get("commission_amount", 0),   # Komisyon Tutarı
                earning.get("net_amount", 0),          # Net Tutar
                status_label                           # Durum
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Number columns right-aligned
                if col_idx in [7, 8, 9, 10, 11]:
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment
            
            total_gross += earning.get("gross_amount", 0)
            total_commission += earning.get("commission_amount", 0)
            total_net += earning.get("net_amount", 0)
        
        # Summary row
        if earnings_list:
            summary_row = len(earnings_list) + 3
            ws.cell(row=summary_row, column=1, value="TOPLAM").font = Font(bold=True)
            ws.cell(row=summary_row, column=7, value=round(total_gross, 2)).font = Font(bold=True)
            ws.cell(row=summary_row, column=10, value=round(total_commission, 2)).font = Font(bold=True)
            ws.cell(row=summary_row, column=11, value=round(total_net, 2)).font = Font(bold=True)
        
        # Adjust column widths
        column_widths = [18, 15, 30, 25, 20, 20, 15, 12, 18, 18, 15, 12]
        for col_idx, width in enumerate(column_widths, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        period_labels = {
            "weekly": "haftalik",
            "monthly": "aylik",
            "yearly": "yillik",
            "all": "tum"
        }
        status_labels = {
            "paid": "odenen",
            "pending": "bekleyen",
            "all": "tum"
        }
        
        safe_user_name = user_name.replace(' ', '_').replace('/', '_')
        filename = f"kazanclar_{safe_user_name}_{period_labels.get(period, 'tum')}_{status_labels.get(status, 'tum')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        logger.info(f"✅ Excel export created for user {user_id}: {filename} with {len(earnings_list)} records")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting earnings: {e}")
        raise HTTPException(status_code=500, detail=str(e))
