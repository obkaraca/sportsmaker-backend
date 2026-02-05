"""
Workflow Engine - İş Akışı Motoru
Tüm sistem parametrelerini ve otomatik eylemleri yönetir
"""

from fastapi import APIRouter, Depends, HTTPException, Request
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from pydantic import BaseModel
import logging
import uuid

from auth import get_current_user

logger = logging.getLogger(__name__)

workflow_router = APIRouter(tags=["workflow"])

# Global db referansı
db = None

def set_database(database):
    """Database referansını ayarla"""
    global db
    db = database


# ==================== VARSAYILAN AYARLAR ====================

DEFAULT_SETTINGS = {
    # Finansal Parametreler
    "finance": {
        "marketplace_commission": {"value": 10, "unit": "percent", "label": "Marketplace Komisyon Oranı", "description": "Satışlardan alınan platform komisyonu"},
        "event_commission": {"value": 10, "unit": "percent", "label": "Etkinlik Komisyon Oranı", "description": "Etkinlik gelirlerinden alınan komisyon"},
        "reservation_commission": {"value": 10, "unit": "percent", "label": "Rezervasyon Komisyon Oranı", "description": "Rezervasyon gelirlerinden alınan komisyon"},
        "min_withdrawal": {"value": 100, "unit": "TRY", "label": "Minimum Çekim Tutarı", "description": "Kullanıcıların çekebileceği minimum tutar"},
        "payment_period": {"value": "weekly", "unit": "select", "options": ["weekly", "biweekly", "monthly"], "label": "Ödeme Periyodu", "description": "Satıcılara ödeme yapılma sıklığı"},
        "iyzico_commission": {"value": 2.79, "unit": "percent", "label": "Iyzico Komisyon Oranı", "description": "Iyzico ödeme sistemi komisyon oranı"},
        "iyzico_min_fee": {"value": 0.35, "unit": "TRY", "label": "Iyzico Asgari Tutar", "description": "Iyzico işlem başına minimum kesinti tutarı"},
        "vakifbank_commission": {"value": 1.89, "unit": "percent", "label": "Vakıfbank Komisyon Oranı", "description": "Vakıfbank POS komisyon oranı"},
        "finansbank_commission": {"value": 1.99, "unit": "percent", "label": "Finansbank Komisyon Oranı", "description": "QNB Finansbank POS komisyon oranı"},
    },
    
    # Zaman Parametreleri
    "timing": {
        "min_reservation_duration": {"value": 60, "unit": "minutes", "label": "Minimum Rezervasyon Süresi", "description": "Tesis rezervasyonu için minimum süre (dakika)"},
        "reservation_cancel_hours": {"value": 24, "unit": "hours", "label": "Rezervasyon İptal Süresi", "description": "Rezervasyon başlangıcından kaç saat önce iptal edilebilir"},
        "event_cancel_hours": {"value": 48, "unit": "hours", "label": "Etkinlik İptal Süresi", "description": "Etkinlik başlangıcından kaç saat önce iptal edilebilir"},
        "marketplace_auto_approve_days": {"value": 1, "unit": "days", "label": "Otomatik Onay Süresi", "description": "Teslimattan sonra otomatik onay süresi"},
        "payment_reminder_hours": {"value": 2, "unit": "hours", "label": "Ödeme Hatırlatma Süresi", "description": "Bekleyen ödemeler için hatırlatma gönderim süresi"},
        "listing_approval_hours": {"value": 24, "unit": "hours", "label": "İlan Onay Bekleme Süresi", "description": "İlanların onay için bekleyeceği maksimum süre"},
    },
    
    # Bildirim Ayarları
    "notifications": {
        "event_reminder_24h": {"value": True, "unit": "boolean", "label": "24 Saat Önce Hatırlatma", "description": "Etkinlikten 24 saat önce bildirim gönder"},
        "event_reminder_1h": {"value": True, "unit": "boolean", "label": "1 Saat Önce Hatırlatma", "description": "Etkinlikten 1 saat önce bildirim gönder"},
        "reservation_reminder": {"value": True, "unit": "boolean", "label": "Rezervasyon Hatırlatması", "description": "Rezervasyonlardan önce hatırlatma gönder"},
        "payment_notification": {"value": True, "unit": "boolean", "label": "Ödeme Bildirimi", "description": "Ödeme işlemlerinde bildirim gönder"},
        "new_message_notification": {"value": True, "unit": "boolean", "label": "Yeni Mesaj Bildirimi", "description": "Yeni mesajlarda bildirim gönder"},
        "marketing_notifications": {"value": False, "unit": "boolean", "label": "Pazarlama Bildirimleri", "description": "Kampanya ve duyuru bildirimleri"},
    },
    
    # İş Kuralları
    "rules": {
        "require_phone_verification": {"value": True, "unit": "boolean", "label": "Telefon Doğrulaması Zorunlu", "description": "Kullanıcıların telefon doğrulaması yapması zorunlu"},
        "require_iban_for_sellers": {"value": True, "unit": "boolean", "label": "Satıcılar İçin IBAN Zorunlu", "description": "Satış yapabilmek için IBAN gerekli"},
        "auto_approve_listings": {"value": False, "unit": "boolean", "label": "İlanları Otomatik Onayla", "description": "Yeni ilanları otomatik olarak onayla"},
        "allow_guest_checkout": {"value": False, "unit": "boolean", "label": "Misafir Alışverişe İzin Ver", "description": "Kayıt olmadan alışveriş yapılabilir"},
        "max_cancellations_per_month": {"value": 3, "unit": "number", "label": "Aylık Maksimum İptal", "description": "Kullanıcı başına aylık maksimum iptal sayısı"},
        "cancellation_penalty_enabled": {"value": True, "unit": "boolean", "label": "İptal Cezası Aktif", "description": "Geç iptallerde ceza uygula"},
        "cancellation_penalty_percent": {"value": 20, "unit": "percent", "label": "İptal Ceza Oranı", "description": "Geç iptallerde uygulanacak ceza oranı"},
    },
    
    # Takvim Ayarları
    "calendar": {
        "work_start_hour": {"value": 8, "unit": "hour", "label": "Çalışma Başlangıç Saati", "description": "Günlük çalışma başlangıç saati"},
        "work_end_hour": {"value": 22, "unit": "hour", "label": "Çalışma Bitiş Saati", "description": "Günlük çalışma bitiş saati"},
        "reservation_slot_minutes": {"value": 60, "unit": "minutes", "label": "Rezervasyon Slot Süresi", "description": "Rezervasyon zaman dilimi (dakika)"},
        "max_advance_booking_days": {"value": 30, "unit": "days", "label": "Maksimum İleri Rezervasyon", "description": "Kaç gün sonrasına rezervasyon yapılabilir"},
        "min_booking_notice_hours": {"value": 2, "unit": "hours", "label": "Minimum Rezervasyon Bildirimi", "description": "En az kaç saat önce rezervasyon yapılmalı"},
    },
    
    # Raporlama Ayarları
    "reporting": {
        "auto_daily_report": {"value": False, "unit": "boolean", "label": "Günlük Otomatik Rapor", "description": "Her gün otomatik rapor oluştur"},
        "auto_weekly_report": {"value": True, "unit": "boolean", "label": "Haftalık Otomatik Rapor", "description": "Her hafta otomatik rapor oluştur"},
        "auto_monthly_report": {"value": True, "unit": "boolean", "label": "Aylık Otomatik Rapor", "description": "Her ay otomatik rapor oluştur"},
        "report_email_enabled": {"value": True, "unit": "boolean", "label": "Rapor E-posta Gönderimi", "description": "Raporları e-posta ile gönder"},
    },
}


# ==================== EYLEMLER ====================

AVAILABLE_ACTIONS = {
    "send_notification": {
        "id": "send_notification",
        "label": "Bildirim Gönder",
        "description": "Kullanıcılara push bildirim gönderir. Metin değişkenler: {user_name}, {user_phone}, {event_name}, {event_date}, {reservation_date}. Tıklanabilir linkler: [link:user_profile], [link:event], [link:reservation], [link:group_chat], [link:event_cancel], [link:reservation_cancel], [link:message]",
        "icon": "notifications",
        "color": "#F59E0B",
        "parameters": [
            {"name": "title", "label": "Bildirim Başlığı", "type": "text", "required": True, "placeholder": "Hoş Geldiniz {user_name}!"},
            {"name": "message", "label": "Bildirim Mesajı", "type": "textarea", "required": True, "placeholder": "Merhaba {user_name}, rezervasyonunuz onaylandı. Detaylar için: [link:reservation]"},
            {"name": "user_type", "label": "Hedef Kullanıcı Tipi", "type": "select", "required": False, "options": ["all", "player", "coach", "referee", "organizer", "admin", "facility_owner", "venue_owner", "club_manager", "event_creator", "reservation_owner"], "default": "all"},
            {"name": "include_link", "label": "Bağlantı Tipi", "type": "select", "required": False, "options": ["none", "event_link", "reservation_link", "event_cancel_link", "reservation_cancel_link", "message_link", "profile_link", "group_chat_link"], "default": "none"}
        ]
    },
    "send_sms": {
        "id": "send_sms",
        "label": "SMS Gönder",
        "description": "Kullanıcılara SMS gönderir. Mesajda {user_name}, {user_phone}, {event_name}, {event_date}, {reservation_date} kullanabilirsiniz",
        "icon": "chatbubble-ellipses",
        "color": "#10B981",
        "parameters": [
            {"name": "message", "label": "SMS Mesajı", "type": "textarea", "required": True, "placeholder": "Merhaba {user_name}, etkinliğiniz {event_date} tarihinde başlayacak."},
            {"name": "user_type", "label": "Hedef Kullanıcı Tipi", "type": "select", "required": False, "options": ["all", "player", "coach", "referee", "organizer", "admin", "facility_owner", "venue_owner", "club_manager", "event_creator", "reservation_owner"], "default": "all"}
        ]
    },
    "send_email": {
        "id": "send_email",
        "label": "E-posta Gönder",
        "description": "Kullanıcılara e-posta gönderir",
        "icon": "mail",
        "color": "#3B82F6",
        "parameters": [
            {"name": "subject", "label": "E-posta Konusu", "type": "text", "required": True, "placeholder": "Hoş Geldiniz"},
            {"name": "template", "label": "E-posta Şablonu", "type": "select", "required": True, "options": ["welcome", "order_confirmation", "event_reminder", "payment_success"]}
        ]
    },
    "create_log": {
        "id": "create_log",
        "label": "Log Kaydı Oluştur",
        "description": "Sistem loguna kayıt ekler",
        "icon": "document-text",
        "color": "#8B5CF6",
        "parameters": [
            {"name": "action_type", "label": "Log Tipi", "type": "text", "required": True, "placeholder": "WORKFLOW_ACTION"}
        ]
    },
    "auto_approve": {
        "id": "auto_approve",
        "label": "Otomatik Onayla",
        "description": "Bekleyen öğeleri otomatik onaylar",
        "icon": "checkmark-circle",
        "color": "#10B981",
        "parameters": [
            {"name": "item_type", "label": "Öğe Tipi", "type": "select", "required": True, "options": ["listing", "event", "venue"]}
        ]
    },
    "auto_cancel": {
        "id": "auto_cancel",
        "label": "Otomatik İptal",
        "description": "Süresi geçen öğeleri otomatik iptal eder",
        "icon": "close-circle",
        "color": "#EF4444",
        "parameters": [
            {"name": "item_type", "label": "Öğe Tipi", "type": "select", "required": True, "options": ["reservation", "order", "event"]},
            {"name": "hours_after", "label": "Kaç Saat Sonra", "type": "number", "required": True, "default": 24}
        ]
    },
    "apply_penalty": {
        "id": "apply_penalty",
        "label": "Ceza Uygula",
        "description": "Kullanıcıya ceza uygular",
        "icon": "warning",
        "color": "#F97316",
        "parameters": [
            {"name": "penalty_type", "label": "Ceza Tipi", "type": "select", "required": True, "options": ["warning", "temporary_ban", "fee"]},
            {"name": "amount", "label": "Miktar (TL)", "type": "number", "required": False, "default": 0}
        ]
    },
    "generate_report": {
        "id": "generate_report",
        "label": "Rapor Oluştur",
        "description": "Otomatik rapor oluşturur",
        "icon": "stats-chart",
        "color": "#06B6D4",
        "parameters": [
            {"name": "report_type", "label": "Rapor Tipi", "type": "select", "required": True, "options": ["daily", "weekly", "monthly"]},
            {"name": "send_email", "label": "E-posta ile Gönder", "type": "boolean", "required": False, "default": False}
        ]
    },
    "update_status": {
        "id": "update_status",
        "label": "Durum Güncelle",
        "description": "Öğenin durumunu günceller",
        "icon": "refresh",
        "color": "#14B8A6",
        "parameters": [
            {"name": "item_type", "label": "Öğe Tipi", "type": "select", "required": True, "options": ["reservation", "event", "listing", "order"]},
            {"name": "new_status", "label": "Yeni Durum", "type": "text", "required": True, "placeholder": "completed"}
        ]
    },
    "create_calendar_event": {
        "id": "create_calendar_event",
        "label": "Takvim Etkinliği Oluştur",
        "description": "Takvime otomatik etkinlik ekler",
        "icon": "calendar",
        "color": "#EC4899",
        "parameters": [
            {"name": "event_type", "label": "Etkinlik Tipi", "type": "select", "required": True, "options": ["reminder", "meeting", "deadline"]},
            {"name": "reminder_minutes", "label": "Hatırlatma (dk)", "type": "number", "required": False, "default": 30}
        ]
    },
    "create_rating_request": {
        "id": "create_rating_request",
        "label": "Puanlama ve Yorum Talebi Oluştur",
        "description": "Kullanıcıya değerlendirme yapması için bildirim gönderir",
        "icon": "star",
        "color": "#FBBF24",
        "parameters": [
            {"name": "target_user", "label": "Hedef Kullanıcı", "type": "select", "required": True, "options": ["reservation_owner", "facility_owner", "event_creator", "event_participant"]},
            {"name": "rating_type", "label": "Değerlendirme Tipi", "type": "select", "required": True, "options": ["facility", "event", "user", "product"]},
            {"name": "message", "label": "Mesaj", "type": "textarea", "required": False, "placeholder": "Deneyiminizi değerlendirin!", "default": "Lütfen deneyiminizi değerlendirin ve yorum yapın."},
            {"name": "delay_hours", "label": "Gecikme (Saat)", "type": "number", "required": False, "default": 24, "description": "İşlem sonrası kaç saat beklensin"}
        ]
    },
    "open_iyzico_payment": {
        "id": "open_iyzico_payment",
        "label": "Iyzico Ödeme Ekranı Aç",
        "description": "Belirtilen kullanıcıya ödeme linki gönderir",
        "icon": "card",
        "color": "#6366F1",
        "parameters": [
            {"name": "target_user", "label": "Ödeme Yapacak Kullanıcı", "type": "select", "required": True, "options": ["reservation_owner", "event_creator", "event_participant", "specific_user"]},
            {"name": "specific_user_id", "label": "Kullanıcı ID (Specific User için)", "type": "text", "required": False, "placeholder": "user-uuid"},
            {"name": "amount", "label": "Tutar (TL)", "type": "number", "required": True, "placeholder": "100"},
            {"name": "description", "label": "Ödeme Açıklaması", "type": "text", "required": True, "placeholder": "Rezervasyon ödemesi"},
            {"name": "payment_type", "label": "Ödeme Tipi", "type": "select", "required": True, "options": ["reservation", "event_fee", "membership", "penalty", "custom"]}
        ]
    },
    "create_reservation": {
        "id": "create_reservation",
        "label": "Rezervasyon Oluştur",
        "description": "Otomatik olarak yeni bir rezervasyon oluşturur",
        "icon": "calendar-outline",
        "color": "#8B5CF6",
        "parameters": [
            {"name": "target_user", "label": "Rezervasyon Sahibi", "type": "select", "required": True, "options": ["event_creator", "specific_user"]},
            {"name": "specific_user_id", "label": "Kullanıcı ID (Specific User için)", "type": "text", "required": False, "placeholder": "user-uuid"},
            {"name": "facility_id", "label": "Tesis ID", "type": "text", "required": True, "placeholder": "facility-uuid"},
            {"name": "field_id", "label": "Saha ID", "type": "text", "required": True, "placeholder": "field-uuid"},
            {"name": "date", "label": "Tarih", "type": "text", "required": True, "placeholder": "2025-06-20"},
            {"name": "start_time", "label": "Başlangıç Saati", "type": "text", "required": True, "placeholder": "14:00"},
            {"name": "end_time", "label": "Bitiş Saati", "type": "text", "required": True, "placeholder": "15:00"},
            {"name": "auto_confirm", "label": "Otomatik Onayla", "type": "boolean", "required": False, "default": False}
        ]
    },
    "cancel_reservation": {
        "id": "cancel_reservation",
        "label": "Rezervasyonu İptal Et",
        "description": "Mevcut bir rezervasyonu iptal eder",
        "icon": "close-circle",
        "color": "#EF4444",
        "parameters": [
            {"name": "reservation_source", "label": "Rezervasyon Kaynağı", "type": "select", "required": True, "options": ["from_context", "specific_reservation"]},
            {"name": "specific_reservation_id", "label": "Rezervasyon ID (Specific için)", "type": "text", "required": False, "placeholder": "reservation-uuid"},
            {"name": "cancellation_reason", "label": "İptal Nedeni", "type": "textarea", "required": True, "placeholder": "İptal nedeni", "default": "Sistem tarafından iptal edildi"},
            {"name": "notify_user", "label": "Kullanıcıyı Bilgilendir", "type": "boolean", "required": False, "default": True},
            {"name": "refund_type", "label": "İade Tipi", "type": "select", "required": False, "options": ["full", "partial", "none"], "default": "none"}
        ]
    },
    "create_iyzico_refund": {
        "id": "create_iyzico_refund",
        "label": "Iyzico Para İadesi Oluştur",
        "description": "Ödeme için para iadesi başlatır",
        "icon": "return-down-back",
        "color": "#F97316",
        "parameters": [
            {"name": "refund_source", "label": "İade Kaynağı", "type": "select", "required": True, "options": ["from_context", "specific_payment"]},
            {"name": "specific_payment_id", "label": "Ödeme ID (Specific için)", "type": "text", "required": False, "placeholder": "payment-uuid"},
            {"name": "refund_type", "label": "İade Tipi", "type": "select", "required": True, "options": ["full", "partial"]},
            {"name": "refund_amount", "label": "İade Tutarı (Partial için)", "type": "number", "required": False, "placeholder": "50"},
            {"name": "refund_reason", "label": "İade Nedeni", "type": "textarea", "required": True, "placeholder": "İade nedeni", "default": "Müşteri talebi"},
            {"name": "notify_user", "label": "Kullanıcıyı Bilgilendir", "type": "boolean", "required": False, "default": True}
        ]
    },
    "create_excel_report": {
        "id": "create_excel_report",
        "label": "Excel Dosyası Oluştur",
        "description": "Belirtilen veriler için Excel raporu oluşturur",
        "icon": "document",
        "color": "#059669",
        "parameters": [
            {"name": "report_type", "label": "Rapor Tipi", "type": "select", "required": True, "options": ["reservations", "payments", "users", "events", "facilities", "transactions", "custom"]},
            {"name": "date_range", "label": "Tarih Aralığı", "type": "select", "required": False, "options": ["today", "yesterday", "this_week", "last_week", "this_month", "last_month", "custom"], "default": "this_month"},
            {"name": "custom_start_date", "label": "Başlangıç Tarihi (Custom için)", "type": "text", "required": False, "placeholder": "2025-06-01"},
            {"name": "custom_end_date", "label": "Bitiş Tarihi (Custom için)", "type": "text", "required": False, "placeholder": "2025-06-30"},
            {"name": "send_to_email", "label": "E-posta Adresi", "type": "text", "required": False, "placeholder": "admin@example.com"},
            {"name": "target_user", "label": "Raporu Gönderilecek Kişi", "type": "select", "required": False, "options": ["admin", "facility_owner", "specific_user"], "default": "admin"},
            {"name": "include_charts", "label": "Grafik Ekle", "type": "boolean", "required": False, "default": False}
        ]
    },
    "send_push_notification": {
        "id": "send_push_notification",
        "label": "Push Bildirim Oluştur",
        "description": "Kullanıcılara mobil push bildirimi gönderir",
        "icon": "phone-portrait",
        "color": "#7C3AED",
        "parameters": [
            {"name": "title", "label": "Bildirim Başlığı", "type": "text", "required": True, "placeholder": "Yeni Bildirim"},
            {"name": "body", "label": "Bildirim Mesajı", "type": "textarea", "required": True, "placeholder": "Bildirim içeriği"},
            {"name": "target_user", "label": "Hedef Kullanıcı", "type": "select", "required": True, "options": ["all", "reservation_owner", "facility_owner", "event_creator", "event_participants", "admin", "specific_user"]},
            {"name": "specific_user_id", "label": "Kullanıcı ID (Specific User için)", "type": "text", "required": False, "placeholder": "user-uuid"},
            {"name": "action_type", "label": "Tıklandığında", "type": "select", "required": False, "options": ["none", "open_reservation", "open_event", "open_payment", "open_profile", "open_url"], "default": "none"},
            {"name": "action_url", "label": "Özel URL (open_url için)", "type": "text", "required": False, "placeholder": "/custom-page"},
            {"name": "sound", "label": "Ses", "type": "select", "required": False, "options": ["default", "none"], "default": "default"},
            {"name": "badge_count", "label": "Badge Sayısı", "type": "number", "required": False, "default": 1}
        ]
    },
}


# ==================== TETİKLEYİCİLER ====================

AVAILABLE_TRIGGERS = {
    # Kullanıcı İşlemleri
    "on_user_register": {"id": "on_user_register", "label": "Kullanıcı Kayıt Olduğunda", "category": "user", "icon": "person-add", "color": "#10B981"},
    "on_user_login": {"id": "on_user_login", "label": "Kullanıcı Giriş Yaptığında", "category": "user", "icon": "log-in", "color": "#10B981"},
    "on_user_deactivate": {"id": "on_user_deactivate", "label": "Kullanıcı Pasif Edildiğinde", "category": "user", "icon": "person-remove", "color": "#EF4444"},
    
    # Etkinlik İşlemleri
    "on_event_create": {"id": "on_event_create", "label": "Etkinlik Oluşturulduğunda", "category": "event", "icon": "calendar", "color": "#3B82F6"},
    "on_event_join": {"id": "on_event_join", "label": "Etkinliğe Katılım Olduğunda", "category": "event", "icon": "people", "color": "#3B82F6"},
    "on_event_cancel": {"id": "on_event_cancel", "label": "Etkinlik İptal Edildiğinde", "category": "event", "icon": "close-circle", "color": "#EF4444"},
    "on_event_approve": {"id": "on_event_approve", "label": "Etkinlik Onaylandığında", "category": "event", "icon": "checkmark-circle", "color": "#10B981"},
    
    # Rezervasyon İşlemleri
    "on_reservation_create": {"id": "on_reservation_create", "label": "Rezervasyon Oluşturulduğunda", "category": "reservation", "icon": "time", "color": "#8B5CF6"},
    "on_reservation_cancel": {"id": "on_reservation_cancel", "label": "Rezervasyon İptal Edildiğinde", "category": "reservation", "icon": "close-circle", "color": "#EF4444"},
    "on_reservation_complete": {"id": "on_reservation_complete", "label": "Rezervasyon Tamamlandığında", "category": "reservation", "icon": "checkmark-done", "color": "#10B981"},
    
    # Ödeme İşlemleri
    "on_payment_success": {"id": "on_payment_success", "label": "Ödeme Başarılı Olduğunda", "category": "payment", "icon": "checkmark-circle", "color": "#10B981"},
    "on_payment_fail": {"id": "on_payment_fail", "label": "Ödeme Başarısız Olduğunda", "category": "payment", "icon": "close-circle", "color": "#EF4444"},
    "on_refund_complete": {"id": "on_refund_complete", "label": "İade Tamamlandığında", "category": "payment", "icon": "return-down-back", "color": "#F59E0B"},
    
    # Marketplace İşlemleri
    "on_listing_create": {"id": "on_listing_create", "label": "İlan Oluşturulduğunda", "category": "marketplace", "icon": "pricetag", "color": "#EC4899"},
    "on_order_create": {"id": "on_order_create", "label": "Sipariş Oluşturulduğunda", "category": "marketplace", "icon": "cart", "color": "#EC4899"},
    "on_order_ship": {"id": "on_order_ship", "label": "Sipariş Kargoya Verildiğinde", "category": "marketplace", "icon": "car", "color": "#06B6D4"},
    "on_order_deliver": {"id": "on_order_deliver", "label": "Sipariş Teslim Edildiğinde", "category": "marketplace", "icon": "checkmark-done", "color": "#10B981"},
    "on_return_request": {"id": "on_return_request", "label": "İade Talebi Oluşturulduğunda", "category": "marketplace", "icon": "return-down-back", "color": "#F97316"},
    
    # Destek İşlemleri
    "on_support_ticket": {"id": "on_support_ticket", "label": "Destek Talebi Oluşturulduğunda", "category": "support", "icon": "help-circle", "color": "#F59E0B"},
    "on_support_reply": {"id": "on_support_reply", "label": "Destek Talebine Yanıt Geldiğinde", "category": "support", "icon": "chatbubble", "color": "#3B82F6"},
    
    # Zamanlanmış İşlemler
    "scheduled_daily": {"id": "scheduled_daily", "label": "Her Gün (Zamanlanmış)", "category": "scheduled", "icon": "today", "color": "#6B7280"},
    "scheduled_weekly": {"id": "scheduled_weekly", "label": "Her Hafta (Zamanlanmış)", "category": "scheduled", "icon": "calendar", "color": "#6B7280"},
    "scheduled_monthly": {"id": "scheduled_monthly", "label": "Her Ay (Zamanlanmış)", "category": "scheduled", "icon": "calendar-outline", "color": "#6B7280"},
}


# ==================== API ENDPOINTS ====================

@workflow_router.get("/workflow/settings")
async def get_workflow_settings(current_user: dict = Depends(get_current_user)):
    """Tüm iş akışı ayarlarını getir"""
    try:
        current_user_id = current_user["id"]
        
        # Admin kontrolü
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        # Veritabanından kayıtlı ayarları al
        saved_settings = await db.workflow_settings.find_one({"type": "global"})
        
        # Varsayılan ayarlarla birleştir
        result = {}
        for category, params in DEFAULT_SETTINGS.items():
            result[category] = {}
            for key, default_value in params.items():
                # Kayıtlı değer varsa onu kullan, yoksa varsayılanı
                if saved_settings and category in saved_settings and key in saved_settings[category]:
                    result[category][key] = {**default_value, "value": saved_settings[category][key]}
                else:
                    result[category][key] = default_value
        
        return {
            "success": True,
            "settings": result,
            "last_updated": saved_settings.get("updated_at") if saved_settings else None
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get workflow settings error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.put("/workflow/settings")
async def update_workflow_settings(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """İş akışı ayarlarını güncelle"""
    try:
        current_user_id = current_user["id"]
        
        # Admin kontrolü
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        body = await request.json()
        category = body.get("category")
        key = body.get("key")
        value = body.get("value")
        
        if not category or not key:
            raise HTTPException(status_code=400, detail="Kategori ve anahtar gerekli")
        
        # Mevcut ayarları al veya oluştur
        saved_settings = await db.workflow_settings.find_one({"type": "global"})
        
        if not saved_settings:
            saved_settings = {"type": "global", "created_at": datetime.utcnow()}
        
        # Kategoriyi kontrol et
        if category not in saved_settings:
            saved_settings[category] = {}
        
        # Değeri güncelle
        saved_settings[category][key] = value
        saved_settings["updated_at"] = datetime.utcnow()
        saved_settings["updated_by"] = current_user_id
        
        # Kaydet
        await db.workflow_settings.update_one(
            {"type": "global"},
            {"$set": saved_settings},
            upsert=True
        )
        
        # Log aktivite
        try:
            from auth_endpoints import log_user_activity
            await log_user_activity(current_user_id, "workflow_setting_update", "success", {
                "category": category,
                "key": key,
                "new_value": value
            })
        except Exception:
            pass
        
        logger.info(f"⚙️ Workflow setting updated: {category}.{key} = {value}")
        
        return {
            "success": True,
            "message": "Ayar güncellendi",
            "category": category,
            "key": key,
            "value": value
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Update workflow settings error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.get("/workflow/public-setting/{category}/{key}")
async def get_public_workflow_setting(category: str, key: str):
    """Belirli bir workflow ayarını getir (public - auth gerektirmez)"""
    try:
        # Sadece belirli kategorilere izin ver (güvenlik için)
        allowed_public_settings = {
            "timing": ["min_reservation_duration", "reservation_cancel_hours", "event_cancel_hours"],
            "rules": ["require_phone_verification", "max_cancellations_per_month"],
        }
        
        if category not in allowed_public_settings:
            raise HTTPException(status_code=403, detail="Bu kategori public erişime kapalı")
        
        if key not in allowed_public_settings[category]:
            raise HTTPException(status_code=403, detail="Bu ayar public erişime kapalı")
        
        value = await get_setting(category, key)
        
        return {
            "success": True,
            "category": category,
            "key": key,
            "value": value
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.get("/workflow/actions")
async def get_available_actions(current_user: dict = Depends(get_current_user)):
    """Kullanılabilir eylemleri getir"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        return {
            "success": True,
            "actions": list(AVAILABLE_ACTIONS.values()),
            "triggers": list(AVAILABLE_TRIGGERS.values())
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.get("/workflow/rules")
async def get_workflow_rules(current_user: dict = Depends(get_current_user)):
    """Tanımlı iş akışı kurallarını getir"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        rules = await db.workflow_rules.find({"is_deleted": {"$ne": True}}).sort("created_at", -1).to_list(100)
        
        # ObjectId'leri temizle
        for rule in rules:
            rule.pop("_id", None)
        
        return {
            "success": True,
            "rules": rules
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.post("/workflow/rules")
async def create_workflow_rule(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Yeni iş akışı kuralı oluştur"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        body = await request.json()
        
        rule = {
            "id": str(uuid.uuid4()),
            "name": body.get("name"),
            "description": body.get("description", ""),
            "trigger": body.get("trigger"),
            "conditions": body.get("conditions", []),
            "actions": body.get("actions", []),
            "is_active": body.get("is_active", True),
            "priority": body.get("priority", 0),
            "created_by": current_user_id,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        await db.workflow_rules.insert_one(rule)
        rule.pop("_id", None)
        
        logger.info(f"⚙️ New workflow rule created: {rule['name']}")
        
        return {
            "success": True,
            "message": "İş akışı kuralı oluşturuldu",
            "rule": rule
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Create workflow rule error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.put("/workflow/rules/{rule_id}")
async def update_workflow_rule(
    rule_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """İş akışı kuralını güncelle"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        body = await request.json()
        
        update_data = {
            "updated_at": datetime.utcnow(),
            "updated_by": current_user_id
        }
        
        # Güncellenebilir alanlar
        for field in ["name", "description", "trigger", "conditions", "actions", "is_active", "priority"]:
            if field in body:
                update_data[field] = body[field]
        
        result = await db.workflow_rules.update_one(
            {"id": rule_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Kural bulunamadı")
        
        return {
            "success": True,
            "message": "İş akışı kuralı güncellendi"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.delete("/workflow/rules/{rule_id}")
async def delete_workflow_rule(
    rule_id: str,
    current_user: dict = Depends(get_current_user)
):
    """İş akışı kuralını sil"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        result = await db.workflow_rules.update_one(
            {"id": rule_id},
            {"$set": {"is_deleted": True, "deleted_at": datetime.utcnow(), "deleted_by": current_user_id}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Kural bulunamadı")
        
        return {
            "success": True,
            "message": "İş akışı kuralı silindi"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.post("/workflow/rules/{rule_id}/execute")
async def execute_workflow_rule(
    rule_id: str,
    current_user: dict = Depends(get_current_user)
):
    """İş akışı kuralını manuel olarak çalıştır"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        rule = await db.workflow_rules.find_one({"id": rule_id, "is_deleted": {"$ne": True}})
        if not rule:
            raise HTTPException(status_code=404, detail="Kural bulunamadı")
        
        # Kuralı çalıştır
        execution_result = await execute_rule_actions(rule)
        
        # Çalıştırma logunu kaydet
        execution_log = {
            "id": str(uuid.uuid4()),
            "rule_id": rule_id,
            "rule_name": rule.get("name"),
            "executed_by": current_user_id,
            "execution_type": "manual",
            "result": execution_result,
            "executed_at": datetime.utcnow()
        }
        await db.workflow_executions.insert_one(execution_log)
        
        return {
            "success": True,
            "message": "İş akışı kuralı çalıştırıldı",
            "result": execution_result
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Execute workflow rule error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.get("/workflow/executions")
async def get_workflow_executions(
    rule_id: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Workflow çalıştırma loglarını getir"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        query = {}
        if rule_id:
            query["rule_id"] = rule_id
        
        executions = await db.workflow_executions.find(query).sort("executed_at", -1).limit(limit).to_list(limit)
        
        # ObjectId'leri temizle
        for execution in executions:
            execution.pop("_id", None)
        
        return {
            "success": True,
            "executions": executions,
            "count": len(executions)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@workflow_router.get("/workflow/stats")
async def get_workflow_stats(current_user: dict = Depends(get_current_user)):
    """Workflow istatistiklerini getir"""
    try:
        current_user_id = current_user["id"]
        user = await db.users.find_one({"id": current_user_id})
        if not user or user.get("user_type") != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        # Toplam kural sayısı
        total_rules = await db.workflow_rules.count_documents({"is_deleted": {"$ne": True}})
        active_rules = await db.workflow_rules.count_documents({"is_active": True, "is_deleted": {"$ne": True}})
        
        # Son 24 saatteki çalıştırmalar
        yesterday = datetime.utcnow() - timedelta(days=1)
        executions_24h = await db.workflow_executions.count_documents({"executed_at": {"$gte": yesterday}})
        
        # Tetikleyici bazlı dağılım
        trigger_pipeline = [
            {"$match": {"is_deleted": {"$ne": True}}},
            {"$group": {"_id": "$trigger", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        trigger_stats = await db.workflow_rules.aggregate(trigger_pipeline).to_list(100)
        
        return {
            "success": True,
            "stats": {
                "total_rules": total_rules,
                "active_rules": active_rules,
                "executions_24h": executions_24h,
                "trigger_distribution": trigger_stats
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def execute_rule_actions(rule: dict) -> dict:
    """Kural eylemlerini çalıştır"""
    results = []
    
    for action in rule.get("actions", []):
        action_type = action.get("type")
        params = action.get("parameters", {})
        
        try:
            if action_type == "send_notification":
                result = await execute_send_notification(params)
            elif action_type == "auto_approve":
                result = await execute_auto_approve(params)
            elif action_type == "generate_report":
                result = await execute_generate_report(params)
            elif action_type == "update_status":
                result = await execute_update_status(params)
            else:
                result = {"status": "skipped", "reason": f"Unknown action: {action_type}"}
            
            results.append({"action": action_type, "result": result})
        except Exception as e:
            results.append({"action": action_type, "error": str(e)})
    
    return {"actions_executed": len(results), "results": results}


async def execute_send_notification(params: dict) -> dict:
    """Bildirim gönderme eylemi"""
    user_type = params.get("user_type", "all")
    title = params.get("title", "Bildirim")
    message = params.get("message", "")
    
    # Hedef kullanıcıları bul
    query = {}
    if user_type != "all":
        query["user_type"] = user_type
    
    users = await db.users.find(query).to_list(10000)
    sent_count = 0
    
    for user in users:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "type": "workflow_notification",
            "title": title,
            "message": message,
            "read": False,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        sent_count += 1
    
    return {"status": "success", "sent_count": sent_count}


async def execute_auto_approve(params: dict) -> dict:
    """Otomatik onaylama eylemi"""
    item_type = params.get("item_type", "listing")
    
    if item_type == "listing":
        result = await db.marketplace_listings.update_many(
            {"status": "pending"},
            {"$set": {"status": "active", "approved_at": datetime.utcnow()}}
        )
        return {"status": "success", "approved_count": result.modified_count}
    
    return {"status": "skipped", "reason": f"Unknown item type: {item_type}"}


async def execute_generate_report(params: dict) -> dict:
    """Rapor oluşturma eylemi"""
    report_type = params.get("report_type", "daily")
    # Rapor oluşturma mantığı buraya eklenebilir
    return {"status": "success", "report_type": report_type, "generated": True}


async def execute_update_status(params: dict) -> dict:
    """Durum güncelleme eylemi"""
    item_type = params.get("item_type")
    new_status = params.get("new_status")
    condition = params.get("condition", {})
    
    collection_map = {
        "reservation": "reservations",
        "event": "events",
        "listing": "marketplace_listings",
        "order": "marketplace_transactions"
    }
    
    collection_name = collection_map.get(item_type)
    if not collection_name:
        return {"status": "skipped", "reason": f"Unknown item type: {item_type}"}
    
    result = await db[collection_name].update_many(
        condition,
        {"$set": {"status": new_status, "updated_at": datetime.utcnow()}}
    )
    
    return {"status": "success", "updated_count": result.modified_count}


# ==================== AYAR OKUMA FONKSİYONLARI ====================

async def get_setting(category: str, key: str, default=None):
    """Belirli bir ayarı oku"""
    try:
        saved_settings = await db.workflow_settings.find_one({"type": "global"})
        
        if saved_settings and category in saved_settings and key in saved_settings[category]:
            return saved_settings[category][key]
        
        # Varsayılan değeri döndür
        if category in DEFAULT_SETTINGS and key in DEFAULT_SETTINGS[category]:
            return DEFAULT_SETTINGS[category][key]["value"]
        
        return default
    except Exception:
        return default


# ==================== WORKFLOW TETİKLEYİCİ FONKSİYONU ====================

async def trigger_workflow(trigger_id: str, context: dict = None):
    """
    Sistem genelinde workflow tetikleyici
    Bu fonksiyon diğer endpoint'lerden çağrılır
    
    Args:
        trigger_id: Tetikleyici ID (örn: "on_user_register", "on_order_create")
        context: Bağlam bilgisi (örn: user_id, event_id, order_id vb.)
    
    Example:
        await trigger_workflow("on_user_register", {"user_id": "123", "user_name": "Ali"})
    """
    global db
    
    # Eğer db başlatılmamışsa, doğrudan bağlan
    if db is None:
        from motor.motor_asyncio import AsyncIOMotorClient
        import os
        mongo_url = os.getenv("MONGO_URL", "mongodb://localhost:27017")
        db_name = os.getenv("DB_NAME", "sports_management")
        client = AsyncIOMotorClient(mongo_url)
        db = client[db_name]
        logger.info(f"Database initialized for workflow trigger: {db_name}")
    
    try:
        # Bu tetikleyiciye sahip aktif kuralları bul
        rules = await db.workflow_rules.find({
            "trigger": trigger_id,
            "is_active": True,
            "is_deleted": {"$ne": True}
        }).sort("priority", -1).to_list(100)
        
        if not rules:
            logger.debug(f"No active workflow rules for trigger: {trigger_id}")
            return {"rules_found": 0, "executed": 0}
        
        logger.info(f"⚙️ Triggering {len(rules)} workflow rules for: {trigger_id}")
        
        executed_count = 0
        for rule in rules:
            try:
                # Koşulları kontrol et
                if not await check_rule_conditions(rule, context):
                    continue
                
                # Eylemleri çalıştır
                execution_result = await execute_rule_actions_with_context(rule, context)
                
                # Çalıştırma logunu kaydet
                execution_log = {
                    "id": str(uuid.uuid4()),
                    "rule_id": rule.get("id"),
                    "rule_name": rule.get("name"),
                    "trigger": trigger_id,
                    "context": context,
                    "execution_type": "automatic",
                    "result": execution_result,
                    "executed_at": datetime.utcnow()
                }
                await db.workflow_executions.insert_one(execution_log)
                executed_count += 1
                
                logger.info(f"✅ Workflow rule executed: {rule.get('name')}")
                
            except Exception as rule_error:
                logger.error(f"Workflow rule error: {rule.get('name')} - {rule_error}")
        
        return {"rules_found": len(rules), "executed": executed_count}
        
    except Exception as e:
        logger.error(f"Workflow trigger error: {e}")
        return {"error": str(e)}


async def check_rule_conditions(rule: dict, context: dict) -> bool:
    """Kural koşullarını kontrol et"""
    conditions = rule.get("conditions", [])
    
    if not conditions:
        return True  # Koşul yoksa her zaman çalıştır
    
    for condition in conditions:
        condition_type = condition.get("type")
        condition_value = condition.get("value")
        
        if condition_type == "user_type":
            user_id = context.get("user_id")
            if user_id:
                user = await db.users.find_one({"id": user_id})
                if user and user.get("user_type") != condition_value:
                    return False
        
        elif condition_type == "amount_greater_than":
            amount = context.get("amount", 0)
            if amount <= float(condition_value):
                return False
        
        elif condition_type == "time_of_day":
            current_hour = datetime.utcnow().hour
            if condition_value == "morning" and not (6 <= current_hour < 12):
                return False
            elif condition_value == "afternoon" and not (12 <= current_hour < 18):
                return False
            elif condition_value == "evening" and not (18 <= current_hour < 22):
                return False
    
    return True


async def execute_rule_actions_with_context(rule: dict, context: dict) -> dict:
    """Kural eylemlerini bağlam bilgisi ile çalıştır"""
    results = []
    
    for action in rule.get("actions", []):
        action_type = action.get("type")
        params = action.get("parameters", {})
        
        # Bağlam bilgilerini parametrelere ekle
        merged_params = {**params, **context} if context else params
        
        try:
            if action_type == "send_notification":
                result = await execute_send_notification_with_context(merged_params, context)
            elif action_type == "send_sms":
                result = await execute_send_sms_with_context(merged_params, context)
            elif action_type == "send_email":
                result = await execute_send_email(merged_params)
            elif action_type == "create_log":
                result = await execute_create_log(merged_params, context)
            elif action_type == "create_calendar_event":
                result = await execute_create_calendar_event(merged_params, context)
            elif action_type == "auto_approve":
                result = await execute_auto_approve(merged_params)
            elif action_type == "generate_report":
                result = await execute_generate_report(merged_params)
            elif action_type == "update_status":
                result = await execute_update_status(merged_params)
            elif action_type == "create_rating_request":
                result = await execute_create_rating_request(merged_params, context)
            elif action_type == "open_iyzico_payment":
                result = await execute_open_iyzico_payment(merged_params, context)
            elif action_type == "create_reservation":
                result = await execute_create_reservation(merged_params, context)
            elif action_type == "cancel_reservation":
                result = await execute_cancel_reservation(merged_params, context)
            elif action_type == "create_iyzico_refund":
                result = await execute_create_iyzico_refund(merged_params, context)
            elif action_type == "create_excel_report":
                result = await execute_create_excel_report(merged_params, context)
            elif action_type == "send_push_notification":
                result = await execute_send_push_notification(merged_params, context)
            else:
                result = {"status": "skipped", "reason": f"Unknown action: {action_type}"}
            
            results.append({"action": action_type, "result": result})
        except Exception as e:
            results.append({"action": action_type, "error": str(e)})
    
    return {"actions_executed": len(results), "results": results}


async def execute_send_notification_with_context(params: dict, context: dict) -> dict:
    """Bağlam bilgisi ile bildirim gönderme"""
    title = params.get("title", "Bildirim")
    message = params.get("message", "")
    include_link = params.get("include_link", "none")
    
    # Tarih değişkenlerini DD.MM.YYYY formatına çevir
    if context:
        for key in list(context.keys()):
            value = context.get(key)
            # Tarih string'lerini formatla (YYYY-MM-DD -> DD.MM.YYYY)
            if isinstance(value, str) and key in ['event_date', 'reservation_date', 'date', 'start_date', 'end_date']:
                try:
                    # YYYY-MM-DD formatını kontrol et
                    if len(value) == 10 and value[4] == '-' and value[7] == '-':
                        parts = value.split('-')
                        context[key] = f"{parts[2]}.{parts[1]}.{parts[0]}"
                except:
                    pass
            # datetime nesnelerini formatla
            elif isinstance(value, datetime):
                context[key] = value.strftime("%d.%m.%Y")
    
    # ✅ Event bilgilerini context'e ekle (event_name ve group_chat_link)
    if context and context.get("event_id"):
        event_id = context.get("event_id")
        # Eğer event_name yoksa, event'ten al
        if not context.get("event_name"):
            try:
                event = await db.events.find_one({"id": event_id})
                if event:
                    context["event_name"] = event.get("title", "Etkinlik")
            except Exception as e:
                logger.warning(f"Could not fetch event name: {e}")
        
        # Etkinlik grup sohbeti linkini ekle
        try:
            group_chat = await db.group_chats.find_one({"event_id": event_id})
            if group_chat:
                group_chat_url = f"/group-chat/{group_chat.get('id', '')}"
                context["event_group_chat_link"] = group_chat_url
                # Markdown formatında tıklanabilir link için
                context["event_group_chat"] = f"[link:group:Grup Sohbeti:{group_chat_url}]"
            else:
                context["event_group_chat_link"] = ""
                context["event_group_chat"] = ""
        except Exception as e:
            logger.warning(f"Could not fetch group chat: {e}")
            context["event_group_chat_link"] = ""
            context["event_group_chat"] = ""
    
    # Link URL'lerini oluştur
    user_profile_url = f"/profile/{context.get('user_id', '')}" if context else ""
    event_url = f"/event/{context.get('event_id', '')}" if context and context.get('event_id') else ""
    reservation_url = f"/reservation-detail?reservationId={context.get('reservation_id', '')}" if context and context.get('reservation_id') else ""
    event_cancel_url = f"/event/{context.get('event_id', '')}?action=cancel" if context and context.get('event_id') else ""
    reservation_cancel_url = f"/reservation-detail?reservationId={context.get('reservation_id', '')}&action=cancel" if context and context.get('reservation_id') else ""
    message_url = "/messages"
    group_chat_url = f"/group-chat/{context.get('group_chat_id', '')}" if context and context.get('group_chat_id') else (context.get('event_group_chat_link', '') if context else '')
    
    # Tıklanabilir linkleri JSON formatında kaydet (frontend parse edecek)
    # Format: [link:type:label:url]
    user_name = context.get('user_name', 'Kullanıcı') if context else 'Kullanıcı'
    event_name = context.get('event_name', 'Etkinlik') if context else 'Etkinlik'
    
    link_replacements = {
        '[link:user_profile]': f'[link:profile:{user_name}:{user_profile_url}]',
        '[link:event]': f'[link:event:{event_name}:{event_url}]',
        '[link:reservation]': f'[link:reservation:Rezervasyon:{reservation_url}]',
        '[link:event_cancel]': f'[link:cancel:Etkinlik İptal:{event_cancel_url}]',
        '[link:reservation_cancel]': f'[link:cancel:Rezervasyon İptal:{reservation_cancel_url}]',
        '[link:message]': f'[link:message:Mesaj:{message_url}]',
        '[link:group_chat]': f'[link:group:Grup Sohbeti:{group_chat_url}]',
    }
    
    # Linkleri değiştir
    for placeholder, replacement in link_replacements.items():
        message = message.replace(placeholder, replacement)
        title = title.replace(placeholder, replacement)
    
    # Context'e link değişkenlerini ekle (eski format uyumluluğu için)
    if context:
        if context.get("reservation_id"):
            context["reservation_cancel_link"] = reservation_cancel_url
            context["reservation_link"] = reservation_url
        if context.get("event_id"):
            context["event_cancel_link"] = event_cancel_url
            context["event_link"] = event_url
            context["group_chat_link"] = group_chat_url
    
    # Normal placeholder'ları değiştir
    if context:
        for key, value in context.items():
            if value is not None and isinstance(value, (str, int, float)):
                message = message.replace(f"{{{key}}}", str(value))
                title = title.replace(f"{{{key}}}", str(value))
    
    # Link URL'sini belirle (action_url için)
    action_url = None
    if include_link and include_link != "none" and context:
        if include_link == "reservation_link":
            action_url = reservation_url
        elif include_link == "event_link":
            action_url = event_url
        elif include_link == "reservation_cancel_link":
            action_url = reservation_cancel_url
        elif include_link == "event_cancel_link":
            action_url = event_cancel_url
        elif include_link == "message_link":
            action_url = message_url
        elif include_link == "profile_link":
            action_url = user_profile_url
        elif include_link == "group_chat_link":
            action_url = group_chat_url
    
    sent_count = 0
    target_user_ids = []
    
    # Hedef kullanıcı tiplerini al
    user_type_filter = params.get("user_type", ["all"])
    if isinstance(user_type_filter, str):
        user_type_filter = [user_type_filter]
    
    logger.info(f"📬 Sending notification - Title: {title}, Target types: {user_type_filter}, Link: {action_url}")
    
    # Özel hedefleri kontrol et
    if "reservation_owner" in user_type_filter and context.get("user_id"):
        target_user_ids.append(context["user_id"])
        logger.info(f"  → Adding reservation owner: {context['user_id']}")
    
    if "facility_owner" in user_type_filter and context.get("owner_id"):
        target_user_ids.append(context["owner_id"])
        logger.info(f"  → Adding facility owner: {context['owner_id']}")
    
    if "event_creator" in user_type_filter and context.get("creator_id"):
        target_user_ids.append(context["creator_id"])
        logger.info(f"  → Adding event creator: {context['creator_id']}")
    
    # Admin kullanıcıları ekle
    if "admin" in user_type_filter:
        admins = await db.users.find({"user_type": "admin"}).to_list(100)
        for admin in admins:
            if admin["id"] not in target_user_ids:
                target_user_ids.append(admin["id"])
                logger.info(f"  → Adding admin: {admin['id']}")
    
    # Diğer kullanıcı tipleri
    other_types = [t for t in user_type_filter if t not in ["all", "reservation_owner", "facility_owner", "event_creator", "admin"]]
    if other_types:
        users = await db.users.find({"user_type": {"$in": other_types}}).to_list(10000)
        for user in users:
            if user["id"] not in target_user_ids:
                target_user_ids.append(user["id"])
    
    # "all" durumu
    if "all" in user_type_filter:
        users = await db.users.find({}).to_list(10000)
        for user in users:
            if user["id"] not in target_user_ids:
                target_user_ids.append(user["id"])
    
    # Bildirimleri gönder
    for user_id in target_user_ids:
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "type": "workflow_notification",
            "title": title,
            "message": message,
            "read": False,
            "is_read": False,
            "action_url": action_url,
            "reservation_id": context.get("reservation_id") if context else None,
            "event_id": context.get("event_id") if context else None,
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        sent_count += 1
        
        # Push notification da gönder
        try:
            user = await db.users.find_one({"id": user_id})
            if user and user.get("push_token"):
                import httpx
                async with httpx.AsyncClient() as client:
                    await client.post(
                        "https://exp.host/--/api/v2/push/send",
                        json={
                            "to": user["push_token"],
                            "title": title,
                            "body": message[:200],
                            "data": {
                                "type": "workflow",
                                "action_url": action_url,
                                "reservation_id": context.get("reservation_id") if context else None
                            }
                        }
                    )
        except Exception as push_err:
            logger.warning(f"Push notification error: {push_err}")
    
    logger.info(f"📬 Notification sent to {sent_count} users")
    return {"status": "success", "sent_count": sent_count}


async def execute_send_email(params: dict) -> dict:
    """E-posta gönderme eylemi (placeholder - gerçek entegrasyon eklenebilir)"""
    # E-posta entegrasyonu için placeholder
    # Gerçek uygulamada SendGrid, SES vb. kullanılabilir
    logger.info(f"📧 Email action triggered: {params}")
    return {"status": "success", "message": "Email action logged (integration pending)"}


async def execute_create_log(params: dict, context: dict) -> dict:
    """Sistem logu oluşturma eylemi"""
    try:
        from auth_endpoints import log_user_activity
        
        user_id = context.get("user_id", "system")
        action_type = params.get("action_type", "WORKFLOW_LOG")
        
        await log_user_activity(user_id, action_type, "success", {
            "workflow_params": params,
            "context": context
        })
        
        return {"status": "success", "logged": True}
    except Exception as e:
        return {"status": "error", "message": str(e)}


async def execute_send_sms_with_context(params: dict, context: dict) -> dict:
    """Bağlam bilgisi ile SMS gönderme (placeholder)"""
    message = params.get("message", "")
    
    # Mesajda placeholder'ları değiştir
    if context:
        for key, value in context.items():
            if isinstance(value, str):
                message = message.replace(f"{{{key}}}", value)
    
    # SMS entegrasyonu için placeholder
    logger.info(f"📱 SMS action triggered: {message}")
    return {"status": "success", "message": "SMS action logged (integration pending)"}


async def execute_create_calendar_event(params: dict, context: dict) -> dict:
    """Takvim etkinliği oluşturma eylemi"""
    try:
        user_id = context.get("user_id")
        if not user_id:
            return {"status": "error", "message": "user_id required"}
        
        event_type = params.get("event_type", "reminder")
        reminder_minutes = params.get("reminder_minutes", 30)
        
        # Rezervasyon bilgilerini al
        reservation_date = context.get("reservation_date")
        start_time = context.get("start_time")
        end_time = context.get("end_time")
        facility_name = context.get("facility_name", "")
        field_name = context.get("field_name", "")
        
        # Takvim etkinliği oluştur
        calendar_event = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "type": event_type,
            "title": f"Rezervasyon: {facility_name} - {field_name}",
            "description": f"Tarih: {reservation_date}\nSaat: {start_time} - {end_time}",
            "date": reservation_date,
            "start_time": start_time,
            "end_time": end_time,
            "reminder_minutes": reminder_minutes,
            "is_read": False,
            "created_at": datetime.utcnow(),
            "source": "workflow"
        }
        
        await db.calendar_items.insert_one(calendar_event)
        logger.info(f"📅 Calendar event created: {calendar_event['id']}")
        
        return {"status": "success", "calendar_event_id": calendar_event["id"]}
    except Exception as e:
        logger.error(f"Calendar event error: {e}")
        return {"status": "error", "message": str(e)}


# ==================== YENİ EYLEM FONKSİYONLARI ====================

async def execute_create_rating_request(params: dict, context: dict) -> dict:
    """Puanlama ve yorum talebi oluştur"""
    try:
        target_user = params.get("target_user", "reservation_owner")
        rating_type = params.get("rating_type", "facility")
        message = params.get("message", "Lütfen deneyiminizi değerlendirin ve yorum yapın.")
        delay_hours = params.get("delay_hours", 24)
        
        # Hedef kullanıcı ID'sini belirle
        target_user_id = None
        if target_user == "reservation_owner":
            target_user_id = context.get("user_id")
        elif target_user == "facility_owner":
            target_user_id = context.get("owner_id")
        elif target_user == "event_creator":
            target_user_id = context.get("creator_id")
        elif target_user == "event_participant":
            target_user_id = context.get("participant_id")
        
        if not target_user_id:
            return {"status": "error", "message": f"Target user not found: {target_user}"}
        
        # Değerlendirme talebi oluştur
        rating_request = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "rating_type": rating_type,
            "reference_id": context.get("reservation_id") or context.get("event_id") or context.get("facility_id"),
            "reference_type": rating_type,
            "message": message,
            "status": "pending",
            "scheduled_at": datetime.utcnow() + timedelta(hours=delay_hours),
            "created_at": datetime.utcnow(),
            "context": context
        }
        
        await db.rating_requests.insert_one(rating_request)
        
        # Bildirim de gönder
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "type": "rating_request",
            "title": "⭐ Değerlendirme Talebi",
            "message": message,
            "read": False,
            "is_read": False,
            "action_url": f"/rate/{rating_type}/{rating_request['reference_id']}",
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"⭐ Rating request created: {rating_request['id']} for user {target_user_id}")
        return {"status": "success", "rating_request_id": rating_request["id"], "target_user_id": target_user_id}
    except Exception as e:
        logger.error(f"Create rating request error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_open_iyzico_payment(params: dict, context: dict) -> dict:
    """Iyzico ödeme ekranı aç - Ödeme linki gönderir"""
    try:
        target_user = params.get("target_user", "reservation_owner")
        specific_user_id = params.get("specific_user_id")
        amount = params.get("amount", 0)
        description = params.get("description", "Ödeme")
        payment_type = params.get("payment_type", "custom")
        
        # Hedef kullanıcı ID'sini belirle
        target_user_id = None
        if target_user == "specific_user" and specific_user_id:
            target_user_id = specific_user_id
        elif target_user == "reservation_owner":
            target_user_id = context.get("user_id")
        elif target_user == "event_creator":
            target_user_id = context.get("creator_id")
        elif target_user == "event_participant":
            target_user_id = context.get("participant_id")
        
        if not target_user_id:
            return {"status": "error", "message": f"Target user not found: {target_user}"}
        
        if amount <= 0:
            return {"status": "error", "message": "Amount must be greater than 0"}
        
        # Ödeme talebi oluştur
        payment_request = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "amount": float(amount),
            "description": description,
            "payment_type": payment_type,
            "status": "pending",
            "reference_id": context.get("reservation_id") or context.get("event_id"),
            "created_at": datetime.utcnow(),
            "context": context
        }
        
        await db.payment_requests.insert_one(payment_request)
        
        # Kullanıcıya bildirim gönder
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "type": "payment_request",
            "title": "💳 Ödeme Talebi",
            "message": f"{description} için {amount} TL ödeme yapmanız bekleniyor.",
            "read": False,
            "is_read": False,
            "action_url": f"/payment/{payment_request['id']}",
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"💳 Payment request created: {payment_request['id']} - {amount} TL for user {target_user_id}")
        return {"status": "success", "payment_request_id": payment_request["id"], "amount": amount, "target_user_id": target_user_id}
    except Exception as e:
        logger.error(f"Open iyzico payment error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_create_reservation(params: dict, context: dict) -> dict:
    """Otomatik rezervasyon oluştur"""
    try:
        target_user = params.get("target_user", "event_creator")
        specific_user_id = params.get("specific_user_id")
        facility_id = params.get("facility_id")
        field_id = params.get("field_id")
        date = params.get("date")
        start_time = params.get("start_time")
        end_time = params.get("end_time")
        auto_confirm = params.get("auto_confirm", False)
        
        # Hedef kullanıcı ID'sini belirle
        target_user_id = None
        if target_user == "specific_user" and specific_user_id:
            target_user_id = specific_user_id
        elif target_user == "event_creator":
            target_user_id = context.get("creator_id") or context.get("user_id")
        
        if not target_user_id:
            return {"status": "error", "message": f"Target user not found: {target_user}"}
        
        if not all([facility_id, field_id, date, start_time, end_time]):
            return {"status": "error", "message": "Missing required fields: facility_id, field_id, date, start_time, end_time"}
        
        # Tesis bilgisini al
        facility = await db.facilities.find_one({"id": facility_id})
        if not facility:
            return {"status": "error", "message": f"Facility not found: {facility_id}"}
        
        # Saha bilgisini al
        field = None
        for f in facility.get("fields", []):
            if f.get("id") == field_id:
                field = f
                break
        
        if not field:
            return {"status": "error", "message": f"Field not found: {field_id}"}
        
        # Rezervasyon oluştur
        reservation = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "facility_id": facility_id,
            "field_id": field_id,
            "facility_name": facility.get("name", ""),
            "field_name": field.get("name", ""),
            "date": date,
            "start_time": start_time,
            "end_time": end_time,
            "price": field.get("price", 0),
            "status": "confirmed" if auto_confirm else "pending",
            "payment_status": "pending",
            "source": "workflow",
            "created_at": datetime.utcnow(),
            "context": context
        }
        
        await db.reservations.insert_one(reservation)
        
        # Bildirim gönder
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": target_user_id,
            "type": "reservation_created",
            "title": "📅 Rezervasyon Oluşturuldu",
            "message": f"{facility.get('name')} - {field.get('name')} için {date} tarihinde {start_time}-{end_time} saatleri arasında rezervasyon oluşturuldu.",
            "read": False,
            "is_read": False,
            "action_url": f"/reservation-detail/{reservation['id']}",
            "created_at": datetime.utcnow()
        }
        await db.notifications.insert_one(notification)
        
        logger.info(f"📅 Reservation created via workflow: {reservation['id']}")
        return {"status": "success", "reservation_id": reservation["id"], "reservation_status": reservation["status"]}
    except Exception as e:
        logger.error(f"Create reservation error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_cancel_reservation(params: dict, context: dict) -> dict:
    """Rezervasyonu iptal et"""
    try:
        reservation_source = params.get("reservation_source", "from_context")
        specific_reservation_id = params.get("specific_reservation_id")
        cancellation_reason = params.get("cancellation_reason", "Sistem tarafından iptal edildi")
        notify_user = params.get("notify_user", True)
        refund_type = params.get("refund_type", "none")
        
        # Rezervasyon ID'sini belirle
        reservation_id = None
        if reservation_source == "specific_reservation" and specific_reservation_id:
            reservation_id = specific_reservation_id
        else:
            reservation_id = context.get("reservation_id")
        
        if not reservation_id:
            return {"status": "error", "message": "Reservation ID not found"}
        
        # Rezervasyonu bul
        reservation = await db.reservations.find_one({"id": reservation_id})
        if not reservation:
            return {"status": "error", "message": f"Reservation not found: {reservation_id}"}
        
        # Rezervasyonu iptal et
        await db.reservations.update_one(
            {"id": reservation_id},
            {
                "$set": {
                    "status": "cancelled",
                    "cancellation_reason": cancellation_reason,
                    "cancelled_at": datetime.utcnow(),
                    "cancelled_by": "workflow"
                }
            }
        )
        
        # Kullanıcıya bildirim gönder
        if notify_user and reservation.get("user_id"):
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": reservation["user_id"],
                "type": "reservation_cancelled",
                "title": "❌ Rezervasyon İptal Edildi",
                "message": f"Rezervasyonunuz iptal edildi. Neden: {cancellation_reason}",
                "read": False,
                "is_read": False,
                "action_url": f"/reservation-detail/{reservation_id}",
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        # İade işlemi başlat
        refund_result = None
        if refund_type != "none" and reservation.get("payment_id"):
            refund_result = await execute_create_iyzico_refund({
                "refund_source": "from_context",
                "refund_type": refund_type,
                "refund_reason": cancellation_reason,
                "notify_user": notify_user
            }, {"payment_id": reservation.get("payment_id"), "user_id": reservation.get("user_id")})
        
        logger.info(f"❌ Reservation cancelled via workflow: {reservation_id}")
        return {"status": "success", "reservation_id": reservation_id, "refund_result": refund_result}
    except Exception as e:
        logger.error(f"Cancel reservation error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_create_iyzico_refund(params: dict, context: dict) -> dict:
    """Iyzico para iadesi oluştur - Gerçek Iyzico servisini kullanır"""
    try:
        from iyzico_service import iyzico_service
        
        refund_source = params.get("refund_source", "from_context")
        specific_payment_id = params.get("specific_payment_id")
        refund_type = params.get("refund_type", "full")
        refund_amount = params.get("refund_amount", 0)
        refund_reason = params.get("refund_reason", "Müşteri talebi")
        notify_user = params.get("notify_user", True)
        
        # Ödeme ID'sini belirle
        payment_id = None
        if refund_source == "specific_payment" and specific_payment_id:
            payment_id = specific_payment_id
        else:
            payment_id = context.get("payment_id")
        
        if not payment_id:
            return {"status": "error", "message": "Payment ID not found"}
        
        # Ödemeyi bul
        payment = await db.payments.find_one({"id": payment_id})
        if not payment:
            # Rezervasyon ödemesini kontrol et
            payment = await db.reservations.find_one({"payment_id": payment_id})
        if not payment:
            # iyzico_payment_id ile de dene
            payment = await db.payments.find_one({"iyzico_payment_id": payment_id})
        
        if not payment:
            return {"status": "error", "message": f"Payment not found: {payment_id}"}
        
        # İade tutarını belirle
        original_amount = payment.get("total_price") or payment.get("amount", 0)
        actual_refund_amount = original_amount if refund_type == "full" else float(refund_amount)
        
        if actual_refund_amount <= 0:
            return {"status": "error", "message": "Refund amount must be greater than 0"}
        
        # Iyzico payment ID'sini al
        iyzico_payment_id = payment.get("iyzico_payment_id") or payment.get("payment_id")
        
        # İade kaydı oluştur
        refund_record = {
            "id": str(uuid.uuid4()),
            "payment_id": payment_id,
            "iyzico_payment_id": iyzico_payment_id,
            "user_id": context.get("user_id") or payment.get("user_id"),
            "original_amount": original_amount,
            "refund_amount": actual_refund_amount,
            "refund_type": refund_type,
            "reason": refund_reason,
            "status": "pending",
            "created_at": datetime.utcnow(),
            "context": context
        }
        
        # Iyzico API ile gerçek iade işlemi yap
        iyzico_result = None
        if iyzico_payment_id:
            try:
                iyzico_result = iyzico_service.refund_payment(
                    payment_id=iyzico_payment_id,
                    amount=actual_refund_amount
                )
                
                if iyzico_result.get("status") == "success":
                    refund_record["status"] = "completed"
                    refund_record["iyzico_refund_id"] = iyzico_result.get("paymentTransactionId")
                    logger.info(f"💰 Iyzico refund successful: {iyzico_result}")
                else:
                    refund_record["status"] = "failed"
                    refund_record["error_message"] = iyzico_result.get("errorMessage", "Unknown error")
                    logger.error(f"💰 Iyzico refund failed: {iyzico_result.get('errorMessage')}")
            except Exception as iyzico_err:
                logger.warning(f"Iyzico refund API error (will be retried manually): {iyzico_err}")
                refund_record["status"] = "pending_manual"
                refund_record["error_message"] = str(iyzico_err)
        else:
            refund_record["status"] = "pending_manual"
            refund_record["note"] = "No Iyzico payment ID found, manual refund required"
        
        await db.refunds.insert_one(refund_record)
        
        # Ödeme kaydını güncelle
        if refund_record["status"] == "completed":
            await db.payments.update_one(
                {"id": payment_id},
                {"$set": {
                    "refund_status": "refunded",
                    "refund_amount": actual_refund_amount,
                    "refund_date": datetime.utcnow()
                }}
            )
        
        # Kullanıcıya bildirim gönder
        if notify_user and refund_record.get("user_id"):
            status_text = "başlatıldı" if refund_record["status"] == "pending" else (
                "tamamlandı" if refund_record["status"] == "completed" else "işleme alındı"
            )
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": refund_record["user_id"],
                "type": "refund_initiated",
                "title": "💰 Para İadesi",
                "message": f"{actual_refund_amount} TL tutarında para iadesi {status_text}. Neden: {refund_reason}",
                "read": False,
                "is_read": False,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"💰 Refund initiated via workflow: {refund_record['id']} - {actual_refund_amount} TL - Status: {refund_record['status']}")
        return {
            "status": "success", 
            "refund_id": refund_record["id"], 
            "refund_amount": actual_refund_amount,
            "refund_status": refund_record["status"],
            "iyzico_result": iyzico_result
        }
    except Exception as e:
        logger.error(f"Create iyzico refund error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_create_excel_report(params: dict, context: dict) -> dict:
    """Excel raporu oluştur"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        import base64
        
        report_type = params.get("report_type", "reservations")
        date_range = params.get("date_range", "this_month")
        custom_start_date = params.get("custom_start_date")
        custom_end_date = params.get("custom_end_date")
        send_to_email = params.get("send_to_email")
        target_user = params.get("target_user", "admin")
        include_charts = params.get("include_charts", False)
        
        # Tarih aralığını hesapla
        now = datetime.utcnow()
        start_date = None
        end_date = now
        
        if date_range == "today":
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_range == "yesterday":
            start_date = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_range == "this_week":
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif date_range == "last_week":
            start_date = now - timedelta(days=now.weekday() + 7)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now - timedelta(days=now.weekday())
        elif date_range == "this_month":
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif date_range == "last_month":
            first_day_this_month = now.replace(day=1)
            end_date = first_day_this_month - timedelta(days=1)
            start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif date_range == "custom" and custom_start_date and custom_end_date:
            start_date = datetime.strptime(custom_start_date, "%Y-%m-%d")
            end_date = datetime.strptime(custom_end_date, "%Y-%m-%d")
        
        # Excel oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Raporu"
        
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Koleksiyon ve alanları belirle
        collection_map = {
            "reservations": {"collection": "reservations", "fields": ["id", "user_id", "facility_name", "field_name", "date", "start_time", "end_time", "status", "price", "created_at"]},
            "payments": {"collection": "payments", "fields": ["id", "user_id", "amount", "status", "payment_type", "created_at"]},
            "users": {"collection": "users", "fields": ["id", "email", "first_name", "last_name", "user_type", "created_at"]},
            "events": {"collection": "events", "fields": ["id", "title", "sport_type", "date", "time", "location", "status", "created_at"]},
            "facilities": {"collection": "facilities", "fields": ["id", "name", "city", "district", "status", "created_at"]},
            "transactions": {"collection": "transactions", "fields": ["id", "user_id", "type", "amount", "status", "created_at"]},
        }
        
        config = collection_map.get(report_type, collection_map["reservations"])
        collection = db[config["collection"]]
        fields = config["fields"]
        
        # Başlıkları yaz
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Verileri çek
        query = {}
        if start_date:
            query["created_at"] = {"$gte": start_date, "$lte": end_date}
        
        records = await collection.find(query).sort("created_at", -1).to_list(10000)
        
        # Verileri yaz
        for row, record in enumerate(records, 2):
            for col, field in enumerate(fields, 1):
                value = record.get(field, "")
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                ws.cell(row=row, column=col, value=str(value) if value else "")
        
        # Dosyayı kaydet
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        file_data = base64.b64encode(file_buffer.getvalue()).decode()
        
        # Rapor kaydı oluştur
        report_record = {
            "id": str(uuid.uuid4()),
            "report_type": report_type,
            "date_range": date_range,
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "record_count": len(records),
            "file_name": f"{report_type}_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx",
            "file_data": file_data,
            "created_at": datetime.utcnow(),
            "created_by": "workflow"
        }
        
        await db.reports.insert_one(report_record)
        
        # Hedef kullanıcıya bildirim gönder
        target_user_id = None
        if target_user == "admin":
            admin = await db.users.find_one({"user_type": "admin"})
            target_user_id = admin["id"] if admin else None
        elif target_user == "facility_owner":
            target_user_id = context.get("owner_id")
        elif target_user == "specific_user":
            target_user_id = params.get("specific_user_id")
        
        if target_user_id:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": target_user_id,
                "type": "report_ready",
                "title": "📊 Rapor Hazır",
                "message": f"{report_type.capitalize()} raporu oluşturuldu. {len(records)} kayıt içeriyor.",
                "read": False,
                "is_read": False,
                "action_url": f"/admin/reports/{report_record['id']}",
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"📊 Excel report created via workflow: {report_record['id']} - {len(records)} records")
        return {"status": "success", "report_id": report_record["id"], "record_count": len(records), "file_name": report_record["file_name"]}
    except Exception as e:
        logger.error(f"Create excel report error: {e}")
        return {"status": "error", "message": str(e)}


async def execute_send_push_notification(params: dict, context: dict) -> dict:
    """Push bildirim gönder"""
    try:
        import httpx
        
        title = params.get("title", "Yeni Bildirim")
        body = params.get("body", "")
        target_user = params.get("target_user", "all")
        specific_user_id = params.get("specific_user_id")
        action_type = params.get("action_type", "none")
        action_url = params.get("action_url")
        sound = params.get("sound", "default")
        badge_count = params.get("badge_count", 1)
        
        # Mesajda placeholder'ları değiştir
        if context:
            for key, value in context.items():
                if value is not None and isinstance(value, (str, int, float)):
                    body = body.replace(f"{{{key}}}", str(value))
                    title = title.replace(f"{{{key}}}", str(value))
        
        # Hedef kullanıcıları belirle
        target_user_ids = []
        
        if target_user == "specific_user" and specific_user_id:
            target_user_ids.append(specific_user_id)
        elif target_user == "reservation_owner":
            if context.get("user_id"):
                target_user_ids.append(context["user_id"])
        elif target_user == "facility_owner":
            if context.get("owner_id"):
                target_user_ids.append(context["owner_id"])
        elif target_user == "event_creator":
            if context.get("creator_id"):
                target_user_ids.append(context["creator_id"])
        elif target_user == "event_participants":
            event_id = context.get("event_id")
            if event_id:
                event = await db.events.find_one({"id": event_id})
                if event:
                    for participant in event.get("participants", []):
                        if participant.get("user_id"):
                            target_user_ids.append(participant["user_id"])
        elif target_user == "admin":
            admins = await db.users.find({"user_type": "admin"}).to_list(100)
            target_user_ids = [admin["id"] for admin in admins]
        elif target_user == "all":
            users = await db.users.find({"push_token": {"$exists": True, "$ne": None}}).to_list(10000)
            target_user_ids = [user["id"] for user in users]
        
        # Action URL'yi belirle
        notification_action_url = action_url
        if action_type != "none" and not notification_action_url:
            if action_type == "open_reservation":
                notification_action_url = f"/reservation-detail/{context.get('reservation_id', '')}"
            elif action_type == "open_event":
                notification_action_url = f"/events/{context.get('event_id', '')}"
            elif action_type == "open_payment":
                notification_action_url = f"/payment/{context.get('payment_id', '')}"
            elif action_type == "open_profile":
                notification_action_url = "/profile"
        
        sent_count = 0
        failed_count = 0
        
        async with httpx.AsyncClient() as client:
            for user_id in target_user_ids:
                try:
                    user = await db.users.find_one({"id": user_id})
                    if not user or not user.get("push_token"):
                        continue
                    
                    push_data = {
                        "to": user["push_token"],
                        "title": title,
                        "body": body[:200],
                        "sound": sound,
                        "badge": badge_count,
                        "data": {
                            "type": "workflow_push",
                            "action_url": notification_action_url,
                            "reservation_id": context.get("reservation_id"),
                            "event_id": context.get("event_id")
                        }
                    }
                    
                    response = await client.post(
                        "https://exp.host/--/api/v2/push/send",
                        json=push_data,
                        timeout=10.0
                    )
                    
                    if response.status_code == 200:
                        sent_count += 1
                    else:
                        failed_count += 1
                        
                except Exception as push_err:
                    logger.warning(f"Push notification error for user {user_id}: {push_err}")
                    failed_count += 1
        
        # Ayrıca veritabanına bildirim kaydet
        for user_id in target_user_ids:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": user_id,
                "type": "push_notification",
                "title": title,
                "message": body,
                "read": False,
                "is_read": False,
                "action_url": notification_action_url,
                "created_at": datetime.utcnow()
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"📱 Push notifications sent via workflow: {sent_count} success, {failed_count} failed")
        return {"status": "success", "sent_count": sent_count, "failed_count": failed_count, "total_targets": len(target_user_ids)}
    except Exception as e:
        logger.error(f"Send push notification error: {e}")
        return {"status": "error", "message": str(e)}
